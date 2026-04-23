from operator import itemgetter
from django.shortcuts import render, redirect
from django.http.response import JsonResponse, FileResponse
from django.contrib.auth.decorators import login_required
from django.db.models.query_utils import Q
from django.db.models import Sum, F, Window
from django.db.models.functions import Concat, Rank
from django.conf import settings
from django.core import exceptions
from reports import models as report_models
from Atlantic.utils import JsonRender, rol_permission
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font
from openpyxl.formatting.rule import Rule, IconSetRule, ColorScaleRule, CellIsRule
import json


# Create your views here.
@login_required
@rol_permission('Control gestion')
def blank_redirect(request):
    
    return redirect('/welcome')

@login_required
@rol_permission('Control gestion')
def landing(request):
    areas = report_models.areas_empresa.objects.all()
    context = {
        'areas': areas
    }
    
    
    if request_is_ajax(request):
        if request.method == 'GET':
            todo = request.GET.get('todo')
            if todo == 'generate_report':
                area = request.GET.get('area')
                month = request.GET.get('month')
                year = request.GET.get('year')
                include_acumulate = request.GET.get('include_acumulate')
                comparate_year = request.GET.get('comparate_year')
                include_details = request.GET.get('include_details')
                
                obj_classification = report_models.clasificacion.objects.all().order_by('pk')
                                
                if area != "TODAS":
                    obj_classification = obj_classification.filter(responsable = area)
                
                i = 2
                report = []
                
                for classify in obj_classification:
                    agg = classify.analysis_by_period(month, int(year), include_acumulate, comparate_year)
                    report.append([
                        classify.pk, classify.tipo, classify.tipo_r_p,classify.cuenta.cuenta,
                        str(classify.cuenta), classify.centro_costo, classify.nit_tercero,
                        classify.nombre_tercero, classify.responsable,
                        agg.get('gastos'),agg.get('rolling'),
                        f"=J{i}-K{i}",
                        f"=J{i}/K{i}",
                        agg.get('gastos_acum'), agg.get('rolling_acum'),
                        f"=N{i}-O{i}",
                        f"=N{i}/O{i}",
                        agg.get('gastos_comp_year'), agg.get('rolling_comp_year'),
                        classify.cuenta.p_y_g
                    ])
                    i+=1
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'Analisis de gastos'
                
                ws.append(["Clasificación", "Tipo", "Real/Ppto", "Cuenta", "Nombre cuenta", "Centro Costo","Nit Tercero","Nombre Tercero","Responsable",
                           "Real","Rolling","Diferencia","% Cumplimiento","Real acumulado","Rolling acumulado","Diferencia","% Cumplimiento",
                           "Gasto año anterior","Rolling año anterior","P&G"])
                
                for row in report:
                    ws.append(row)
                
                if include_details == 'true':                    
                    ws_detail = wb.create_sheet('Detalle de gastos')
                    ws_detail.append([
                        "Id","Fecha","Cuenta","Nit Tercero","Nombre Tercero","Importe","Referencia","Cebe","Texto",
                        "Ceco","Num Doc","Posicion","Clasificacion"
                    ])
                    
                    obj_gastos = report_models.gastos.objects.filter(fecha__year = year
                                                                     ).order_by('fecha').values_list(
                                                                         "pk","fecha","cuenta","nit_tercero","nombre_tercero","importe",
                                                                         "referencia","cebe","texto","ceco","n_doc","posicion",
                                                                         "clasificacion"
                                                                     )
                                                                     
                    if area != "TODAS":
                        obj_gastos = obj_gastos.filter(clasificacion__responsable = area)
                    
                    for row in obj_gastos:
                        ws_detail.append(row)
                    
                    for cell in ws_detail['F']:
                        cell.style = 'Comma'

                    ws_detail.freeze_panes = "A2" 
                    
                    ws_detail.column_dimensions['B'].width = 10
                    ws_detail.column_dimensions['C'].width = 12
                    ws_detail.column_dimensions['D'].width = 13
                    ws_detail.column_dimensions['E'].width = 24
                    ws_detail.column_dimensions['F'].width = 17
                    ws_detail.column_dimensions['G'].width = 18
                    ws_detail.column_dimensions['I'].width = 55
                    ws_detail.column_dimensions['K'].width = 11
                    ws_detail.column_dimensions['M'].width = 11
                    
                    
                    ws_detail.sheet_view.zoomScale = 80
                    
                    
                rule = IconSetRule('5Arrows', 'num', [0, 0.9, 0.99, 1.01, 1.1], showValue=None, percent=None, reverse=True)
                
                ws.column_dimensions['B'].width = 9.4
                ws.column_dimensions['C'].width = 9.2
                ws.column_dimensions['D'].width = 14.1
                ws.column_dimensions['E'].width = 49
                ws.column_dimensions['F'].width = 22
                ws.column_dimensions['G'].width = 24
                ws.column_dimensions['H'].width = 41.5
                ws.column_dimensions['I'].width = 23
                ws.column_dimensions['J'].width = 20
                ws.column_dimensions['K'].width = 20
                ws.column_dimensions['L'].width = 19.2
                ws.column_dimensions['M'].width = 18.7
                ws.column_dimensions['N'].width = 18.7
                ws.column_dimensions['O'].width = 19.3
                ws.column_dimensions['P'].width = 18.7
                ws.column_dimensions['Q'].width = 18.7
                ws.column_dimensions['R'].width = 18.7
                ws.column_dimensions['S'].width = 18.7
                ws.column_dimensions['T'].width = 18.7
                
                ws[f'J{obj_classification.count()+2}'] = f'=SUM(J2:J{obj_classification.count()+1})'
                ws[f'K{obj_classification.count()+2}'] = f'=SUM(K2:K{obj_classification.count()+1})'
                
                rows_currency = ['J','K','L','N','P','O']
                rows_percent = ['M','Q']
                
                for row in rows_currency:
                    for cell in ws[row]:
                        cell.style = 'Comma'
                            
                for row in rows_percent:
                    for cell in ws[row]:
                        cell.style = 'Percent'
                
                
                k = obj_classification.count()+1
                ws.conditional_formatting.add(f"M2:M{k}", rule)
                ws.conditional_formatting.add(f"Q2:Q{k}", rule)
                
                ws.conditional_formatting.add(f'L2:L{k}', CellIsRule(operator='lessThanOrEqual', formula=['0'],font = Font(color = '008E40')))
                ws.conditional_formatting.add(f'P2:P{k}', CellIsRule(operator='lessThanOrEqual', formula=['0'],font = Font(color = '008E40')))

                ws.conditional_formatting.add(f'L2:L{k}', CellIsRule(operator='greaterThan', formula=['0'],font = Font(color = '00FF0000')))
                ws.conditional_formatting.add(f'P2:P{k}', CellIsRule(operator='greaterThan', formula=['0'],font = Font(color = '00FF0000')))

                
                last_row = obj_classification.count()+1
                        
                for row in ws['A1:T1']:
                    for cell in row:
                        cell.font = Font(bold=True)
                
                for row in ws[f'A{last_row + 1}:Q{last_row + 1}']:
                    for cell in row:
                        cell.font = Font(bold=True)
                
                ws.freeze_panes = "F2"                    
                
                ws.sheet_view.zoomScale = 70
                
                area = 'GENERAL' if area == 'TODAS' else area
                
                filename = f'Reporte_gastos_{area}_{month}_{year}.xlsx'
                    
                wb.save(settings.MEDIA_ROOT / f"tmp/{filename}")
                
                data = {
                    'href':f'tmp/{filename}',
                    'data': report,
                }
                
                
                return JsonResponse(data)
        
        
    return render(request,'landing.html',context)

@login_required
@rol_permission('Control gestion')
def analyze_expenses(request,date_from,date_to):
    
    context = {
        'from':date_from,
        'to':date_to
    }
    
    if request_is_ajax(request):
        if request.method == 'GET':
            todo = request.GET.get('todo')
            
            if todo == 'generate_report':
                month = request.GET.get('month')
                year = request.GET.get('year')
                include_acumulate = request.GET.get('include_acumulate')
                comparate_year = request.GET.get('comparate_year')
                
                obj_classification = report_models.clasificacion.objects.all()
                
                
                report = []
                for classify in obj_classification:
                    agg = classify.analysis_by_period(month, year, include_acumulate, comparate_year)
                    report.append([
                        classify.pk, classify.tipo, classify.tipo_r_p,
                        classify.cuenta.cuenta, classify.centro_costo, classify.nit_tercero,
                        classify.nombre_tercero, classify.responsable,
                        agg.get('gastos'),agg.get('rolling'),
                        agg.get('gastos_acum'), agg.get('rolling_acum'),
                        agg.get('gastos_comp_year'), agg.get('rolling_comp_year'),
                    ])
                
                wb = l.Workbook()
                ws = wb.active
                
                ws.append(["Id", "Tipo", "Real/Ppto", "Cuenta", "Centro Costo","Nit Tercero","Nombre Tercero","Responsable",
                           "Real","Rolling","Gasto acumulado","Rolling acumulado","Gasto año anterior","Rolling año anterior"])
                
                for row in report:
                    ws.append(row)
                
                wb.save(settings.MEDIA_ROOT / "tmp/report.xlsx")
                
                data = {
                    'href':'tmp/report.xlsx',
                    'data': report,
                }
                
                return JsonResponse(data)
                           
    
    return render(request, 'look_expenses.html', context)

@login_required
@rol_permission('Control gestion')
def expenses_detail(request):
    
    if request.method == 'GET' and request.GET:
        
        if request_is_ajax(request):
            todo = request.GET.get('todo')
            
            if todo == 'getexpenses':
                date_from = request.GET.get('date_from')
                date_to = request.GET.get('date_to')
                campo = request.GET.get('campo')
                valor_campo = request.GET.get('valor')
                """ print(campo, valor_campo) """
                
                cuenta = request.GET.get('0')
                fecha = request.GET.get('1')
                nit  = request.GET.get('2')
                nombre = request.GET.get('3')
                ceco = request.GET.get('4')
                centro = request.GET.get('5')
                clasif = request.GET.get('6')
                
                is_not_searching = cuenta == None and fecha == None and nit == None and \
                                nombre == None and ceco == None and centro == None and \
                                    clasif == None
                
                draw = request.GET.get('draw')
                start = request.GET.get('start')
                length = request.GET.get('length')
                search_val = request.GET.get('search[value]')
                
                
                
                obj_expenses = report_models.gastos.objects.filter(
                    fecha__gte = date_from, fecha__lte = date_to,
                ).order_by('fecha')  
                
                if campo!= None and campo != '':
                    
                    obj_expenses = obj_expenses.filter(
                        **{campo+"__icontains":valor_campo}
                    )
                                
                
                if not is_not_searching:
                    if cuenta != None and cuenta != "":
                        obj_expenses = obj_expenses.filter(
                            cuenta__pk__icontains = cuenta
                        )
                    if fecha != None and fecha != "":
                        obj_expenses = obj_expenses.filter(
                            fecha__icontains = fecha
                        )
                    if nit != None and nit != "":
                        obj_expenses = obj_expenses.filter(
                            nit_tercero__icontains = nit
                        )
                    if nombre != None and nombre != "":
                        obj_expenses = obj_expenses.filter(
                            nombre_tercero__icontains = nombre
                        )
                    if ceco != None and ceco != "":
                        obj_expenses = obj_expenses.filter(
                            ceco__pk__icontains = ceco
                        )
                    if centro != None and centro != "":
                        obj_expenses = obj_expenses.filter(
                            ceco__cedo__icontains = centro
                        )
                    if clasif != None and clasif != "":
                        obj_expenses = obj_expenses.filter(
                            clasificacion__pk__icontains = clasif
                        )
                
                total_expenses = obj_expenses.aggregate(total=Sum('importe')).get('total')
                if total_expenses == None: total_expenses = 0
                
                if int(length) != -1:
                    defered_list = obj_expenses[int(start):int(start)+int(length)]
                else:
                    defered_list = obj_expenses
                
                """ data={
                    'data':JsonRender(obj_expenses).render(),
                    } """
                
                data={
                    "draw": int(draw),
                    "recordsTotal": obj_expenses.count(),
                    "recordsFiltered": obj_expenses.count(),
                    'data':JsonRender(defered_list).render(),
                    'totals':total_expenses,
                }
                
                return JsonResponse(data)
            
            elif todo == 'getexpensesbyclassify':
                month = request.GET.get('month')
                year = request.GET.get('year')
                classify = request.GET.get('classify')
                data = {'data':[]}
                
                if classify:
                    obj_gastos = report_models.gastos.objects.filter(
                        fecha__month = month, fecha__year = year,
                        clasificacion = classify
                    ).order_by('fecha')
                    
                    json_gastos = JsonRender(obj_gastos)

                    data = {
                        'data':json_gastos.render()
                    }
                
                return JsonResponse(data)
                    
            elif todo == 'rolling':
                month = request.GET.get('month')
                year = request.GET.get('year')
                
                rolling_objects = report_models.rollling.objects.filter(mes=month, anio=year)
                
                json_rolling = JsonRender(rolling_objects,query_functions=['real',])
                
                data = {
                    'data': json_rolling.render()
                }
                
                return JsonResponse(data)
            
            elif todo == 'getrules':
                obj_rules = report_models.rules.objects.all()
                
                data = {
                    'data': JsonRender(obj_rules).render()
                }
                
                return JsonResponse(data)
            
            elif todo == 'getrollingfile':
                obj_classify = report_models.clasificacion.objects.filter(activo=True).order_by('pk')
                
                wb = openpyxl.Workbook()
                ws = wb.active
                
                ws.append(["Id", "Tipo","Real/Ppto","Responsable","Cuenta","Ceco/Orden","Nit","Tercero","Valor"])
                
                for row in obj_classify:
                    ws.append([row.pk,row.tipo,row.tipo_r_p,row.responsable,row.cuenta.pk,row.centro_costo,row.nit_tercero,row.nombre_tercero,""])
                
                filepath = settings.MEDIA_ROOT / "tmp/rolling_to_upload.xlsx"
                
                wb.save(filepath)
                
                data = {
                  'url': settings.MEDIA_URL + 'tmp/rolling_to_upload.xlsx'
                }
                        
                return JsonResponse(data)
                
                
        
    elif request.method == 'POST':
        if request_is_ajax(request):
            todo = request.POST.get('todo')
            
            
            
            if todo == 'reclasify':
                id_expense = request.POST.get('id_expense')
                new_classify = request.POST.get('new_classify')
                
                obj_expense = report_models.gastos.objects.get(pk=id_expense)
                obj_classify = report_models.clasificacion.objects.get(pk=new_classify)
                
                obj_expense.clasificacion = obj_classify
                obj_expense.save()
                
                data = {
                    
                }
                
                return JsonResponse(data)
            
            elif todo == 'reclasify-multiple':
                expenses_to_reclassify = request.POST.getlist('classify_list[]')
                new_classify = request.POST.get('new_classify')
                obj_classify = report_models.clasificacion.objects.get(pk=new_classify)

                for i in expenses_to_reclassify:
                    expense = report_models.gastos.objects.get(pk=i)
                    expense.clasificacion = obj_classify
                    expense.save()
                
                data = {
                    
                }
                
                return JsonResponse(data)
                
            elif todo == 'add-rule':
                
                rules_if = request.POST.getlist('rule-if-field')
                rules_conditions = request.POST.getlist('rule-condition')
                rules_values = request.POST.getlist('rule-if-value')
                and_or = request.POST.getlist('and-or-rule')
                statement_field = request.POST.get('statement_field_name')
                statement_value = request.POST.get('statement_field_value')
                
                rule = ''
                for i in range(len(rules_if)):
                    if 'pk' in rules_if[i]:
                        if rules_conditions[i] == 'in':
                            rule += f'''{rules_values[i]} {rules_conditions[i]} object_expense.{rules_if[i]}'''
                        else:
                            rule += f'''{rules_values[i]} {rules_conditions[i]} object_expense.{rules_if[i]}'''
                    else: 
                        if rules_conditions[i] == 'in':
                            rule += f'''"{rules_values[i].lower()}" {rules_conditions[i]} object_expense.{rules_if[i]}.lower()'''
                        else:
                            rule += f'''"{rules_values[i]}" {rules_conditions[i]} object_expense.{rules_if[i]}'''
                    
                    if and_or[i] !="":
                        rule += f' {and_or[i]} '
                        
                if statement_field == 'clasificacion':
                    statement_value = f'report_models.clasificacion.objects.get(pk={statement_value})'
                elif statement_field == 'ceco':
                    statement_value = f'report_models.centro_costo.objects.get(pk={statement_value})'
                else:
                    statement_value = f'"{statement_value}"'
                
                statement = f'object_expense.{statement_field} = {statement_value}'
                
                report_models.rules.objects.create(
                    premise = rule, statement = statement, active = True
                )
                
                data = {
                    'response': 'Se creo la regla'
                }
                
                return JsonResponse(data)

            elif todo == 'close-rolling':
                month = request.POST.get('month')
                year = request.POST.get('year')
                
                obj_rolling = report_models.rollling.objects.filter(
                        mes = month, anio = year
                )
                
                for rolling in obj_rolling:
                    agg = rolling.clasificacion.analysis_by_period(month, int(year), False, False)
                    total_gasto = agg.get('gastos')
                    rolling.valor_previo = rolling.valor
                    rolling.valor = total_gasto
                    rolling.save()
                
                    
                data = {
                    
                }
                
                return JsonResponse(data)
            
        else:
            expenses_file = request.FILES.get('expenses-file')
            rolling_file = request.FILES.get('rolling-file')
            
            if expenses_file:
                book = openpyxl.load_workbook(expenses_file)        
                sheet = book.active
                
                has_info = True
                i = 2
                no_upload_lines = []
                
                while has_info:            
                    
                    cuenta = sheet.cell(i,1).value
                    if cuenta == "" or cuenta == None:                
                        has_info = False
                        break
                    
                    posicion = sheet.cell(i,14).value
                    n_doc = sheet.cell(i,8).value
                    
                    obj_line = report_models.gastos.objects.filter(
                        posicion = posicion, n_doc = n_doc
                    )
                    
                    if obj_line.exists():
                        no_upload_lines.append([i,f'La linea con el documento {n_doc} y la posición {posicion} ya existe en la BD'
                            ])
                        i+=1
                        continue
                    
                    fecha = sheet.cell(i,7).value
                    
                    obj_cierre = report_models.cierre_de_mes.objects.filter(
                        mes = fecha.month, annio = fecha.year
                    )
                    
                    if obj_cierre.exists():
                        estado = obj_cierre[0].estado
                        if estado == 'Cerrado':
                            no_upload_lines.append([i,'El periodo de gasto para esta linea ya está cerrado'])
                    else:
                        report_models.cierre_de_mes.objects.create(
                            mes = fecha.month, annio = fecha.year,
                            estado = 'Abierto'
                        )                    
                    
                    centro_costo = sheet.cell(i,5).value
                    if centro_costo == "" or centro_costo == None:
                        centro_costo = sheet.cell(i,6).value
                            
                    nit_tercero = sheet.cell(i,2).value
                    nombre_tercero = sheet.cell(i,3).value
                    importe = sheet.cell(i,4).value
                    fecha = sheet.cell(i,7).value
                    referencia = sheet.cell(i,9).value
                    importe_usd = sheet.cell(i,10).value
                    cebe = sheet.cell(i,11).value
                    texto = sheet.cell(i,12).value
                    asignacion = sheet.cell(i,13).value
                    
                    try:
                        obj_cuenta = report_models.cuentas_contables.objects.get(
                            pk=cuenta
                        )
                        obj_cc = report_models.centro_costo.objects.get(
                            pk=centro_costo
                        )
                        report_models.gastos.objects.create(
                            cuenta = obj_cuenta, nit_tercero = nit_tercero,
                            nombre_tercero = nombre_tercero, importe = importe,
                            ceco = obj_cc, fecha = fecha, n_doc = n_doc, posicion = posicion,
                            referencia = referencia, importe_usd = importe_usd,
                            cebe = cebe, texto = texto, asignacion = asignacion,
                            usuario_carga = request.user, 
                        )
                    except exceptions.ObjectDoesNotExist:
                        no_upload_lines.append([i,'La cuenta o el centro de costo/orden no existe en la base de datos'])
                    i+=1
                
                wb = openpyxl.Workbook()
                ws = wb.active
                
                ws.append(["Line", "Issue"])
                
                for row in no_upload_lines:
                    ws.append(row)
                
                filepath = settings.MEDIA_ROOT / "tmp/no_upload_lines.xlsx"
                
                wb.save(filepath)
                
                return FileResponse(open(filepath, "rb"), as_attachment= True, filename = "upload_results.xlsx")
            
            elif rolling_file:
                month = request.POST.get('month-rolling')
                year = request.POST.get('year-rolling')
                not_override = request.POST.get('override_uploaded_already')
                
                book = openpyxl.load_workbook(rolling_file)        
                sheet = book.active
                
                has_info = True
                i = 2
                no_upload_lines = []
                
                while has_info:   
                          
                    id_classify = sheet.cell(i,1).value
                    value = sheet.cell(i,9).value
                    
                    if id_classify == "" or id_classify == None:                
                        has_info = False
                        break
                    
                    obj_classify = report_models.clasificacion.objects.filter(pk=id_classify)
                    
                    if obj_classify.exists():
                        c = obj_classify[0]
                    
                        rolling_line = report_models.rollling.objects.filter(
                            mes = month, anio = year, clasificacion = c)
                        if rolling_line.exists():
                            if not_override == 'on':
                                i+=1
                                continue
                            else:
                                rolling_line[0].valor = value
                                rolling_line[0].save()
                        else:                            
                            report_models.rollling.objects.create(
                                mes = month, anio = year,
                                clasificacion = c, valor = value,
                            )
                    else:
                        no_upload_lines.append([
                            i, 'La calsificacion no existe en la base de datos'
                        ])
                    i+=1
                       
                wb = openpyxl.Workbook()
                ws = wb.active
                
                ws.append(["Line", "Issue"])
                
                for row in no_upload_lines:
                    ws.append(row)
                
                filepath = settings.MEDIA_ROOT / "tmp/no_upload_lines.xlsx"
                
                wb.save(filepath)
                
                return FileResponse(open(filepath, "rb"), as_attachment= True, filename = "upload_results.xlsx")
                        
        
    obj_classify = report_models.clasificacion.objects.all(
                    ).order_by('pk')
    responsables =   report_models.areas_empresa.objects.all().order_by('nombre_area')
                                                        
    campos = [f for f in report_models.gastos._meta.get_fields(include_parents=True)]
    to_pop = (0,3,4,4,5,4,-1,-2,-2,0,2,4)
    for _ in to_pop: campos.pop(_)
    campos.append({
        'name': 'cuenta.pk',
        'verbose_name': 'cuenta'
    })
    campos.append(
        {
        'name': 'ceco.pk',
        'verbose_name': 'centro de costo/orden'
    }
    )
    campos.append(
        {
        'name': 'clasificacion.pk',
        'verbose_name': 'clasificacion'
    }
    )
    
    cedos = report_models.centro_costo.objects.distinct('cedo')
    
    
    
    expenses = report_models.gastos.objects.filter(
        fecha__year__lte = datetime.now().year,
        fecha__month__lte=7#datetime.now().month
    ).values('fecha','nit_tercero','nombre_tercero','texto','importe','clasificacion__pk','cuenta__cuenta')
    
    
    data = {
        'data': list(expenses)
    }
    
    campos_no_tabla = [
        {'name':'n_doc','verbose_name':'Numero documento'},
        {'name':'referencia','verbose_name':'referencia'},
        {'name':'texto','verbose_name':'texto'},
        {'name':'asignacion','verbose_name':'asignacion'},
    ]
    
    context = {
        'classify':obj_classify,
        'cuentas':report_models.cuentas_contables.objects.all().order_by('pk'),
        'cedos': cedos,#report_models.centro_costo.objects.all().order_by('pk'),
        'responsables': responsables,
        'gastos_campos': campos,
        'campos_adicionales':campos_no_tabla,
        'areas': report_models.areas_empresa.objects.all().order_by('nombre_area'),
        'dashboard': {
            'expenses': data
        }
    }
    
    return render(request, 'expenses_detail.html', context)

@login_required
@rol_permission('Control gestion')
def classifications(request):
    
    #if request_is_ajax(request):
        
    if request.method == 'GET' and request.GET:
        obj_classify = report_models.clasificacion.objects.all().order_by('pk')
        api = request.GET.get('api')
        
        if api == 'datatable':
            
            data = {
                'data': JsonRender(obj_classify).render()
            }
            
            
            return JsonResponse(data)
        
        elif api == 'dropdown':
            results = []
            for c in obj_classify:
                cc = f'| {c.centro_costo}' if c.centro_costo != None else ""
                tercero = f' | {c.nit_tercero}-{c.nombre_tercero}' if c.nit_tercero != None else ""
                results.append({
                    'name': f'({c.id}) {c.responsable} | {c.cuenta}{cc}{tercero}',
                    'value': c.id
                })
            
            data = {
                'success':True,
                'results':results
            }
            
            return JsonResponse(data)
    
    elif request.method == 'POST':
        tipo = request.POST.get('tipo_classify')
        real_ppto = request.POST.get('tipo_r_p_classify')
        cuenta = request.POST.get('cuenta_classify')
        cedo = request.POST.get('ceco_classify')
        nit_tercero = request.POST.get('nit_tercero_classify')
        tercero = request.POST.get('nombre_tercero_classify')        
        responsable = request.POST.get('responsable_classify')
        
        obj_cuenta = report_models.cuentas_contables.objects.get(pk=cuenta)
        
        
        check_classify = report_models.clasificacion.objects.filter(
            tipo = tipo, tipo_r_p = real_ppto,
            cuenta = obj_cuenta, centro_costo = cedo,
            nit_tercero = nit_tercero, nombre_tercero = tercero,
            responsable = responsable
        )
        if check_classify.exists(): 
            msj = f'Ups :( La clasificacion que intentas crear ya existe, id={check_classify[0].pk}'
            class_css = 'error'
        else:
            report_models.clasificacion.objects.create(
                tipo = tipo, tipo_r_p = real_ppto,
                cuenta = obj_cuenta, centro_costo = cedo,
                nit_tercero = nit_tercero, nombre_tercero = tercero,
                responsable = responsable
            )
            msj = f'Magnifico!! la clasificación fue creada sin problemas :)'
            class_css = 'success'
        
        data = {
            'msj': msj,
            'class': class_css
        }
        
        return JsonResponse(data)

@login_required       
@rol_permission('Control gestion') 
def expenses(request):
    if request.method == 'GET':
        to_do = request.GET.get('todo')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        if to_do == 'pivot':

            obj_gastos = report_models.gastos.objects.filter(
                fecha__gte = date_from, fecha__lte=date_to
            )
            
            json_gastos = {
                'data':list(obj_gastos.values(
                    'pk','cuenta__pk','cuenta__nombre_cuenta','fecha','ceco__pk','nombre_tercero','nit_tercero',
                    'ceco__cedo','clasificacion__responsable','importe_usd','texto',
                    'importe','clasificacion__pk'
                ))
                }
            
            return JsonResponse(json_gastos)
        
    if request.method == 'POST':
        if request_is_ajax(request):
            to_do = request.POST.get('todo')
            date_from = request.POST.get('date_from_classify')
            date_to = request.POST.get('date_to_classify')
                    
            if to_do == 'classify':
                override = request.POST.get('override_classified_already')
                
                obj_gastos = report_models.gastos.objects.filter(
                    fecha__gte = date_from, fecha__lte = date_to)
                
                if override == 'on':
                    obj_gastos = obj_gastos.filter(clasificacion__isnull=True)
                
                no_clasificados = []
                
                obj_rules = report_models.rules.objects.filter(active=True)
                
                for gasto in obj_gastos:
                    rule_is_classify = check_rules(gasto,obj_rules)
                    
                    if rule_is_classify:
                        continue
                    
                    clasificador = clasificar_gasto(gasto)
                    
                    if clasificador[0]:
                        gasto.clasificacion = clasificador[1]
                        gasto.save()
                    else:
                        
                        no_clasificados.append([
                            gasto.pk, gasto.cuenta.pk, gasto.nit_tercero,
                            gasto.nombre_tercero, gasto.importe,
                            gasto.ceco.pk, gasto.texto,
                        ])
                        
                
                wb = openpyxl.Workbook()
                ws = wb.active
                
                ws.append(["Id", "cuenta","nit","nombre","importe","centro_costo-orden","texto"])
                
                for row in no_clasificados:
                    ws.append(row)
                
                filepath = settings.MEDIA_ROOT / "tmp/no_classified_expenses.xlsx"
                
                wb.save(filepath)
                
                data = {
                  'url': settings.MEDIA_URL + 'tmp/no_classified_expenses.xlsx'
                }
                        
                return JsonResponse(data)
                
        
        
        filename = 'gastos.json'
        filepath = settings.MEDIA_ROOT/f'tmp/{filename}'
        
@login_required        
@rol_permission('Control gestion')
def delete_classify(request):
    
    if request.method == 'GET':
        actual = request.GET.get('desde')
        nueva = request.GET.get('nueva')
        
        status='error'
        if actual == nueva:
            msj = 'La clasificación nueva no puede ser igual a la actual'
        elif actual == '' or actual == None:
            msj = 'Debes enviar una clasificación actual'
        elif nueva == '' or nueva == None:
            msj = 'Debes enviar una clasificación nueva'
        else:
            status='success'
            msj = f'La clasificación N°{actual} fue eliminada, los gastos y rollings asociados fueron trasladados a la clasificación N°{nueva}'
            unificar_clasificaciones(actual,nueva)
            
        data = {
            'status':status,
            'response': msj
        }
        
        return JsonResponse(data)

@login_required        
@rol_permission('Control gestion')
def dashboarddata(request):
    
    if request.method == 'GET':
        todo = request.GET.get('todo')
        year = request.GET.get('year')
        month = request.GET.get('month')
        
        if todo == 'expenses':
            
            if year is None: year = datetime.now().year
            if month is None: month = datetime.now().month
            
            
            expenses = report_models.gastos.objects.filter(
                fecha__year__lte = year,
            ).values('fecha__year','fecha__month'
            ).order_by('fecha__year','fecha__month'
            ).annotate(total=Sum('importe'))
            
            data = []
            data.append(['fecha','gasto_ejecutado','rolling'])
            for expense in expenses:
                year = expense.get('fecha__year')
                month = expense.get('fecha__month')
                rolling = report_models.rollling.objects.filter(
                    mes = month, anio = year
                ).values('mes','anio').order_by('mes','anio').annotate(total=Sum('valor'))
                rolling = rolling.first()
                rolling = 0 if rolling.get('total') is None else rolling.get('total')
                
                
                
                data.append([
                    int(datetime(year, month, 1, 0,0,0,0).timestamp() * 1000),
                    expense.get('total'),
                    rolling
                ])
                
            
            return JsonResponse(data, safe=False)
            
        elif todo == 'topexpensesareas':
            
            if year is None: year = datetime.now().year
            if month is None: month = datetime.now().month
            
            expenses = report_models.gastos.objects.filter(
                fecha__year = year,
            ).values('fecha__year','clasificacion__responsable'
            ).order_by('fecha__year','clasificacion__responsable'
            ).annotate(total=Sum('importe')
            ).order_by('-total')
            
            data = []
            data.append(['year','area','real','rolling','desviacion'])
            
            for exp in expenses[:5]:
                year = exp.get('fecha__year')
                responsable = exp.get('clasificacion__responsable')
                total = exp.get('total',0)
                
                rolling = report_models.rollling.objects.filter(
                    anio = year, clasificacion__responsable = responsable
                ).aggregate(total=Sum('valor'))
                
                rolling = 0 if rolling.get('total') is None else rolling.get('total')
                desv = round((total - rolling)*100/rolling,2)
                data.append([
                    year,
                    responsable,
                    total,
                    rolling,
                    desv
                ])
            return JsonResponse(data, safe=False)
        
        elif todo == 'topdifferencerolling':
            
            if year is None: year = datetime.now().year
            if month is None: month = datetime.now().month
            
            expenses = report_models.gastos.objects.filter(
                fecha__year = year, clasificacion__responsable__isnull = False,
            ).values('fecha__year','clasificacion__responsable'
            ).order_by('fecha__year','clasificacion__responsable'
            ).annotate(total=Sum('importe')
            ).order_by('-total')
            
            data = []
            
            for exp in expenses:
                year = exp.get('fecha__year')
                responsable = exp.get('clasificacion__responsable')
                total = exp.get('total',0)
                rolling = report_models.rollling.objects.filter(
                    anio = year, clasificacion__responsable = responsable
                ).aggregate(total=Sum('valor'))
                
                rolling = 0 if rolling.get('total') is None else rolling.get('total')
                
                difference = total - rolling
                
                data.append([
                    year,
                    responsable,
                    total,
                    rolling,
                    difference
                ])
            
            data = sorted(data,key=itemgetter(4),reverse=True)
            data = data[:5]
            data.insert(0,['year','area','real','rolling','diferencia'])
            
            return JsonResponse(data, safe=False)
 
def unificar_clasificaciones(actual, hacia):
     
    nueva_clasificacion = report_models.clasificacion.objects.get(pk=hacia)
    clasificacion_actual = report_models.clasificacion.objects.get(pk=actual)
    
    #Cambio los gastos asociados de la clasificacion anterior a la nueva
    gastos_asociados = report_models.gastos.objects.filter(clasificacion=actual)
    for gasto in gastos_asociados:
        gasto.clasificacion = nueva_clasificacion
        gasto.save()
    
    #Sumo los rollings de la clasificacion anterior a la nueva y elimino los rollings
    rollings = report_models.rollling.objects.filter(clasificacion = actual)
    
    for r in rollings:
        mes  = r.mes
        annio = r.anio
        valor = r.valor
        
        r_hacia = report_models.rollling.objects.get(mes=mes, anio = annio, clasificacion = hacia)
        
        nuevo_valor = r_hacia.valor + valor
        
        
        r_hacia.valor = nuevo_valor
        r_hacia.save()
        
        r.delete()
    
    #Elimino la clasificacion
    clasificacion_actual.delete()
 
def clasificar_gasto(obj_gasto):    
    
    if obj_gasto.ceco != None:
        check_1 = report_models.clasificacion.objects.filter(
            cuenta = obj_gasto.cuenta, centro_costo = obj_gasto.ceco.cedo,
            nit_tercero = obj_gasto.nit_tercero
        )
        
        if check_1.exists():
            return True, check_1[0]

        check_2 = report_models.clasificacion.objects.filter(
            Q(nit_tercero = "")|Q(nit_tercero = None),
            cuenta = obj_gasto.cuenta, centro_costo = obj_gasto.ceco.cedo
        )
        
        if check_2.exists():
            return True, check_2[0]
    
    check_3 = report_models.clasificacion.objects.filter(
        Q(centro_costo = "")|Q(centro_costo = None),
        cuenta = obj_gasto.cuenta, nit_tercero = obj_gasto.nit_tercero
    )
    
    if check_3.exists():
        return True, check_3[0]
    
    check_4 = report_models.clasificacion.objects.filter(
        Q(nit_tercero = "")|Q(nit_tercero = None),
        Q(centro_costo = "")|Q(centro_costo = None),
        cuenta = obj_gasto.cuenta
    )
    
    if check_4.exists():
        return True, check_4[0]
    
    return False, None

def check_rules(object_expense,rules):
    change_clasification = False
    for rule in rules:

        if eval(rule.premise):
            stmt = rule.statement.split(" = ")
            stmt_obj = stmt[0]
            stmt_key = stmt[1].replace('"','')
            if stmt_obj == 'object_expense.clasificacion.pk':
                change_clasification = True
                object_expense.clasificacion = report_models.clasificacion.objects.get(pk=stmt_key)
            else:
                exec(rule.statement)
            object_expense.save()
        
                
    return change_clasification

def request_is_ajax(request):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest': 
        return True
    return False