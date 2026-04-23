import argparse
import json
import os
import sys
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'Atlantic.settings')

import django

django.setup()

from openpyxl import load_workbook
from django.db import transaction
from human_resources.models import area, cargo, estructura, historico_base_teorica

DEFAULT_INPUT = '/code/data de Personal Easy Data (2).xlsx'
DEFAULT_SHEET = 'BD PERSONAL'
DEFAULT_BACKUP_DIR = '/code/static_media/tmp'

STRUCTURE_MAP = {
    'OPERATIVO': 'OPERACIONES',
    'ADMON': 'ADMIN Y FRA',
    'COMERCIAL': 'COMERCIAL',
    'GENERAL': 'GENERAL',
    'ABASTECIMIENTO': 'ABASTECIMIENTO',
}


def clean_text(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def normalize_header(value):
    if value is None:
        return ''
    value = str(value)
    value = ''.join(ch for ch in unicodedata.normalize('NFKD', value) if not unicodedata.combining(ch))
    return re.sub(r'\s+', ' ', value).strip().upper()


def most_common_value(counter):
    if not counter:
        return None
    return counter.most_common(1)[0][0]


def load_snapshot(path, sheet_name):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    raw_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {normalize_header(h): pos for pos, h in enumerate(raw_headers)}

    area_rows = Counter()
    all_cargo_keys = set()       # todos los cargos vistos (para crear el registro)
    cargo_rows = Counter()       # solo activos+vacantes (para cantidad_aprobada)
    cargo_tap = defaultdict(Counter)
    cargo_tipo_posicion = defaultdict(Counter)

    NULL_VALS = {'', 'NONE', 'N/A', 'NA', '#N/A'}
    CONTABLE_ESTADOS = {'ACTIVO', 'VACANTE'}

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_structure = clean_text(row[idx['ESTRUCTURA']])
        raw_area = clean_text(row[idx['AREA']])
        raw_cargo = clean_text(row[idx['CARGO']])
        raw_estado = (clean_text(row[idx['ESTADO']]) or '').upper() if 'ESTADO' in idx else ''
        if not raw_structure or not raw_area:
            continue
        mapped_structure = STRUCTURE_MAP.get(raw_structure)
        if not mapped_structure:
            continue
        area_rows[(raw_structure, mapped_structure, raw_area)] += 1
        if raw_cargo:
            key = (raw_structure, mapped_structure, raw_area, raw_cargo)
            all_cargo_keys.add(key)
            # cantidad_aprobada = activos + vacantes (dotacion ideal)
            if raw_estado in CONTABLE_ESTADOS:
                cargo_rows[key] += 1
            tap_val = clean_text(row[idx.get('TAP', -1)]) if idx.get('TAP') is not None else None
            pos_val = clean_text(row[idx.get('TIPO DE POSICION', -1)]) if idx.get('TIPO DE POSICION') is not None else None
            if tap_val and tap_val.upper() not in NULL_VALS:
                cargo_tap[key][tap_val.upper()] += 1
            if pos_val and pos_val.upper() not in NULL_VALS:
                cargo_tipo_posicion[key][pos_val.upper()] += 1

    return area_rows, cargo_rows, all_cargo_keys, cargo_tap, cargo_tipo_posicion


def serialize_current_catalog():
    return {
        'timestamp_utc': datetime.utcnow().isoformat(),
        'areas': list(area.objects.select_related('estructura').values('id', 'descripcion', 'estructura_id', 'estructura__descripcion')),
        'cargos': list(cargo.objects.select_related('area').values('id', 'descripcion', 'area_id', 'area__descripcion', 'activo', 'cantidad_aprobada', 'criticidad')),
        'historico_base_teorica': list(historico_base_teorica.objects.values()),
    }


def ensure_backup_dir(path):
    Path(path).mkdir(parents=True, exist_ok=True)


def write_backup(backup_dir, payload):
    ensure_backup_dir(backup_dir)
    backup_path = Path(backup_dir) / f"hr_catalog_backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
    backup_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str))
    return str(backup_path)


def summarize(area_rows, cargo_rows, all_cargo_keys):
    by_structure = Counter()
    for (_, mapped_structure, _), count in area_rows.items():
        by_structure[mapped_structure] += count
    return {
        'excel_unique_areas': len(area_rows),
        'excel_unique_cargos': len(all_cargo_keys),
        'excel_cargos_con_dotacion': len(cargo_rows),
        'excel_rows_by_structure': dict(by_structure),
        'current_areas': area.objects.count(),
        'current_cargos': cargo.objects.count(),
        'current_historico_base_teorica': historico_base_teorica.objects.count(),
    }


def rebuild(area_rows, cargo_rows, all_cargo_keys, cargo_tap, cargo_tipo_posicion):
    structures = {item.descripcion: item for item in estructura.objects.all()}
    created_areas = {}
    created_cargos = 0

    with transaction.atomic():
        print('step=delete_start', flush=True)
        historico_base_teorica.objects.all().delete()
        cargo.objects.all().delete()
        area.objects.all().delete()
        print('step=delete_done', flush=True)

        print('step=areas_start', flush=True)
        for (raw_structure, mapped_structure, raw_area), _count in sorted(area_rows.items(), key=lambda item: (item[0][1], item[0][2])):
            structure_obj = structures[mapped_structure]
            area_obj = area.objects.create(descripcion=raw_area, estructura=structure_obj)
            created_areas[(mapped_structure, raw_area)] = area_obj

        print('step=areas_done', flush=True)
        print('step=cargos_start', flush=True)
        for key in sorted(all_cargo_keys, key=lambda k: (k[1], k[2], k[3])):
            _raw_structure, mapped_structure, raw_area, raw_cargo = key
            area_obj = created_areas[(mapped_structure, raw_area)]
            cantidad = cargo_rows.get(key, 0)
            tap_val = most_common_value(cargo_tap.get(key))
            tipo_posicion_val = most_common_value(cargo_tipo_posicion.get(key)) or 'HC'
            cargo.objects.create(
                descripcion=raw_cargo,
                area=area_obj,
                activo=True,
                cantidad_aprobada=cantidad,
                criticidad='MEDIA',
                tap=tap_val,
                tipo_posicion=tipo_posicion_val,
            )
            created_cargos += 1

        print('step=cargos_done', flush=True)
    return len(created_areas), created_cargos


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', default=DEFAULT_INPUT)
    parser.add_argument('--sheet', default=DEFAULT_SHEET)
    parser.add_argument('--backup-dir', default=DEFAULT_BACKUP_DIR)
    parser.add_argument('--apply', action='store_true')
    args = parser.parse_args()

    area_rows, cargo_rows, all_cargo_keys, cargo_tap, cargo_tipo_posicion = load_snapshot(args.input, args.sheet)
    summary = summarize(area_rows, cargo_rows, all_cargo_keys)
    print(json.dumps({'mode': 'apply' if args.apply else 'dry-run', 'summary': summary}, ensure_ascii=False, indent=2))

    if not args.apply:
        return

    print('step=backup_start', flush=True)
    backup_payload = serialize_current_catalog()
    backup_path = write_backup(args.backup_dir, backup_payload)
    print('step=backup_done', flush=True)
    created_areas, created_cargos = rebuild(area_rows, cargo_rows, all_cargo_keys, cargo_tap, cargo_tipo_posicion)
    print(json.dumps({
        'backup_path': backup_path,
        'created_areas': created_areas,
        'created_cargos': created_cargos,
        'remaining_historico_base_teorica': historico_base_teorica.objects.count(),
        'final_areas': area.objects.count(),
        'final_cargos': cargo.objects.count(),
    }, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
