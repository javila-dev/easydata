import argparse
import json
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "Atlantic.settings")

import django

django.setup()

from django.db import transaction
from django.db.models import Count
from django.db.models.deletion import ProtectedError
from openpyxl import load_workbook

from human_resources.models import area, cargo

DEFAULT_INPUT = "/code/Data personal activo.xlsx"
DEFAULT_SHEET = "BD PERSONAL"
DEFAULT_BACKUP_DIR = "/code/static_media/tmp"
SAMPLE_LIMIT = 20

STRUCTURE_MAP = {
    "OPERATIVO": "OPERACIONES",
    "ADMON": "ADMIN Y FRA",
    "COMERCIAL": "COMERCIAL",
    "GENERAL": "GENERAL",
    "ABASTECIMIENTO": "ABASTECIMIENTO",
}

NULL_VALS = {"", "NONE", "N/A", "NA", "#N/A"}


def clean_text(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def normalize_header(value):
    if value is None:
        return ""
    value = str(value)
    value = "".join(
        ch for ch in unicodedata.normalize("NFKD", value) if not unicodedata.combining(ch)
    )
    return re.sub(r"\s+", " ", value).strip().upper()


def normalize_text(value):
    if value is None:
        return ""
    value = "".join(
        ch for ch in unicodedata.normalize("NFKD", str(value)) if not unicodedata.combining(ch)
    )
    return re.sub(r"\s+", " ", value).strip().upper()


def most_common_value(counter):
    if not counter:
        return None
    return counter.most_common(1)[0][0]


def is_countable_row(raw_estado, raw_nombre):
    estado = normalize_text(raw_estado)
    nombre = normalize_text(raw_nombre)
    return estado == "ACTIVO" or "VACANTE" in estado or "VACANTE" in nombre


def build_snapshot(path, sheet_name):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    raw_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {normalize_header(h): pos for pos, h in enumerate(raw_headers)}

    required_headers = {"ESTRUCTURA", "AREA", "CARGO"}
    missing_headers = sorted(required_headers - set(idx.keys()))
    if missing_headers:
        raise ValueError(f"Headers faltantes en el Excel: {', '.join(missing_headers)}")

    rows_by_structure = Counter()
    area_keys = set()
    all_cargo_keys = set()
    countable_by_key = Counter()
    tap_by_key = defaultdict(Counter)
    tipo_posicion_by_key = defaultdict(Counter)

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_structure = clean_text(row[idx["ESTRUCTURA"]])
        raw_area = clean_text(row[idx["AREA"]])
        raw_cargo = clean_text(row[idx["CARGO"]])
        raw_estado = row[idx["ESTADO"]] if "ESTADO" in idx else None
        raw_nombre = row[idx["NOMBRE COMPLETO"]] if "NOMBRE COMPLETO" in idx else None

        if not raw_structure or not raw_area:
            continue

        mapped_structure = STRUCTURE_MAP.get(raw_structure)
        if not mapped_structure:
            continue

        area_key = (mapped_structure, raw_area)
        rows_by_structure[mapped_structure] += 1
        area_keys.add(area_key)

        if not raw_cargo:
            continue

        cargo_key = (mapped_structure, raw_area, raw_cargo)
        all_cargo_keys.add(cargo_key)

        if is_countable_row(raw_estado, raw_nombre):
            countable_by_key[cargo_key] += 1

        if "TAP" in idx:
            tap_val = clean_text(row[idx["TAP"]])
            if tap_val and normalize_text(tap_val) not in NULL_VALS:
                tap_by_key[cargo_key][tap_val.upper()] += 1

        if "TIPO DE POSICION" in idx:
            tipo_posicion_val = clean_text(row[idx["TIPO DE POSICION"]])
            if tipo_posicion_val and normalize_text(tipo_posicion_val) not in NULL_VALS:
                tipo_posicion_by_key[cargo_key][tipo_posicion_val.upper()] += 1

    return {
        "area_keys": area_keys,
        "all_cargo_keys": all_cargo_keys,
        "countable_by_key": countable_by_key,
        "tap_by_key": tap_by_key,
        "tipo_posicion_by_key": tipo_posicion_by_key,
        "rows_by_structure": dict(rows_by_structure),
    }


def serialize_current_catalog():
    return {
        "timestamp_utc": datetime.utcnow().isoformat(),
        "cargos": list(
            cargo.objects.select_related("area__estructura").values(
                "id",
                "descripcion",
                "area_id",
                "area__descripcion",
                "area__estructura__descripcion",
                "activo",
                "cantidad_aprobada",
                "criticidad",
                "tipo_posicion",
                "tap",
            )
        ),
    }


def ensure_backup_dir(path):
    Path(path).mkdir(parents=True, exist_ok=True)


def write_backup(backup_dir, payload):
    ensure_backup_dir(backup_dir)
    backup_path = Path(backup_dir) / f"hr_cargos_backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
    backup_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str))
    return str(backup_path)


def key_to_dict(key):
    estructura_desc, area_desc, cargo_desc = key
    return {
        "estructura": estructura_desc,
        "area": area_desc,
        "cargo": cargo_desc,
    }


def cargo_to_key(cargo_obj):
    return (
        cargo_obj.area.estructura.descripcion,
        cargo_obj.area.descripcion,
        cargo_obj.descripcion,
    )


def plan_sync(snapshot):
    db_areas = {
        (obj.estructura.descripcion, obj.descripcion): obj
        for obj in area.objects.select_related("estructura").all()
    }
    db_cargos = {
        cargo_to_key(obj): obj
        for obj in cargo.objects.select_related("area__estructura").all()
    }
    used_counts = {
        row["id"]: row["used_contracts"]
        for row in cargo.objects.annotate(
            used_contracts=Count("contratos_personal", distinct=True)
        ).values("id", "used_contracts")
    }

    missing_areas = sorted(snapshot["area_keys"] - set(db_areas.keys()))
    to_create = []
    to_update = []
    to_delete = []
    kept_used_outside_excel = []

    excel_keys = snapshot["all_cargo_keys"]
    db_keys = set(db_cargos.keys())

    for key in sorted(excel_keys - db_keys):
        area_key = (key[0], key[1])
        area_obj = db_areas.get(area_key)
        if area_obj is None:
            continue
        to_create.append({
            **key_to_dict(key),
            "cantidad_aprobada": snapshot["countable_by_key"].get(key, 0),
            "tap": most_common_value(snapshot["tap_by_key"].get(key)),
            "tipo_posicion": most_common_value(snapshot["tipo_posicion_by_key"].get(key)) or "HC",
            "area_id": area_obj.id,
        })

    for key in sorted(excel_keys & db_keys):
        cargo_obj = db_cargos[key]
        new_cantidad = snapshot["countable_by_key"].get(key, 0)
        if cargo_obj.cantidad_aprobada != new_cantidad:
            to_update.append({
                **key_to_dict(key),
                "id": cargo_obj.id,
                "current_cantidad_aprobada": cargo_obj.cantidad_aprobada,
                "new_cantidad_aprobada": new_cantidad,
            })

    for key in sorted(db_keys - excel_keys):
        cargo_obj = db_cargos[key]
        used_contracts = used_counts.get(cargo_obj.id, 0)
        payload = {
            **key_to_dict(key),
            "id": cargo_obj.id,
            "used_contracts": used_contracts,
            "cantidad_aprobada": cargo_obj.cantidad_aprobada,
        }
        if used_contracts > 0:
            kept_used_outside_excel.append(payload)
        else:
            to_delete.append(payload)

    return {
        "summary": {
            "excel_unique_areas": len(snapshot["area_keys"]),
            "excel_unique_cargos": len(snapshot["all_cargo_keys"]),
            "excel_cargos_con_dotacion": len(snapshot["countable_by_key"]),
            "excel_rows_by_structure": snapshot["rows_by_structure"],
            "current_areas": area.objects.count(),
            "current_cargos": cargo.objects.count(),
            "used_cargos": sum(1 for value in used_counts.values() if value > 0),
            "unused_cargos": sum(1 for value in used_counts.values() if value == 0),
            "areas_missing_in_db": len(missing_areas),
            "to_create": len(to_create),
            "to_update": len(to_update),
            "to_delete": len(to_delete),
            "used_outside_excel": len(kept_used_outside_excel),
        },
        "missing_areas": [
            {"estructura": estructura_desc, "area": area_desc}
            for estructura_desc, area_desc in missing_areas[:SAMPLE_LIMIT]
        ],
        "to_create": to_create,
        "to_update": to_update,
        "to_delete": to_delete,
        "kept_used_outside_excel": kept_used_outside_excel,
        "samples": {
            "to_create": to_create[:SAMPLE_LIMIT],
            "to_update": to_update[:SAMPLE_LIMIT],
            "to_delete": to_delete[:SAMPLE_LIMIT],
            "used_outside_excel": kept_used_outside_excel[:SAMPLE_LIMIT],
        },
    }


def apply_sync(sync_plan, backup_dir):
    if sync_plan["summary"]["areas_missing_in_db"] > 0:
        raise ValueError("No se puede aplicar la sincronización: hay áreas del Excel que no existen en la DB.")

    backup_path = write_backup(backup_dir, serialize_current_catalog())
    updated = 0
    created = 0
    deleted = 0
    skipped_protected_deletes = []

    with transaction.atomic():
        for item in sync_plan["to_update"]:
            cargo_obj = cargo.objects.get(pk=item["id"])
            cargo_obj.cantidad_aprobada = item["new_cantidad_aprobada"]
            cargo_obj.save(update_fields=["cantidad_aprobada"])
            updated += 1

        for item in sync_plan["to_create"]:
            cargo.objects.create(
                descripcion=item["cargo"],
                area_id=item["area_id"],
                activo=True,
                cantidad_aprobada=item["cantidad_aprobada"],
                criticidad="MEDIA",
                tipo_posicion=item["tipo_posicion"],
                tap=item["tap"],
            )
            created += 1

        for item in sync_plan["to_delete"]:
            cargo_obj = cargo.objects.get(pk=item["id"])
            try:
                cargo_obj.delete()
                deleted += 1
            except ProtectedError:
                skipped_protected_deletes.append(item)

    return {
        "backup_path": backup_path,
        "updated_cargos": updated,
        "created_cargos": created,
        "deleted_cargos": deleted,
        "skipped_protected_deletes": skipped_protected_deletes,
        "final_cargos": cargo.objects.count(),
    }


def build_output(mode, sync_plan, apply_result=None):
    payload = {
        "mode": mode,
        "summary": sync_plan["summary"],
        "missing_areas": sync_plan["missing_areas"],
        "samples": sync_plan["samples"],
    }
    if apply_result is not None:
        payload["result"] = apply_result
    return payload


def main():
    parser = argparse.ArgumentParser(description="Sincroniza incrementalmente el catálogo de cargos desde Excel.")
    parser.add_argument("--input", default=DEFAULT_INPUT)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--backup-dir", default=DEFAULT_BACKUP_DIR)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    snapshot = build_snapshot(args.input, args.sheet)
    sync_plan = plan_sync(snapshot)

    if not args.apply:
        print(json.dumps(build_output("dry-run", sync_plan), ensure_ascii=False, indent=2, default=str))
        return

    result = apply_sync(sync_plan, args.backup_dir)
    print(json.dumps(build_output("apply", sync_plan, result), ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
