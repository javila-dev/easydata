import argparse
import csv
import json
import os
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'Atlantic.settings')

import django

if not django.apps.registry.apps.ready:
    django.setup()

from openpyxl import load_workbook
from human_resources.models import (
    BONIFICACION_TYPE_CHOICES,
    arl,
    area,
    cargo,
    cceco,
    ccf,
    ciudad,
    contratos_personal,
    empleadores,
    eps,
    estructura,
    fondo_cesantias,
    fondo_pensiones,
    motivos_retiro,
    sede,
    temporales,
)

HEADER_ROW = 1
DATA_SHEET = 'BD PERSONAL'
DEFAULT_INPUT = '/code/data de Personal Easy Data (2).xlsx'
DEFAULT_OUTPUT = '/code/static_media/tmp/hr_snapshot_mapping.json'
DEFAULT_CSV_OUTPUT = '/code/static_media/tmp/hr_snapshot_mapping.csv'

COLUMN_MAP = {
    'empleador_empresa': 'EMPLEADOR',
    'codigo': 'CODIGO',
    'tipo_ingreso': 'TIPO INGRESO',
    'tipo_id': 'TIPO ID',
    'numero_id': 'ID',
    'nombre_completo': 'NOMBRE COMPLETO',
    'cargo': 'CARGO',
    'tap': 'TAP',
    'estado': 'ESTADO',
    'empalme': 'EMPALME',
    'tipo_empalme': 'TIPO DE EMPALME',
    'fecha_limite_empalme': 'F.LIMITE EMPALME',
    'persona_reemplaza': 'PERSONA QUE REEMPLAZA',
    'tipo_posicion': 'TIPO DE POSICION',
    'modalidad_ingreso': 'MODALIDAD DE INGRESO',
    'temporal': 'EMPLEADOR_2',
    'tipo_contrato': 'TIPO DE CONTRATO',
    'estructura': 'ESTRUCTURA',
    'gerencia': 'GERENCIA',
    'area': 'AREA',
    'codigo_centro_costos': 'CODIGO CENTRO DE COSTOS',
    'centro_costos': 'CENTRO DE COSTOS',
    'jefe_inmediato': 'JEFE INMEDIATO',
    'fecha_ingreso': 'F.INGRESO',
    'periodo_prueba': 'PERIODO DE PRUEBA',
    'ubicacion': 'UBICACION',
    'ciudad_laboral': 'CIUDAD',
    'tipo_salario': 'TIPO DE SALARIO',
    'auxilio_transporte': 'AUX TTE',
    'auxilio_movilizacion': 'AUX MOVILIZACION',
    'valor_auxilio': 'AUXILIOS',
    'tipo_auxilio': 'TIPO AUXILIO',
    'tipo_bonificacion_raw': 'TIPO BONIFICACION',
    'base_variable': 'BASE VARIABLE',
    'eps': 'EPS',
    'afp': 'AFP',
    'cesantias': 'CESANTIAS',
    'arl': 'ARL',
    'riesgo': '% RIESGO',
    'ccf': 'CCF',
    'email': 'EMAIL',
    'direccion': 'DIR.RESIDENCIA',
    'ciudad_residencia': 'CIUDAD RESIDENCIA',
    'celular': 'CELULAR',
    'otro_contacto': 'OTRO CONTACTO',
    'vivienda': '¿VIVIENDA PROPIA?',
    'nivel_educativo': 'NIVEL EDUCATIVO',
    'profesion': 'PROFESION/CARRERA',
    'sexo': 'SEXO',
    'rh': 'RH',
    'estado_civil': 'ESTADO CIVIL',
    'fecha_nacimiento': 'F NACIM',
    'talla_camisa': 'TALLA CAMISA',
    'talla_pantalon': 'TALLA PANTALON',
    'talla_calzado': 'TALLA CALZADO',
    'fecha_retiro': 'FECHA RETIRO',
    'motivo_retiro': 'MOTIVO DEL RETIRO',
    'motivo_retiro_real': 'MOTIVO REAL TERMINACION(CONFIDENCIAL)',
    'salario_base': 'SALARIO',
}

TIPO_INGRESO_MAP = {
    'T': 'T',
    'TEMPORAL': 'T',
    'V': 'V',
    'DIRECTO': 'V',
    'A': 'A',
    'APRENDIZAJE': 'A',
    'APRENDIZ': 'A',
}

MODALIDAD_INGRESO_MAP = {
    'NUEVO INGRESO POR TEMPORAL': 'NUEVO INGRESO POR TEMPORAL',
    'NUEVO INGRESO DIRECTO': 'NUEVO INGRESO DIRECTO',
    'NUEVO INGRESO DIRECTO (TEMPORAL A DIRECTO)': 'NUEVO INGRESO DIRECTO (TEMPORAL A DIRECTO)',
    'NUEVO INGRESO DIRECTO TEMPORAL A DIRECTO': 'NUEVO INGRESO DIRECTO (TEMPORAL A DIRECTO)',
    'CAMBIO TEMPORAL A DIRECTO': 'NUEVO INGRESO DIRECTO (TEMPORAL A DIRECTO)',
    'REINGRESO': 'REINGRESO',
}

TIPO_CONTRATO_MAP = {
    'OBRA O LABOR': 'Obra o Labor',
    'OBRA LABOR': 'Obra o Labor',
    'INDEFINIDO': 'Indefinido',
    'FIJO': 'Indefinido',
    'APRENDIZAJE': 'Aprendizaje',
}

TIPO_POSICION_MAP = {
    'HC': 'HC',
    'HC PLUS': 'HC PLUS',
    'ESTRATEGIA ROTACION': 'ESTRATEGIA ROTACION',
    'APRENDIZ': 'APRENDIZ',
    'NO PLUS': 'NO PLUS',
    'PTE RECUPERAR': 'PTE RECUPERAR',
    'MATERNIDAD': 'MATERNIDAD',
}

TIPO_VIVIENDA_MAP = {
    'SI': 'Propia',
    'NO': 'Familiar',
    'PROPIA': 'Propia',
    'FAMILIAR': 'Familiar',
    'ARRENDADA': 'Arrendada',
}

SEXO_MAP = {'M': 'M', 'F': 'F', 'O': 'O'}

BONIFICACION_TYPE_ALIASES = {
    'RENDIMIENTO PLANTA': 'BONIFICACION POR RENDIMIENTO DE PLANTA',
    'BONIFICACION RENDIMIENTO DE PLANTA': 'BONIFICACION POR RENDIMIENTO DE PLANTA',
    'BONFICACION POR RENDIMIENTO DE PLANTA': 'BONIFICACION POR RENDIMIENTO DE PLANTA',
    'COMISIONES COMERCIALES': 'COMISION COMERCIALES',
    'COMSION': 'COMISION',
    '-': None,
}

ESTADO_CIVIL_MAP = {
    'SOLTERO(A)': 'SOLTERO (A)',
    'SOLTERO (A)': 'SOLTERO (A)',
    'CASADO(A)': 'CASADO (A)',
    'CASADO (A)': 'CASADO (A)',
    'VIUDO(A)': 'VIUDO (A)',
    'VIUDO (A)': 'VIUDO (A)',
    'DIVORCIADO(A)': 'DIVORCIADO (A)',
    'DIVORCIADO (A)': 'DIVORCIADO (A)',
    'SEPARADO (A)': 'DIVORCIADO (A)',
    'UNION LIBRE': 'UNION LIBRE',
}

NIVEL_EDUCATIVO_MAP = {
    'NINGUNO': None,
    'BACHILLER': 'BACHILLER',
    'BACHILLER ACADEMICO': 'BACHILLER',
    'TECNICO': 'TECNICO',
    'TECNICO LABORAL': 'TECNICO',
    'TECNOLOGO': 'TECNOLOGO',
    'PROFESIONAL': 'UNIVERSITARIO/PROFESIONAL',
    'UNIVERSITARIO': 'UNIVERSITARIO/PROFESIONAL',
    'ADMINISTRADOR': 'UNIVERSITARIO/PROFESIONAL',
    'INGENIERO': 'INGENIERO',
    'ESPECIALISTA': 'ESPECIALISTA',
    'MAGISTER': 'MAGISTER',
    'DOCTORADO': 'DOCTORADO',
}

RIESGO_MAP = {
    '0': '0',
    '0.522': '0.522',
    '0,522': '0.522',
    '1.044': '1.044',
    '1.04': '1.044',
    '1044': '1.044',
    '4.35': '4.35',
    '4,35': '4.35',
    '4.350': '4.35',
    '2436': 'OTRO',
    '2.436': 'OTRO',
    'NO APLICA': 'NO APLICA',
    'OTRO': 'OTRO',
}

ESTADO_ACTIVO_MAP = {
    'ACTIVO': True,
    'INACTIVO': False,
}

ESTRUCTURA_ALIASES = {
    'OPERATIVO': 'OPERACIONES',
    'OPERACIONES': 'OPERACIONES',
    'COMERCIAL': 'COMERCIAL',
    'ABASTECIMIENTO': 'ABASTECIMIENTO',
    'ADMON': 'ADMIN Y FRA',
    'ADMIN Y FRA': 'ADMIN Y FRA',
    'GENERAL': 'GENERAL',
}

AREA_ALIASES = {}

SEDE_ALIASES = {
    'OFICINA BOGOTA': 'BOGOTA',
    'OFICINA CALI': 'CALI',
    'OFICINA CENTRAL': 'OFICINA CENTRAL',
    'VTAS BOGOTA': 'VTAS BOGOTA',
    'VTAS MEDELLIN': 'VTAS MEDELLIN',
    'VTAS CARTAGENA': 'CARTAGENA',
    'PROCESOS VA S.A.S.': 'OFICINA CENTRAL',
    'OFICINA SABANETA': 'CEDI MEDELLIN SABANETA',
    'CONSUMIDOR FINAL CEDI': 'CONSUMIDOR FINAL OF CENTRAL',
    'CONSUMIDOR FINAL CEDI ': 'CONSUMIDOR FINAL OF CENTRAL',
    'OFICINA CUCUTA': 'VTAS CUCUTA',
    'MERCADEO CEDI': 'MERCADEO OF CENTRAL',
    'B2B CUCUTA': 'B2B BUCARAMANGA',
    'RETAIL POBLADO': 'RETAIL OF CENTRAL',
    'B2C POBLADO': 'RETAIL OF CENTRAL',
    'PVA CEDI CARTAGENA': 'CEDI CARTAGENA',
    'COTA': 'OFICINA COTA',
    'RETAIL BOGOTA': 'BOGOTA',
    'RETAIL BARRANQUILLA': 'BARRANQUILLA',
    'RETAIL CALI': 'CALI',
    'REATIL CEDI': 'RETAIL CEDI',
    'OFICINA BARANQUILLA': 'OFICINA BARRANQUILLA',
    'B2C BOGOTA': 'VTAS BOGOTA',
}

CCF_ALIASES = {
    'COMFENALCO ANTIOQUIA': 'COMFAMA',
    'COMFAMILIAR ATLANTICO': 'COMFAMILIAR DEL ATLANTICO SA',
    'COMFAMILIAR DEL ATLANTCIO SAS': 'COMFAMILIAR DEL ATLANTICO SA',
    'COMFEN': 'COMFENALCO CARTAGENA',
    'COFMAMA': 'COMFAMA',
}

# CCF por defecto según sede, para cuando el campo CCF está vacío en el Excel.
# Clave: normalize_text del valor de UBICACION en el Excel.
SEDE_CCF_DEFAULT = {
    'OFICINA CENTRAL':          'COMFAMA',
    'PVA MEDELLIN':             'COMFAMA',
    'RETAIL POBLADO':           'COMFAMA',
    'OFICINA CALDAS':           'COMFAMA',
    'CEDI MEDELLIN SABANETA':   'COMFAMA',
    'OFICINA CALI':             'COMFANDI',
    'CEDI CALI':                'COMFANDI',
    'RETAIL CALI':              'COMFANDI',
    'OFICINA BOGOTA':           'COMPENSAR',
    'OFICINA COTA':             'COMPENSAR',
    'RETAIL BOGOTA':            'COMPENSAR',
    'VTAS BOGOTA':              'COMPENSAR',
    'OFICINA MANIZALES':        'COFREM',
    'OFICINA IBAGUE':           'COMFATOLIMA',
    'OFICINA MONTERIA':         'COMFACOR',
    'OFICINA ARMENIA':          'COMFANALCO QUINDIO',
    'OFICINA PEREIRA':          'COMFAMILIAR RISARALDA',
    'OFICINA TUNJA':            'COMFABOY',
    'OFICINA BARRANQUILLA':     'COMFAMILIAR DEL ATLANTICO SA',
    'VTAS BARRANQUILLA':        'COMFAMILIAR DEL ATLANTICO SA',
    'OFICINA CARTAGENA':        'COMFENALCO CARTAGENA',
    'CEDI CARTAGENA':           'COMFENALCO CARTAGENA',
    'OFICINA BUCARAMANGA':      'COMFENALCO SANTANDER',
    'VTAS BUCARAMANGA':         'COMFENALCO SANTANDER',
    'B2B BUCARAMANGA':          'COMFENALCO SANTANDER',
    'OFICINA IBAGUE':           'COMFATOLIMA',
    'OFICINA MANIZALES':        'COFREM',
    'OFICINA ARMENIA':          'COMFANALCO QUINDIO',
    'OFICINA TUNJA':            'COMFABOY',
    'OFICINA MONTERIA':         'COMFACOR',
    'OFICINA SAN GIL':          'COMFENALCO SANTANDER',
    'OFICINA CNCH':             'COMPENSAR',
    'B2C BOGOTA':               'COMPENSAR',
}

ARL_ALIASES = {
    'ARL POSITIVA': 'ARL POSITIVA',
    'AR.L SURA': 'ARL SURA',
    'NO APLICA': 'OTRA',
}

GLOBAL_AREA_ALIASES = {}

EPS_ALIASES = {
    'EPS S.O.S SERVICIO OCCIDENTAL': 'EPS SOS SERVICIO OCCIDENTAL',
    'EPS SOS SERVICIOS OCCIDENTE': 'EPS SOS SERVICIO OCCIDENTAL',
}

NULL_TEMPORAL_VALUES = {
    '',
    'NONE',
    'NO REGISTRA',
    'N/A',
    'NA',
    'DIRECTA',
    'VACANTE',
}

CITY_ALIASES = {
    'BOGOTA': 'Bogotá',
    'MEDELLIN': 'Medellín',
    'CARTAGENA': 'Cartagena',
    'CALI': 'Cali',
    'BARRANQUILLA': 'Barranquilla',
    'BUCARAMANGA': 'Bucaramanga',
    'PEREIRA': 'Pereira',
    'CUCUTA': 'Cúcuta',
    'DOS QUEBRADAS': 'Dosquebradas',
    'SANTIAGO DE CALI': 'Cali',
    'MALALMBO': 'Malambo',
    'SUBA': 'Bogotá',
    'BOSA': 'Bogotá',
    'NIQUIA': 'Bello',
    'SAN ANTONIO DE PRADO': 'Medellín',
    'EL RETIRO': 'Retiro',
    'QUINDIO': 'Armenia',
}

# Cuando el nombre de ciudad es ambiguo (existe en varios dptos),
# indica qué departamento preferir. Clave: normalize_text del valor raw.
CITY_DEPT_HINTS = {
    'QUINDIO': 'Quindio',
}


def strip_accents(value):
    if value is None:
        return ''
    value = str(value)
    return ''.join(ch for ch in unicodedata.normalize('NFKD', value) if not unicodedata.combining(ch))


def normalize_text(value):
    if value is None:
        return ''
    value = strip_accents(value)
    value = value.replace('\t', ' ')
    value = value.replace('\n', ' ')
    value = re.sub(r'\s+', ' ', value).strip()
    return value.upper()


def normalize_choice(value, mapping):
    key = normalize_text(value)
    return mapping.get(key)


BONIFICACION_TYPE_MAP = {
    normalize_text(value): value for value, _label in BONIFICACION_TYPE_CHOICES
}


def normalize_bonificacion_type(value):
    key = normalize_text(value)
    if not key:
        return None
    candidate = BONIFICACION_TYPE_ALIASES.get(key, value)
    if candidate is None:
        return None
    return BONIFICACION_TYPE_MAP.get(normalize_text(candidate))


def strip_leading_code(value):
    cleaned = normalize_text(value)
    if not cleaned:
        return ''
    return re.sub(r'^[A-Z]+\d+\s+', '', cleaned).strip()


def clean_optional(value):
    cleaned = normalize_text(value)
    if cleaned in {'', 'NONE', 'NO REGISTRA', 'N/A', 'NA', '#N/A'}:
        return None
    return str(value).strip() if isinstance(value, str) else value


def parse_numeric(value):
    if value in (None, '', 'None'):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace('$', '').replace(' ', '')
    if not cleaned:
        return None
    if ',' in cleaned and '.' in cleaned:
        cleaned = cleaned.replace('.', '').replace(',', '.')
    else:
        cleaned = cleaned.replace(',', '.')
    try:
        return float(cleaned)
    except (TypeError, ValueError):
        return None


def split_full_name(full_name):
    cleaned = normalize_text(full_name)
    if not cleaned:
        return {
            'primer_nombre': None,
            'segundo_nombre': None,
            'primer_apellido': None,
            'segundo_apellido': None,
            'issues': ['nombre_completo_vacio'],
        }

    parts = cleaned.split(' ')
    issues = []
    if len(parts) == 2:
        primer_nombre, primer_apellido = parts
        segundo_nombre = None
        segundo_apellido = None
    elif len(parts) == 3:
        common_first_names = {'JOSE','JUAN','LUIS','MARIA','ANA','ANDRES','JULIO','JHON','CARLOS','DAVID'}
        if parts[1] in common_first_names:
            primer_apellido = parts[0]
            segundo_apellido = None
            primer_nombre = parts[1]
            segundo_nombre = parts[2]
        else:
            primer_apellido = parts[0]
            segundo_apellido = parts[1]
            primer_nombre = parts[2]
            segundo_nombre = None
    else:
        primer_apellido = parts[0]
        segundo_apellido = parts[1] if len(parts) > 1 else None
        first_names = parts[2:]
        primer_nombre = first_names[0] if first_names else None
        segundo_nombre = ' '.join(first_names[1:]) if len(first_names) > 1 else None
    return {
        'primer_nombre': primer_nombre,
        'segundo_nombre': segundo_nombre,
        'primer_apellido': primer_apellido,
        'segundo_apellido': segundo_apellido,
        'issues': issues,
    }


class CatalogLookup:
    def __init__(self, queryset, attr):
        self.by_normalized = {}
        self.attr = attr
        for obj in queryset:
            key = normalize_text(getattr(obj, attr))
            self.by_normalized[key] = obj

    def get(self, value):
        if value is None:
            return None
        key = normalize_text(value)
        return self.by_normalized.get(key)

    def get_with_alias(self, value, aliases=None):
        if value is None:
            return None
        candidate = aliases.get(normalize_text(value), value) if aliases else value
        return self.get(candidate)


class CecoLookup:
    def __init__(self):
        self.by_code = {}
        self.by_desc = {}
        for obj in cceco.objects.all():
            if obj.codigo_sap:
                self.by_code[normalize_text(obj.codigo_sap)] = obj
            self.by_desc[normalize_text(obj.descripcion)] = obj

    def get(self, code, description):
        if code:
            obj = self.by_code.get(normalize_text(code))
            if obj:
                return obj
        if description:
            return self.by_desc.get(normalize_text(description))
        return None


class AreaLookup:
    def __init__(self):
        self.by_pair = {}
        self.by_name = defaultdict(list)
        for obj in area.objects.select_related('estructura').all():
            key = (normalize_text(obj.estructura.descripcion), normalize_text(obj.descripcion))
            self.by_pair[key] = obj
            self.by_name[normalize_text(obj.descripcion)].append(obj)

    def get(self, estructura_name, area_name):
        estructura_key = normalize_text(estructura_name)
        area_key = normalize_text(area_name)
        area_candidates = [area_key]
        stripped_area_key = strip_leading_code(area_name)
        if stripped_area_key and stripped_area_key not in area_candidates:
            area_candidates.append(stripped_area_key)

        for candidate_key in area_candidates:
            alias = AREA_ALIASES.get((estructura_key, candidate_key)) or GLOBAL_AREA_ALIASES.get(candidate_key)
            resolved_key = normalize_text(alias) if alias else candidate_key
            match = self.by_pair.get((estructura_key, resolved_key))
            if match:
                return match
            candidates = self.by_name.get(resolved_key, [])
            if len(candidates) == 1:
                return candidates[0]
            for candidate in candidates:
                if normalize_text(candidate.estructura.descripcion) == estructura_key:
                    return candidate
        return None


class CargoLookup:
    def __init__(self):
        self.by_area = defaultdict(dict)
        for obj in cargo.objects.select_related('area').all():
            self.by_area[normalize_text(obj.area.descripcion)][normalize_text(obj.descripcion)] = obj

    def get(self, area_obj, cargo_name):
        if area_obj is None or not cargo_name:
            return None
        cargo_key = normalize_text(cargo_name)
        return self.by_area[normalize_text(area_obj.descripcion)].get(cargo_key)


NULL_CITY_VALUES = {'', 'NONE', 'NO REGISTRA', 'N/A', 'NA', '#N/A', '#NA'}
CITY_NULL_DEFAULT = 'Medellín'

class CityLookup:
    def __init__(self):
        self.by_name = defaultdict(list)
        for obj in ciudad.objects.select_related('departamento').all():
            self.by_name[normalize_text(obj.nombre)].append(obj)
        default_matches = self.by_name.get(normalize_text(CITY_NULL_DEFAULT), [])
        self._default = default_matches[0] if default_matches else None

    def get(self, city_name):
        if city_name is None or normalize_text(city_name) in NULL_CITY_VALUES:
            return self._default
        raw_key = normalize_text(city_name)
        alias = CITY_ALIASES.get(raw_key, city_name)
        matches = self.by_name.get(normalize_text(alias), [])
        if not matches:
            return None
        if len(matches) == 1:
            return matches[0]
        dept_hint = CITY_DEPT_HINTS.get(raw_key)
        if dept_hint:
            hint_key = normalize_text(dept_hint)
            for candidate in matches:
                if normalize_text(candidate.departamento.nombre) == hint_key:
                    return candidate
        return matches[0]


class Mapper:
    def __init__(self):
        self.eps_lookup = CatalogLookup(eps.objects.all(), 'nombre')
        self.pension_lookup = CatalogLookup(fondo_pensiones.objects.all(), 'nombre')
        self.cesantias_lookup = CatalogLookup(fondo_cesantias.objects.all(), 'nombre')
        self.arl_lookup = CatalogLookup(arl.objects.all(), 'nombre')
        self.ccf_lookup = CatalogLookup(ccf.objects.all(), 'nombre')
        self.empleador_lookup = CatalogLookup(empleadores.objects.all(), 'nombre')
        self.temporal_lookup = CatalogLookup(temporales.objects.all(), 'nombre')
        self.estructura_lookup = CatalogLookup(estructura.objects.all(), 'descripcion')
        self.sede_lookup = CatalogLookup(sede.objects.all(), 'descripcion')
        self.city_lookup = CityLookup()
        self.area_lookup = AreaLookup()
        self.cargo_lookup = CargoLookup()
        self.ceco_lookup = CecoLookup()
        self.motivo_lookup = CatalogLookup(motivos_retiro.objects.all(), 'descripcion')

    def map_row(self, raw_row, row_number):
        issues = []
        mapped = {'row_number': row_number}

        name_parts = split_full_name(raw_row.get('nombre_completo'))
        mapped.update({k: v for k, v in name_parts.items() if k != 'issues'})
        issues.extend(name_parts['issues'])

        raw_estado = normalize_text(raw_row.get('estado'))
        raw_nombre = normalize_text(raw_row.get('nombre_completo'))
        is_vacante_row = (
            (raw_estado == 'VACANTE' or 'VACANTE' in raw_nombre)
            and clean_optional(raw_row.get('numero_id')) in (None, 0, '0')
        )
        if is_vacante_row:
            mapped.update({'ignored_row': True})
            issues.append('fila_vacante_ignorada')
            mapped['issues'] = issues
            mapped['can_import'] = False
            mapped['source_snapshot'] = raw_row
            return mapped

        activo = normalize_choice(raw_row.get('estado'), ESTADO_ACTIVO_MAP)
        if activo is None:
            issues.append('estado_no_mapeado')

        tipo_ingreso = normalize_choice(raw_row.get('tipo_ingreso'), TIPO_INGRESO_MAP)
        if tipo_ingreso is None:
            tipo_ingreso = 'V'
        modalidad = normalize_choice(raw_row.get('modalidad_ingreso'), MODALIDAD_INGRESO_MAP)
        if modalidad is None and clean_optional(raw_row.get('modalidad_ingreso')) is None:
            modalidad = 'NUEVO INGRESO DIRECTO'
        tipo_contrato = normalize_choice(raw_row.get('tipo_contrato'), TIPO_CONTRATO_MAP)
        tipo_posicion = normalize_choice(raw_row.get('tipo_posicion'), TIPO_POSICION_MAP)
        tipo_vivienda = normalize_choice(raw_row.get('vivienda'), TIPO_VIVIENDA_MAP)
        sexo = normalize_choice(raw_row.get('sexo'), SEXO_MAP)
        estado_civil = normalize_choice(raw_row.get('estado_civil'), ESTADO_CIVIL_MAP)
        tipo_riesgo = normalize_choice(raw_row.get('riesgo'), RIESGO_MAP)
        if tipo_riesgo is None:
            tipo_riesgo = '0'
        nivel_educativo = normalize_choice(raw_row.get('nivel_educativo'), NIVEL_EDUCATIVO_MAP)
        auxilio_transporte = parse_numeric(raw_row.get('auxilio_transporte'))
        auxilio_movilizacion = parse_numeric(raw_row.get('auxilio_movilizacion'))
        valor_auxilio = parse_numeric(raw_row.get('valor_auxilio'))
        base_bonificacion = parse_numeric(raw_row.get('base_variable'))
        tipo_auxilio_raw = clean_optional(raw_row.get('tipo_auxilio'))
        tipo_auxilio = tipo_auxilio_raw.strip() if isinstance(tipo_auxilio_raw, str) else tipo_auxilio_raw
        bonificacion_tipo = normalize_bonificacion_type(raw_row.get('tipo_bonificacion_raw'))
        tipo_auxilio_para_comparacion = normalize_bonificacion_type(tipo_auxilio_raw) or tipo_auxilio
        mismo_concepto_aux_y_bonificacion = (
            bonificacion_tipo is not None
            and tipo_auxilio_para_comparacion is not None
            and normalize_text(bonificacion_tipo) == normalize_text(tipo_auxilio_para_comparacion)
        )

        bonificacion = None
        if mismo_concepto_aux_y_bonificacion and base_bonificacion not in (None, 0):
            bonificacion = bonificacion_tipo
        elif not mismo_concepto_aux_y_bonificacion:
            base_bonificacion = None

        tipo_id = clean_optional(raw_row.get('tipo_id'))
        if tipo_id is None:
            tipo_id = 'CC'

        mapped.update({
            'numero_identificacion': raw_row.get('numero_id'),
            'tipo_id': tipo_id,
            'codigo_sap': clean_optional(raw_row.get('codigo')),
            'email': clean_optional(raw_row.get('email')),
            'direccion_residencia': clean_optional(raw_row.get('direccion')),
            'contacto': clean_optional(raw_row.get('celular')),
            'contacto_otro': clean_optional(raw_row.get('otro_contacto')),
            'tipo_vivienda': tipo_vivienda,
            'nivel_educativo': nivel_educativo,
            'titulo': clean_optional(raw_row.get('profesion')),
            'sexo': sexo,
            'rh': clean_optional(raw_row.get('rh')),
            'estado_civil': estado_civil,
            'fecha_nacimiento': raw_row.get('fecha_nacimiento'),
            'talla_camisa': clean_optional(raw_row.get('talla_camisa')),
            'talla_pantalon': clean_optional(raw_row.get('talla_pantalon')),
            'talla_calzado': clean_optional(raw_row.get('talla_calzado')),
            'activo': activo,
            'tipo_riesgo': tipo_riesgo,
            'tipo_ingreso': tipo_ingreso,
            'modalidad_ingreso': modalidad,
            'tipo_contrato': tipo_contrato,
            'tipo_posicion': tipo_posicion,
            'auxilio_transporte': auxilio_transporte,
            'base_bonificacion': base_bonificacion,
            'bonificacion': bonificacion,
            'tipo_bonificacion_raw': bonificacion_tipo,
            'fecha_inicio': raw_row.get('fecha_ingreso'),
            'fecha_periodo_prueba': raw_row.get('periodo_prueba'),
            'fecha_retiro': raw_row.get('fecha_retiro'),
            'motivo_retiro_real': clean_optional(raw_row.get('motivo_retiro_real')),
        })

        auxilios = []
        if auxilio_movilizacion not in (None, 0):
            auxilios.append({
                'descripcion': 'AUXILIO DE MOVILIZACION',
                'valor': auxilio_movilizacion,
                'source': 'auxilio_movilizacion',
            })
        if valor_auxilio not in (None, 0):
            suprimir_auxilio_por_bonificacion = (
                mismo_concepto_aux_y_bonificacion and base_bonificacion not in (None, 0)
            )
            if tipo_auxilio is None and not suprimir_auxilio_por_bonificacion:
                issues.append('auxilio_sin_tipo')
            elif not suprimir_auxilio_por_bonificacion:
                auxilios.append({
                    'descripcion': tipo_auxilio,
                    'valor': valor_auxilio,
                    'source': 'auxilios',
                })
        elif clean_optional(raw_row.get('tipo_auxilio')) is not None:
            issues.append('tipo_auxilio_sin_valor')
        mapped['auxilios'] = auxilios

        for field_name, field_value in [
            ('tipo_id', mapped['tipo_id']),
            ('sexo', mapped['sexo']),
            ('tipo_ingreso', mapped['tipo_ingreso']),
            ('modalidad_ingreso', mapped['modalidad_ingreso']),
            ('tipo_contrato', mapped['tipo_contrato']),
            ('tipo_riesgo', mapped['tipo_riesgo']),
        ]:
            if field_value in (None, ''):
                issues.append(f'{field_name}_no_mapeado')

        raw_temporal = raw_row.get('temporal')
        raw_temporal_key = normalize_text(raw_temporal)
        temporal_source = None if raw_temporal_key in NULL_TEMPORAL_VALUES else raw_temporal

        eps_obj = self.eps_lookup.get_with_alias(raw_row.get('eps'), EPS_ALIASES)
        pension_obj = self.pension_lookup.get(raw_row.get('afp'))
        cesantias_obj = self.cesantias_lookup.get(raw_row.get('cesantias'))
        arl_obj = self.arl_lookup.get_with_alias(raw_row.get('arl'), ARL_ALIASES)
        ccf_obj = self.ccf_lookup.get_with_alias(raw_row.get('ccf'), CCF_ALIASES)
        if ccf_obj is None and clean_optional(raw_row.get('ccf')) is None:
            sede_key = normalize_text(raw_row.get('ubicacion'))
            ccf_default_name = SEDE_CCF_DEFAULT.get(sede_key)
            if ccf_default_name:
                ccf_obj = self.ccf_lookup.get(ccf_default_name)
        empleador_obj = self.empleador_lookup.get(raw_row.get('empleador_empresa'))
        temporal_obj = self.temporal_lookup.get(temporal_source)
        estructura_name = ESTRUCTURA_ALIASES.get(normalize_text(raw_row.get('estructura')), raw_row.get('estructura'))
        estructura_obj = self.estructura_lookup.get(estructura_name)
        area_obj = self.area_lookup.get(estructura_name, raw_row.get('area'))
        cargo_obj = self.cargo_lookup.get(area_obj, raw_row.get('cargo'))
        sede_obj = self.sede_lookup.get_with_alias(raw_row.get('ubicacion'), SEDE_ALIASES)
        ceco_obj = self.ceco_lookup.get(raw_row.get('codigo_centro_costos'), raw_row.get('centro_costos'))
        ciudad_obj = self.city_lookup.get(raw_row.get('ciudad_residencia'))
        ciudad_laboral_obj = self.city_lookup.get(raw_row.get('ciudad_laboral'))
        motivo_obj = self.motivo_lookup.get(raw_row.get('motivo_retiro'))

        fk_pairs = {
            'eps_id': eps_obj.id if eps_obj else None,
            'pension_id': pension_obj.id if pension_obj else None,
            'cesantias_id': cesantias_obj.id if cesantias_obj else None,
            'arl_id': arl_obj.id if arl_obj else None,
            'ccf_id': ccf_obj.id if ccf_obj else None,
            'empleador_id': empleador_obj.id if empleador_obj else None,
            'temporal_id': temporal_obj.id if temporal_obj else None,
            'estructura_id': estructura_obj.id if estructura_obj else None,
            'area_id': area_obj.id if area_obj else None,
            'cargo_id': cargo_obj.id if cargo_obj else None,
            'sede_id': sede_obj.id if sede_obj else None,
            'cceco_id': ceco_obj.id if ceco_obj else None,
            'ciudad_id': ciudad_obj.id if ciudad_obj else None,
            'departamento_id': ciudad_obj.departamento_id if ciudad_obj else None,
            'ciudad_laboral_id': ciudad_laboral_obj.id if ciudad_laboral_obj else None,
            'motivo_retiro_id': motivo_obj.id if motivo_obj else None,
        }
        mapped.update(fk_pairs)

        issue_checks = [
            ('eps_sin_match', 'eps_id', raw_row.get('eps')),
            ('afp_sin_match', 'pension_id', raw_row.get('afp')),
            ('cesantias_sin_match', 'cesantias_id', raw_row.get('cesantias')),
            ('arl_sin_match', 'arl_id', raw_row.get('arl')),
            ('ccf_sin_match', 'ccf_id', raw_row.get('ccf')),
            ('empleador_sin_match', 'empleador_id', raw_row.get('empleador_empresa')),
            ('temporal_sin_match', 'temporal_id', temporal_source),
            ('estructura_sin_match', 'estructura_id', raw_row.get('estructura')),
            ('area_sin_match', 'area_id', raw_row.get('area')),
            ('cargo_sin_match', 'cargo_id', raw_row.get('cargo')),
            ('ubicacion_sin_match', 'sede_id', raw_row.get('ubicacion')),
            ('ciudad_residencia_sin_match', 'ciudad_id', raw_row.get('ciudad_residencia')),
            ('ciudad_laboral_sin_match', 'ciudad_laboral_id', raw_row.get('ciudad_laboral')),
            ('motivo_retiro_sin_match', 'motivo_retiro_id', raw_row.get('motivo_retiro')),
        ]
        for issue_name, fk_key, raw_value in issue_checks:
            if mapped.get(fk_key) is None and clean_optional(raw_value) is not None:
                issues.append(issue_name)

        required_checks = [
            ('primer_nombre_requerido', mapped.get('primer_nombre')),
            ('primer_apellido_requerido', mapped.get('primer_apellido')),
            ('numero_identificacion_requerido', mapped.get('numero_identificacion')),
            ('tipo_id_requerido', mapped.get('tipo_id')),
            ('sexo_requerido', mapped.get('sexo')),
            ('ciudad_id_requerido', mapped.get('ciudad_id')),
            ('departamento_id_requerido', mapped.get('departamento_id')),
            ('arl_id_requerido', mapped.get('arl_id')),
            ('ccf_id_requerido', mapped.get('ccf_id')),
            ('tipo_riesgo_requerido', mapped.get('tipo_riesgo')),
            ('empleador_id_requerido', mapped.get('empleador_id')),
            ('estructura_id_requerido', mapped.get('estructura_id')),
            ('area_id_requerido', mapped.get('area_id')),
            ('cargo_id_requerido', mapped.get('cargo_id')),
            ('sede_id_requerido', mapped.get('sede_id')),
            ('tipo_ingreso_requerido', mapped.get('tipo_ingreso')),
            ('modalidad_ingreso_requerido', mapped.get('modalidad_ingreso')),
            ('tipo_contrato_requerido', mapped.get('tipo_contrato')),
            ('fecha_inicio_requerida', mapped.get('fecha_inicio')),
        ]
        for issue_name, field_value in required_checks:
            if field_value in (None, ''):
                issues.append(issue_name)

        if mapped.get('activo') is False:
            if mapped.get('fecha_retiro') in (None, ''):
                issues.append('fecha_retiro_requerida_inactivo')
            if clean_optional(raw_row.get('motivo_retiro')) is not None and mapped.get('motivo_retiro_id') is None:
                issues.append('motivo_retiro_requerido_inactivo')

        raw_salario = raw_row.get('salario_base')
        try:
            mapped['salario_base'] = int(float(raw_salario)) if raw_salario not in (None, '', 0) else 0
        except (ValueError, TypeError):
            mapped['salario_base'] = 0

        mapped['source_snapshot'] = {
            key: (value.isoformat() if hasattr(value, 'isoformat') else value)
            for key, value in raw_row.items()
        }
        mapped['issues'] = sorted(set(issues))
        blocking_prefixes = ('_sin_match', '_no_mapeado', '_requerido', '_requerida')
        mapped['can_import'] = not any(issue.endswith(blocking_prefixes) for issue in mapped['issues'])
        return mapped


def load_rows(path, limit=None):
    wb = load_workbook(path, read_only=True, data_only=True)
    if DATA_SHEET not in wb.sheetnames:
        raise ValueError(f'La hoja requerida {DATA_SHEET!r} no existe en {path}')

    ws = wb[DATA_SHEET]
    row_iter = ws.iter_rows(min_row=1, values_only=True)
    header_values = list(next(row_iter))
    header_values[15] = 'EMPLEADOR_2'
    normalized_headers = [normalize_text(h) for h in header_values]
    index_by_header = {header: idx for idx, header in enumerate(normalized_headers)}

    rows = []
    for offset, values in enumerate(row_iter, start=2):
        if limit is not None and offset > HEADER_ROW + limit:
            break
        values = list(values)
        if not any(value not in (None, '') for value in values):
            continue
        raw = {}
        for target, source_header in COLUMN_MAP.items():
            source_index = index_by_header.get(normalize_text(source_header))
            raw[target] = values[source_index] if source_index is not None else None
        rows.append((offset, raw))
    return rows


def write_outputs(records, json_output, csv_output):
    json_path = Path(json_output)
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(records, ensure_ascii=False, indent=2, default=str))

    csv_path = Path(csv_output)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        'row_number', 'numero_identificacion', 'primer_nombre', 'segundo_nombre',
        'primer_apellido', 'segundo_apellido', 'activo', 'tipo_ingreso',
        'modalidad_ingreso', 'tipo_contrato', 'tipo_posicion', 'empleador_id', 'temporal_id',
        'estructura_id', 'area_id', 'cargo_id', 'sede_id', 'cceco_id',
        'ciudad_id', 'departamento_id', 'ciudad_laboral_id', 'eps_id', 'pension_id', 'cesantias_id',
        'arl_id', 'ccf_id', 'motivo_retiro_id', 'can_import', 'issues'
    ]
    with csv_path.open('w', newline='', encoding='utf-8') as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            row = {key: record.get(key) for key in fieldnames}
            row['issues'] = '|'.join(record.get('issues', []))
            writer.writerow(row)


def main():
    parser = argparse.ArgumentParser(description='Mapea el snapshot de personal Excel a FKs/choices de la BD.')
    parser.add_argument('--input', default=DEFAULT_INPUT)
    parser.add_argument('--output', default=DEFAULT_OUTPUT)
    parser.add_argument('--csv-output', default=DEFAULT_CSV_OUTPUT)
    parser.add_argument('--limit', type=int, default=None)
    args = parser.parse_args()

    mapper = Mapper()
    source_rows = load_rows(args.input, limit=args.limit)
    records = [mapper.map_row(raw_row, row_number) for row_number, raw_row in source_rows]
    write_outputs(records, args.output, args.csv_output)

    total = len(records)
    importables = sum(1 for record in records if record['can_import'])
    with_issues = total - importables
    print(f'Filas procesadas: {total}')
    print(f'Filas importables sin faltantes de mapeo: {importables}')
    print(f'Filas con incidencias: {with_issues}')
    print(f'JSON: {args.output}')
    print(f'CSV: {args.csv_output}')

    issue_counts = defaultdict(int)
    for record in records:
        for issue in record['issues']:
            issue_counts[issue] += 1
    if issue_counts:
        print('Incidencias mas frecuentes:')
        for issue, count in sorted(issue_counts.items(), key=lambda item: (-item[1], item[0]))[:20]:
            print(f'  - {issue}: {count}')


if __name__ == '__main__':
    main()
