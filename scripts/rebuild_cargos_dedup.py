"""
Orquesta reset de cargos/contratos/personal y reconstrucción del catálogo
de cargos con deduplicación por acentos/mayúsculas/espacios.

Pasos:
  1. (--reset) Borra personal, contratos y cargos de la BD (mantiene áreas/estructuras)
  2. Lee BD PERSONAL del Excel (todos los estados: activos, inactivos, vacantes)
  3. Deduplica cargos que difieren solo en tildes u otras diferencias menores
  4. Crea cargos con cantidad_aprobada = filas ACTIVO + VACANTE por cargo

Uso:
  python3 scripts/rebuild_cargos_dedup.py              # dry-run (solo muestra plan)
  python3 scripts/rebuild_cargos_dedup.py --apply      # aplica creación de cargos
  python3 scripts/rebuild_cargos_dedup.py --reset --apply  # reset + creación
"""
import argparse
import json
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "Atlantic.settings")

import django
django.setup()

from django.db import transaction
from openpyxl import load_workbook

from human_resources.models import (
    area,
    auxilios_contrato,
    base_personal,
    cambios_salario,
    cargo,
    contratos_personal,
    descargos,
    empalmes,
    historico_base_teorica,
    importacion_personal_job,
)

DEFAULT_INPUT = "/code/Data personal activo.xlsx"
DEFAULT_SHEET = "BD PERSONAL"
DEFAULT_BACKUP_DIR = "/code/static_media/tmp"

# Mapeo nombre columna ESTRUCTURA del Excel → descripcion de estructura en BD
STRUCTURE_MAP = {
    "OPERATIVO": "OPERACIONES",
    "ADMON": "ADMIN Y FRA",
    "COMERCIAL": "COMERCIAL",
    "GENERAL": "GENERAL",
    "ABASTECIMIENTO": "ABASTECIMIENTO",
}

NULL_VALS = {"", "NONE", "N/A", "NA", "#N/A"}


# ---------------------------------------------------------------------------
# Helpers de texto
# ---------------------------------------------------------------------------

def clean_text(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def normalize_header(value):
    if value is None:
        return ""
    value = "".join(
        ch for ch in unicodedata.normalize("NFKD", str(value))
        if not unicodedata.combining(ch)
    )
    return re.sub(r"\s+", " ", value).strip().upper()


def normalize_text(value):
    """Quita tildes, pasa a mayúsculas, colapsa espacios — solo para comparar."""
    if value is None:
        return ""
    value = "".join(
        ch for ch in unicodedata.normalize("NFKD", str(value))
        if not unicodedata.combining(ch)
    )
    return re.sub(r"\s+", " ", value).strip().upper()


def most_common_value(counter):
    if not counter:
        return None
    return counter.most_common(1)[0][0]


def is_countable_row(raw_estado, raw_nombre):
    """Filas que suman a la dotación ideal: estado ACTIVO o cualquier VACANTE."""
    estado = normalize_text(raw_estado)
    nombre = normalize_text(raw_nombre)
    return estado == "ACTIVO" or "VACANTE" in estado or "VACANTE" in nombre


# ---------------------------------------------------------------------------
# Paso 1 – Reset de datos
# ---------------------------------------------------------------------------

def collect_counts():
    return {
        "cargos": cargo.objects.count(),
        "base_personal": base_personal.objects.count(),
        "contratos_personal": contratos_personal.objects.count(),
        "auxilios_contrato": auxilios_contrato.objects.count(),
        "cambios_salario": cambios_salario.objects.count(),
        "empalmes": empalmes.objects.count(),
        "descargos": descargos.objects.count(),
        "historico_base_teorica": historico_base_teorica.objects.count(),
        "importacion_personal_job": importacion_personal_job.objects.count(),
    }


def reset_hr_data():
    before = collect_counts()
    with transaction.atomic():
        # Limpia FK circular antes de borrar base_personal
        area.objects.exclude(responsable__isnull=True).update(responsable=None)
        auxilios_contrato.objects.all().delete()
        cambios_salario.objects.all().delete()
        empalmes.objects.all().delete()
        descargos.objects.all().delete()
        contratos_personal.objects.all().delete()
        base_personal.objects.all().delete()
        historico_base_teorica.objects.all().delete()
        cargo.objects.all().delete()
        importacion_personal_job.objects.all().delete()
    after = collect_counts()
    return {"before": before, "after": after}


# ---------------------------------------------------------------------------
# Paso 2 – Lectura y normalización del Excel
# ---------------------------------------------------------------------------

def load_excel_snapshot(path, sheet_name):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    raw_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {normalize_header(h): pos for pos, h in enumerate(raw_headers)}

    required = {"ESTRUCTURA", "AREA", "CARGO"}
    missing = sorted(required - set(idx.keys()))
    if missing:
        raise ValueError(f"Headers faltantes en el Excel: {', '.join(missing)}")

    # norm_key = (mapped_structure, norm_area, norm_cargo)
    # Agrupamos los nombres RAW de cargo por clave normalizada para elegir canónico
    raw_names_by_norm: dict[tuple, Counter] = defaultdict(Counter)
    countable_by_norm: Counter = Counter()
    tap_by_norm: dict[tuple, Counter] = defaultdict(Counter)
    tipo_posicion_by_norm: dict[tuple, Counter] = defaultdict(Counter)

    total_rows = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_structure = clean_text(row[idx["ESTRUCTURA"]])
        raw_area = clean_text(row[idx["AREA"]])
        raw_cargo_val = clean_text(row[idx["CARGO"]])

        if not raw_structure or not raw_area:
            skipped += 1
            continue

        mapped_structure = STRUCTURE_MAP.get(raw_structure)
        if not mapped_structure:
            skipped += 1
            continue

        total_rows += 1

        if not raw_cargo_val:
            continue

        norm_key = (mapped_structure, normalize_text(raw_area), normalize_text(raw_cargo_val))

        # Guardamos el nombre tal como aparece en el Excel (en mayúsculas y sin espacios extra)
        canonical_candidate = re.sub(r"\s+", " ", raw_cargo_val).strip().upper()
        raw_names_by_norm[norm_key][canonical_candidate] += 1

        raw_estado = row[idx["ESTADO"]] if "ESTADO" in idx else None
        raw_nombre = row[idx["NOMBRE COMPLETO"]] if "NOMBRE COMPLETO" in idx else None
        if is_countable_row(raw_estado, raw_nombre):
            countable_by_norm[norm_key] += 1

        if "TAP" in idx:
            tap_val = clean_text(row[idx["TAP"]])
            if tap_val and normalize_text(tap_val) not in NULL_VALS:
                tap_by_norm[norm_key][tap_val.upper()] += 1

        if "TIPO DE POSICION" in idx:
            pos_val = clean_text(row[idx["TIPO DE POSICION"]])
            if pos_val and normalize_text(pos_val) not in NULL_VALS:
                tipo_posicion_by_norm[norm_key][pos_val.upper()] += 1

    return {
        "raw_names_by_norm": raw_names_by_norm,
        "countable_by_norm": countable_by_norm,
        "tap_by_norm": tap_by_norm,
        "tipo_posicion_by_norm": tipo_posicion_by_norm,
        "stats": {
            "total_rows_procesadas": total_rows,
            "rows_omitidas": skipped,
            "cargos_unicos_normalizados": len(raw_names_by_norm),
            "cargos_con_dotacion": len(countable_by_norm),
        },
    }


# ---------------------------------------------------------------------------
# Paso 3 – Plan de creación con deduplicación
# ---------------------------------------------------------------------------

def plan_cargo_creation(snapshot):
    # Indexa áreas de la BD por (norm_estructura, norm_area)
    db_areas: dict[tuple, area] = {}
    for a in area.objects.select_related("estructura").all():
        key = (normalize_text(a.estructura.descripcion), normalize_text(a.descripcion))
        db_areas[key] = a

    to_create = []
    duplicates_merged = []
    missing_areas = []

    for norm_key, raw_counter in snapshot["raw_names_by_norm"].items():
        mapped_structure, norm_area, _ = norm_key
        area_key = (normalize_text(mapped_structure), norm_area)

        area_obj = db_areas.get(area_key)
        if area_obj is None:
            missing_areas.append({
                "estructura": mapped_structure,
                "area": norm_area,
                "cargo_normalizado": norm_key[2],
            })
            continue

        # Nombre canónico = el más frecuente en el Excel
        canonical = most_common_value(raw_counter)
        all_variants = list(raw_counter.keys())

        if len(all_variants) > 1:
            duplicates_merged.append({
                "canonico": canonical,
                "variantes_fusionadas": all_variants,
                "estructura": area_obj.estructura.descripcion,
                "area": area_obj.descripcion,
            })

        cantidad = snapshot["countable_by_norm"].get(norm_key, 0)
        tap = most_common_value(snapshot["tap_by_norm"].get(norm_key))
        tipo_posicion = most_common_value(snapshot["tipo_posicion_by_norm"].get(norm_key)) or "HC"

        to_create.append({
            "descripcion": canonical,
            "area_id": area_obj.id,
            "area_desc": area_obj.descripcion,
            "estructura_desc": area_obj.estructura.descripcion,
            "cantidad_aprobada": cantidad,
            "tap": tap,
            "tipo_posicion": tipo_posicion,
        })

    return {
        "to_create": to_create,
        "duplicates_merged": duplicates_merged,
        "missing_areas": missing_areas,
        "summary": {
            **snapshot["stats"],
            "areas_en_bd": len(db_areas),
            "cargos_a_crear": len(to_create),
            "areas_sin_match_en_bd": len(missing_areas),
            "duplicados_fusionados": len(duplicates_merged),
        },
    }


# ---------------------------------------------------------------------------
# Paso 4 – Aplicación
# ---------------------------------------------------------------------------

def apply_cargo_creation(plan):
    created = 0
    errors = []

    with transaction.atomic():
        for item in plan["to_create"]:
            try:
                cargo.objects.create(
                    descripcion=item["descripcion"],
                    area_id=item["area_id"],
                    activo=True,
                    cantidad_aprobada=item["cantidad_aprobada"],
                    criticidad="MEDIA",
                    tipo_posicion=item["tipo_posicion"],
                    tap=item["tap"],
                )
                created += 1
            except Exception as exc:
                errors.append({
                    "cargo": item["descripcion"],
                    "area": item["area_desc"],
                    "error": str(exc),
                })

    return {
        "creados": created,
        "errores": errors,
        "cargos_final": cargo.objects.count(),
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Reset + reconstrucción de cargos con deduplicación por acentos."
    )
    parser.add_argument("--input", default=DEFAULT_INPUT, help="Ruta al Excel")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Nombre de la hoja")
    parser.add_argument(
        "--reset",
        action="store_true",
        help="Borra cargos, contratos y personal antes de crear (mantiene áreas/estructuras)",
    )
    parser.add_argument("--apply", action="store_true", help="Ejecuta los cambios (sin --apply es dry-run)")
    args = parser.parse_args()

    mode = "apply" if args.apply else "dry-run"

    snapshot = load_excel_snapshot(args.input, args.sheet)
    plan = plan_cargo_creation(snapshot)

    output: dict = {
        "mode": mode,
        "reset_solicitado": args.reset,
        "resumen": plan["summary"],
        "duplicados_fusionados": plan["duplicates_merged"],
        "areas_sin_match": plan["missing_areas"][:30],
        "muestra_a_crear": plan["to_create"][:15],
    }

    if not args.apply:
        print(json.dumps(output, ensure_ascii=False, indent=2))
        return

    # -- Aplicar --
    if args.reset:
        print("step=reset_start", flush=True)
        reset_result = reset_hr_data()
        output["reset_result"] = reset_result
        print("step=reset_done", flush=True)
        # Recalcula el plan con BD limpia (cargos borrados → todas las áreas matchean igual)
        plan = plan_cargo_creation(snapshot)
        output["resumen"] = plan["summary"]

    print("step=cargos_start", flush=True)
    result = apply_cargo_creation(plan)
    output["resultado"] = result
    print("step=cargos_done", flush=True)

    print(json.dumps(output, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
