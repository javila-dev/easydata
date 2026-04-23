import calendar
import datetime
import json
import os
import re
import threading
import traceback
import unicodedata
import dateutil
from django.conf import settings
from django.http import FileResponse, HttpResponse, JsonResponse
from django.db.models.query_utils import Q
from django.db.models import Sum, Count, F, Case, When, IntegerField, Func, Prefetch
from django.shortcuts import render
from django.db import transaction
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import close_old_connections
from django.utils import timezone

import openpyxl
import openpyxl.workbook
import openpyxl.workbook.defined_name
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from requests import delete

from Atlantic.utils import JsonRender
from reports.models import areas_empresa
from useraccounts.utils import rol_permission
from human_resources.forms import auxiliosForm, checkworkersForm, descargosForms, importMasiveForm, retiroForm, workersForm
from human_resources.models import (area, arl, auxilios_contrato, base_personal, cambios_salario, cargo as cargos_personal, 
                                    ccf, ciudad, contratos_personal, departamento, descargos, empalmes, empleadores, 
                                    eps, estructura, fondo_cesantias, fondo_pensiones, historial,
                                    canal as canal_model, historico_base_teorica, importacion_personal_job, motivos_retiro, sede as sede_model, cceco, temporales, tipos_auxilio)


# Create your views here.
@login_required
@rol_permission('Gestion humana')
def landing(request):
    
    context = {}

    return render(request, 'landing_rh.html', context)

@login_required
@rol_permission('Gestion humana')
def human_resources(request):
    
    if request_is_ajax(request):
        if request.method == 'GET':
            todo = request.GET.get('todo')
            
            if todo == 'getworkers':
                active = True if request.GET.get('active') == 'true' else False
                retired = True if request.GET.get('retired') == 'true' else False
                
                if active and retired:
                    personal = base_personal.objects.all()
                elif active:
                    personal = base_personal.objects.filter(activo=True)
                elif retired:
                    personal = base_personal.objects.filter(activo=False)
                else:
                    personal = base_personal.objects.none()

                personal = personal.select_related(
                    'eps',
                    'pension',
                    'cesantias',
                    'arl',
                    'ccf',
                    'ciudad',
                    'departamento',
                ).prefetch_related(
                    Prefetch(
                        'trabajador',
                        queryset=contratos_personal.objects.select_related(
                            'empleador',
                            'temporal',
                            'cargo',
                            'area',
                            'area__estructura',
                            'canal',
                            'cceco',
                            'sede',
                            'ciudad_laboral',
                            'ciudad_laboral__departamento',
                            'jefe_inmediato',
                            'motivo_retiro',
                        ).prefetch_related(
                            'auxilios_contrato_set__tipo',
                        ).order_by('fecha_inicio', 'id'),
                        to_attr='prefetched_contracts',
                    )
                )
                    
                render_json = JsonRender(personal,
                            query_functions = ('get_full_name','contrato_activo',))
                data = {
                    'data': render_json.render()
                }
                
                return JsonResponse(data)

            elif todo == 'getworkerhistory':
                numero_id = request.GET.get('numero_id')
                worker = base_personal.objects.filter(
                    numero_identificacion=numero_id
                ).select_related(
                    'eps',
                    'pension',
                    'cesantias',
                    'arl',
                    'ccf',
                    'ciudad',
                    'departamento',
                ).prefetch_related(
                    Prefetch(
                        'trabajador',
                        queryset=contratos_personal.objects.select_related(
                            'empleador',
                            'temporal',
                            'cargo',
                            'area',
                            'area__estructura',
                            'canal',
                            'cceco',
                            'sede',
                            'ciudad_laboral',
                            'ciudad_laboral__departamento',
                            'jefe_inmediato',
                            'motivo_retiro',
                        ).prefetch_related(
                            'auxilios_contrato_set__tipo',
                        ).order_by('fecha_inicio', 'id'),
                        to_attr='prefetched_contracts',
                    )
                ).first()

                if worker is None:
                    return JsonResponse({'msj': 'Empleado no encontrado'}, status=404)

                return JsonResponse({
                    'historico_contratos': worker.historico_contratos(),
                    'historico_acciones': worker.historico_acciones(),
                })
            
            elif todo == 'get_subareas_by_cargo':
                cargo = request.GET.get('cargo')
                
                obj_cargo = cargos_personal.objects.get(pk=cargo)
                
                subareas = subarea.objects.filter(
                    area = obj_cargo.area
                ).order_by('descripcion')
                
                subareas_list = []
                
                for i in subareas:
                    subareas_list.append({
                        'value':i.pk,
                        'text':i.descripcion,
                        'name':i.descripcion
                        })
                
                data = {
                    'data':subareas_list
                }                
                
                return JsonResponse(data)
            
        elif request.method == 'POST':
            todo = request.POST.get('todo')        
            
            if todo == 'add_new_worker':
                numero_id = request.POST.get('numero_id')
                
                if base_personal.objects.filter(numero_identificacion = numero_id).exists():
                    data = {
                        'msj': 'El numero de identificación ingresado ya existe en la base de datos',
                        'class': 'error',
                        'errortype':'existalready'
                    }
                    
                    return JsonResponse(data)
                
                numero_id = request.POST.get('numero_id')
                tipo_id = request.POST.get('tipo_id')
                codigo_sap = request.POST.get('codigo_sap')
                primer_nombre = request.POST.get('primer_nombre')
                segundo_nombre = request.POST.get('segundo_nombre')
                primer_apellido = request.POST.get('primer_apellido')
                segundo_apellido = request.POST.get('segundo_apellido')
                email = request.POST.get('email')
                direccion = request.POST.get('direccion')
                _departamento = request.POST.get('departamento')
                _ciudad = request.POST.get('ciudad')
                celular = request.POST.get('celular')
                telefono = request.POST.get('telefono')
                tipo_vivienda = request.POST.get('tipo_vivienda')
                sexo = request.POST.get('sexo')
                rh = request.POST.get('rh')
                estado_civil = request.POST.get('estado_civil')
                fecha_nacimiento = request.POST.get('fecha_nacimiento')
                _eps = request.POST.get('eps')
                _pension = request.POST.get('pension')
                _cesantias = request.POST.get('cesantias')
                _arl = request.POST.get('arl')
                tipo_riesgo = request.POST.get('tipo_riesgo')
                _ccf = request.POST.get('ccf')
                talla_camisa = request.POST.get('talla_camisa')
                talla_pantalon = request.POST.get('talla_pantalon')
                talla_calzado = request.POST.get('talla_calzado')
                nivel_educativo = request.POST.get('nivel_educativo')
                tipo_ingreso = request.POST.get('tipo_ingreso')
                titulo = request.POST.get('titulo')
                
                modalidad_ingreso = request.POST.get('modalidad_ingreso')
                tipo_contrato = request.POST.get('tipo_contrato')
                empleador = request.POST.get('empleador')
                temporal = request.POST.get('temporal')
                fecha_inicio = request.POST.get('fecha_inicio')
                fecha_fin_pp = request.POST.get('fecha_fin_pp')
                fecha_fin_cto = request.POST.get('fecha_fin_cto')
                _area = request.POST.get('area')
                cargo = request.POST.get('cargo')
                ceco = request.POST.get('cceco')
                sede = request.POST.get('sede')
                jefe_inmediato = request.POST.get('jefe_inmediato')
                salario_base = request.POST.get('salario_base')
                auxilio_transporte = request.POST.get('auxilio_transporte')
                bonificacion = request.POST.get('bonificacion')
                base_bonificacion = request.POST.get('base_bonificacion')
                
                _cargo = cargos_personal.objects.get(pk=cargo)
                _area = area.objects.get(pk=_area)
                
                area_personal_count = contratos_personal.objects.filter(
                    area = _area, activo = True
                ).count()
                cargo_personal_count = contratos_personal.objects.filter(
                    cargo = _cargo, activo = True
                ).count()
                es_empalme = True if request.POST.get('es_empalme') == 'on' else False
                if cargo_personal_count >= _cargo.cantidad_aprobada:
                    if not es_empalme:
                        data = {
                            'msj': f'La cantidad de contrataciones aprobadas para este cargo ({_cargo.cantidad_aprobada}) ya está cubierta, ¿Quizás se trata de un Empalme o Licencia?',
                            'class': 'error',
                            'errortype':'quanty'
                        }
                        
                        return JsonResponse(data)
                
                
                _eps = None if _eps == '' else eps.objects.get(pk=_eps)
                pension = None if _pension == '' else fondo_pensiones.objects.get(pk=_pension)
                cesantias = None if _cesantias == '' else fondo_cesantias.objects.get(pk=_cesantias)
                _arl = None if _arl == '' else arl.objects.get(pk=_arl)
                _ccf = None if _ccf == '' else ccf.objects.get(pk=_ccf)
                jefe_inmediato = None if jefe_inmediato == '' else base_personal.objects.get(pk=jefe_inmediato)
                
                
                _ceco = None if ceco == '' else cceco.objects.get(pk=ceco)
                fecha_fin_pp = None if fecha_fin_pp == '' else fecha_fin_pp
                fecha_fin_cto = None if fecha_fin_cto == '' else fecha_fin_cto
                auxilio_transporte = None if auxilio_transporte == '' else auxilio_transporte.replace(',','')
                base_bonificacion = None if base_bonificacion == '' else base_bonificacion.replace(',','')
                temporal = None if temporal == '' else temporales.objects.get(pk=temporal)
                
                worker = base_personal.objects.create(
                    numero_identificacion = numero_id,
                    tipo_id = tipo_id,
                    codigo_sap = codigo_sap,
                    primer_nombre = primer_nombre,
                    segundo_nombre = segundo_nombre,
                    primer_apellido = primer_apellido,
                    segundo_apellido = segundo_apellido,
                    email = email,
                    direccion_residencia = direccion,
                    departamento = departamento.objects.get(pk=_departamento),
                    ciudad = ciudad.objects.get(pk=_ciudad),
                    contacto = celular,
                    contacto_otro = telefono,
                    tipo_vivienda = tipo_vivienda,
                    sexo = sexo,
                    rh = rh,
                    estado_civil = estado_civil,
                    fecha_nacimiento = fecha_nacimiento,
                    eps = _eps,
                    pension = pension,
                    cesantias = cesantias,
                    arl = _arl,
                    tipo_riesgo = tipo_riesgo,
                    talla_calzado = talla_calzado,
                    talla_camisa = talla_camisa,
                    talla_pantalon = talla_pantalon,
                    ccf = _ccf,
                    nivel_educativo = nivel_educativo,
                    titulo = titulo
                
                )
                
                contrato = contratos_personal.objects.create(
                    trabajador = worker,
                    fecha_inicio = fecha_inicio,
                    fecha_periodo_prueba = fecha_fin_pp,
                    fecha_fin = fecha_fin_cto,
                    tipo_ingreso = tipo_ingreso,
                    modalidad_ingreso = modalidad_ingreso,
                    tipo_contrato = tipo_contrato,
                    empleador = empleadores.objects.get(pk=empleador),
                    area = _area,
                    cargo = _cargo,
                    sede = sede_model.objects.get(pk=sede),
                    cceco = _ceco,
                    jefe_inmediato = jefe_inmediato,
                    salario_base = salario_base.replace(',',''),
                    auxilio_transporte = auxilio_transporte,
                    bonificacion = bonificacion,
                    base_bonificacion = base_bonificacion,
                    activo_desde = fecha_inicio,
                )
                
                auxilios = request.POST.getlist('tipo_auxilio')
                valores = request.POST.getlist('valor_auxilio')
                
                if auxilios is not None:
                    for auxilio, valor in zip(auxilios,valores):
                        auxilios_contrato.objects.create(
                            contrato = contrato, 
                            tipo = tipos_auxilio.objects.get(pk=auxilio), 
                            valor= valor.replace(',','')
                        )
                
                add_history(request.user, f'Agregó un nuevo trabajador {worker.get_full_name().upper()} ({worker.numero_identificacion})')
                      
                
                if es_empalme:
                    empleado_sale = request.POST.get('empleado_sale')
                    empleado_sale = base_personal.objects.get(pk=empleado_sale)
                    fecha_inicio_empalme = request.POST.get('fecha_inicio_empalme')
                    fecha_fin_empalme = request.POST.get('fecha_fin_empalme')
                    motivo_empalme = request.POST.get('motivo_empalme')
                    
                    empalmes.objects.create(
                        quien_ingresa = worker, 
                        quien_sale = empleado_sale,
                        contrato_sale = empleado_sale.contrato_activo(type='object'),
                        fecha_inicio = fecha_inicio_empalme,
                        fecha_fin = fecha_fin_empalme,
                        motivo = motivo_empalme,
                    )
                    
                    add_history(request.user, 
                                f'Registró un empalme entre SALE: {empleado_sale.get_full_name()} ({empleado_sale.numero_identificacion}) INGRESA: {worker.get_full_name().upper()} ({worker.numero_identificacion}) para el cargo {_cargo.descripcion}')
                    
                data = {
                        'msj': f'Se creó el empleado {worker.get_full_name().upper()} con exito',
                        'class': 'success'
                    }
                
                return JsonResponse(data)
            
            elif todo == 'modify_worker':
                
                numero_id = request.POST.get('numero_id')
                codigo_sap = request.POST.get('codigo_sap')
                primer_nombre = request.POST.get('primer_nombre')
                segundo_nombre = request.POST.get('segundo_nombre')
                primer_apellido = request.POST.get('primer_apellido')
                segundo_apellido = request.POST.get('segundo_apellido')
                email = request.POST.get('email')
                direccion = request.POST.get('direccion')
                _departamento = request.POST.get('departamento')
                _ciudad = request.POST.get('ciudad')
                celular = request.POST.get('celular')
                telefono = request.POST.get('telefono')
                tipo_vivienda = request.POST.get('tipo_vivienda')
                sexo = request.POST.get('sexo')
                rh = request.POST.get('rh')
                estado_civil = request.POST.get('estado_civil')
                fecha_nacimiento = request.POST.get('fecha_nacimiento')
                _eps = request.POST.get('eps')
                _pension = request.POST.get('pension')
                _cesantias = request.POST.get('cesantias')
                _arl = request.POST.get('arl')
                tipo_riesgo = request.POST.get('tipo_riesgo')
                _ccf = request.POST.get('ccf')
                talla_camisa = request.POST.get('talla_camisa')
                talla_pantalon = request.POST.get('talla_pantalon')
                talla_calzado = request.POST.get('talla_calzado')
                nivel_educativo = request.POST.get('nivel_educativo')
                jefe_inmediato = request.POST.get('jefe_inmediato')
                contrato_activo = request.POST.get('activecontract')
                titulo = request.POST.get('titulo')
                
                _eps = None if _eps == '' else eps.objects.get(pk=_eps)
                pension = None if _pension == '' else fondo_pensiones.objects.get(pk=_pension)
                cesantias = None if _cesantias == '' else fondo_cesantias.objects.get(pk=_cesantias)
                _arl = None if _arl == '' else arl.objects.get(pk=_arl)
                _ccf = None if _ccf == '' else ccf.objects.get(pk=_ccf)
                jefe_inmediato = None if jefe_inmediato == '' else base_personal.objects.get(pk=jefe_inmediato)
                
                
                worker = base_personal.objects.get(numero_identificacion=numero_id)

                worker.codigo_sap = codigo_sap
                worker.primer_nombre = primer_nombre
                worker.segundo_nombre = segundo_nombre
                worker.primer_apellido = primer_apellido
                worker.segundo_apellido = segundo_apellido
                worker.email = email
                worker.direccion_residencia = direccion
                worker.departamento = departamento.objects.get(pk=_departamento)
                worker.ciudad = ciudad.objects.get(pk=_ciudad)
                worker.contacto = celular
                worker.contacto_otro = telefono
                worker.tipo_vivienda = tipo_vivienda
                worker.sexo = sexo
                worker.rh = rh
                worker.estado_civil = estado_civil
                worker.fecha_nacimiento = fecha_nacimiento
                worker.eps = _eps
                worker.pension = pension
                worker.cesantias = cesantias
                worker.arl = _arl
                worker.tipo_riesgo = tipo_riesgo
                worker.talla_calzado = talla_calzado
                worker.talla_camisa = talla_camisa
                worker.talla_pantalon = talla_pantalon
                worker.ccf = _ccf
                worker.nivel_educativo = nivel_educativo
                worker.titulo = titulo
                
                worker.save()
                
                add_history(request.user, f'Modificó los datos del trabajador {worker.get_full_name().upper()} ({worker.numero_identificacion})')
                
                data = {
                        'msj': f'Se modificaron los datos del empleado {worker.get_full_name().upper()} con exito',
                        'class': 'success'
                    }
                
                return JsonResponse(data)
    
            elif todo == 'retire_worker':
                numero_id = request.POST.get('numero_id')
                worker = base_personal.objects.get(numero_identificacion = numero_id)
                fecha_retiro = request.POST.get('fecha_retiro')
                motivo_retiro = request.POST.get('motivo_retiro')
                motivo_real = request.POST.get('motivo_real')
                
                contratos = contratos_personal.objects.filter(
                    trabajador = worker.pk,
                    activo = True
                )
                if contratos.exists():
                    contrato = contratos.last()
                    worker.activo = False
                    contrato.fecha_retiro = fecha_retiro
                    contrato.motivo_retiro = motivos_retiro.objects.get(pk=motivo_retiro)
                    contrato.motivo_retiro_real = motivo_real
                    contrato.activo = False
                    contrato.activo_hasta = fecha_retiro
                    contrato.save()
                    worker.save()
                
                    data = {
                            'msj': f'Se retiró el empleado {worker.get_full_name()} con exito',
                            'class': 'success'
                        }
                    
                    add_history(request.user, f'Retiró al trabajador {worker.get_full_name()} ({worker.numero_identificacion})')
                
                else:
                    data = {
                            'msj': f'El empleado {worker.get_full_name()} no tiene ningun contrato activo',
                            'class': 'error'
                        }
                
                return JsonResponse(data)

            elif todo == 'add_new_contract':
                
                numero_id = request.POST.get('numero_id')
                id_cto = request.POST.get('activecontract')
                
                worker = base_personal.objects.get(numero_identificacion=numero_id)
                old_contrato = contratos_personal.objects.get(pk=id_cto)
                tipo = 'cambio de contrato'
                
                fecha_registro = worker.fecha_registro.date()
                yday = datetime.date.today() - datetime.timedelta(days=1)
                
                tipo_ingreso = request.POST.get('tipo_ingreso')
                modalidad_ingreso = request.POST.get('modalidad_ingreso')
                tipo_contrato = request.POST.get('tipo_contrato')
                empleador = request.POST.get('empleador')
                temporal = request.POST.get('temporal')
                fecha_inicio = request.POST.get('fecha_inicio')
                fecha_fin_pp = request.POST.get('fecha_fin_pp')
                fecha_fin_cto = request.POST.get('fecha_fin_cto')
                _area = request.POST.get('area')
                cargo = request.POST.get('cargo')
                ceco = request.POST.get('cceco')
                sede = request.POST.get('sede')
                jefe_inmediato = request.POST.get('jefe_inmediato')
                salario_base = request.POST.get('salario_base')
                auxilio_transporte = request.POST.get('auxilio_transporte')
                bonificacion = request.POST.get('bonificacion')
                base_bonificacion = request.POST.get('base_bonificacion')
                fecha_cambio = request.POST.get('fecha_inicio_cambio')
                fecha_fmt = datetime.datetime.strptime(fecha_cambio,'%Y-%m-%d').date()
                
                
                _cargo = cargos_personal.objects.get(pk=cargo)
                _area = area.objects.get(pk=_area)
                
                
                area_personal_count = contratos_personal.objects.filter(
                    area = _area, activo = True
                ).count()
                cargo_personal_count = contratos_personal.objects.filter(
                    cargo = _cargo, activo = True
                ).count()
                es_empalme = True if request.POST.get('es_empalme') == 'on' else False
                
                if _cargo.pk != old_contrato.cargo.pk and \
                    cargo_personal_count >= _cargo.cantidad_aprobada:
                        if not es_empalme:
                            data = {
                                'msj': f'La cantidad de contrataciones aprobadas para este cargo ({_cargo.cantidad_aprobada}) ya está cubierta, ¿Quizás se trata de un Empalme o Licencia?',
                                'class': 'error',
                                'errortype':'quanty'
                            }
                            
                            return JsonResponse(data)
                
                
                date_fecha_inicio = datetime.datetime.strptime(fecha_inicio,'%Y-%m-%d')
                
                if old_contrato.fecha_inicio > date_fecha_inicio.date():
                    data = {
                        'msj': 'El nuevo contrato no puede tener una fecha de inicio menor al último contrato activo',
                        'class': 'error'
                    }
                
                    return JsonResponse(data)
                
                fecha_fin_pp = None if fecha_fin_pp == '' else fecha_fin_pp
                fecha_fin_cto = None if fecha_fin_cto == '' else fecha_fin_cto
                auxilio_transporte = None if auxilio_transporte == '' else auxilio_transporte.replace(',','')
                base_bonificacion = None if base_bonificacion == '' else base_bonificacion.replace(',','')
                jefe_inmediato = None if jefe_inmediato == '' else base_personal.objects.get(pk=jefe_inmediato)
                temporal = None if temporal == '' else temporales.objects.get(pk=temporal)
                _ceco = None if ceco == '' else cceco.objects.get(pk=ceco)
                
                auxilios = request.POST.getlist('tipo_auxilio')
                valores = request.POST.getlist('valor_auxilio')
                
                if fecha_registro >= yday:
                    contrato = old_contrato
                    contrato.trabajador = worker
                    contrato.fecha_inicio = fecha_inicio
                    contrato.fecha_periodo_prueba = fecha_fin_pp
                    contrato.fecha_fin = fecha_fin_cto
                    contrato.tipo_ingreso = tipo_ingreso
                    contrato.modalidad_ingreso = modalidad_ingreso
                    contrato.tipo_contrato = tipo_contrato
                    contrato.empleador = empleadores.objects.get(pk=empleador)
                    contrato.temporal = temporal
                    contrato.area = _area
                    contrato.cargo = cargos_personal.objects.get(pk=cargo)
                    contrato.sede = sede_model.objects.get(pk=sede)
                    contrato.cceco = _ceco
                    contrato.jefe_inmediato = jefe_inmediato
                    contrato.salario_base = salario_base.replace(',', '')
                    contrato.auxilio_transporte = auxilio_transporte
                    contrato.bonificacion = bonificacion
                    contrato.base_bonificacion = base_bonificacion
                    contrato.activo_desde = fecha_cambio
                    contrato.save()
                    
                    auxilios_obj = auxilios_contrato.objects.filter(
                        contrato = contrato
                    )
                    
                    for aux in auxilios_obj:
                        aux.delete()
                            
                    emp = empalmes.objects.filter(
                        quien_ingresa = worker
                    )
                    
                    for e in emp: e.delete()
                                                
                    
                    msj = 'Modificó el contrato inicial'
                    
                else:
                    contrato = contratos_personal.objects.create(
                        trabajador = worker,
                        fecha_inicio = fecha_inicio,
                        fecha_periodo_prueba = fecha_fin_pp,
                        fecha_fin = fecha_fin_cto,
                        tipo_ingreso = tipo_ingreso,
                        modalidad_ingreso = modalidad_ingreso,
                        tipo_contrato = tipo_contrato,
                        empleador = empleadores.objects.get(pk=empleador),
                        temporal = temporal,
                        area = _area,
                        cargo = cargos_personal.objects.get(pk=cargo),
                        sede = sede_model.objects.get(pk=sede),
                        cceco = _ceco,
                        jefe_inmediato = jefe_inmediato,
                        salario_base = salario_base.replace(',',''),
                        auxilio_transporte = auxilio_transporte,
                        bonificacion = bonificacion,
                        base_bonificacion = base_bonificacion,
                        activo_desde = fecha_cambio
                    )
                    
                    if not worker.activo:
                        worker.activo = True
                        worker.save()
                        tipo = 'reintegro'
                    
                    
                    if old_contrato.activo:
                        old_contrato.activo = False
                        old_contrato.activo_hasta = fecha_fmt - datetime.timedelta(days=1)
                        old_contrato.save()
                    
                    if old_contrato.salario_base != contrato.salario_base:
                        tipo = 'cambio de salario'
                        motivo = request.POST.get('motivo_cambio_salario')
                        cambios_salario.objects.create(
                            trabajador = worker,
                            contrato = contrato,
                            fecha = fecha_fmt,
                            salario_anterior = old_contrato.salario_base,
                            nuevo_salario = contrato.salario_base,
                            motivo = motivo
                        )
                    
                        msj = f'Creó un nuevo contrato de tipo {tipo} para {worker.get_full_name()} ({worker.numero_identificacion})'
                
                add_history(request.user, msj)
                
                if es_empalme:
                    empleado_sale = request.POST.get('empleado_sale')
                    empleado_sale = base_personal.objects.get(pk=empleado_sale)
                    fecha_inicio_empalme = request.POST.get('fecha_inicio_empalme')
                    fecha_fin_empalme = request.POST.get('fecha_fin_empalme')
                    motivo_empalme = request.POST.get('motivo_empalme')
                    
                    empalmes.objects.create(
                        quien_ingresa = worker, 
                        quien_sale = empleado_sale,
                        contrato_sale = empleado_sale.contrato_activo(type='object'),
                        fecha_inicio = fecha_inicio_empalme,
                        fecha_fin = fecha_fin_empalme,
                        motivo = motivo_empalme,
                    )
                    
                    add_history(request.user, 
                            f'Registró un empalme entre SALE: {empleado_sale.get_full_name()} ({empleado_sale.numero_identificacion}) INGRESA: {worker.get_full_name().upper()} ({worker.numero_identificacion}) para el cargo {_cargo.descripcion}')
                    
                                
                if auxilios is not None:
                    for auxilio, valor in zip(auxilios,valores):
                        auxilios_contrato.objects.create(
                            contrato = contrato, 
                            tipo = tipos_auxilio.objects.get(pk=auxilio), 
                            valor= valor.replace(',','')
                        )    
                                
                data = {
                        'msj': msj,
                        'class': 'success'
                    }
                
                return JsonResponse(data)
                
            elif todo == 'check_workers':
                
                file = request.FILES.get('archivo_compara')
                sheet_names = request.POST.getlist('nombre_hoja')
                ids = request.POST.getlist('col_cc')
                fechas = request.POST.getlist('col_fecharetiro')
                today = datetime.date.today()
                book = openpyxl.load_workbook(file)
                
                    
                # 🔹 Validar que el archivo sea un Excel válido
                try:
                    book = openpyxl.load_workbook(file, data_only=True)
                except Exception as e:
                    return JsonResponse({'class': 'error', 'msj': 'El archivo no es un Excel válido.'})
                
                no_exist = []
                cedulas = set()  # Usar set para búsquedas más rápidas

                for j, sheet_name in enumerate(sheet_names):
                    # 🔹 Validar que la hoja existe en el archivo
                    if sheet_name not in book.sheetnames:
                        return JsonResponse({'class': 'error', 'msj': f'La hoja "{sheet_name}" no existe en el archivo.'})

                    sheet = book[sheet_name]
                    col_id = int(ids[j])
                    col_fecha = int(fechas[j])
                    
                    # 🔹 Leer todas las filas de una vez
                    rows = list(sheet.iter_rows(min_row=2, values_only=True))
                    
                    for row in rows:
                        if len(row) < max(col_id, col_fecha):  # Validar que la fila tenga suficientes columnas
                            continue
                        
                        cc = row[col_id - 1]  # Ajuste para índice basado en 1
                        fecha_retiro = row[col_fecha - 1]

                        if cc is None or cc == '':
                            continue  # Evitar procesar filas vacías

                        cedulas.add(cc)
                        worker_obj = base_personal.objects.filter(numero_identificacion=cc)

                        if worker_obj.exists():
                            worker = worker_obj.first()
                            if worker.activo:
                                if fecha_retiro:  # Cualquier fecha de retiro implica retiro
                                    no_exist.append([cc, sheet_name, 'RETIRAR'])
                            else:
                                if not fecha_retiro:  # Si no tiene fecha de retiro y está inactivo -> Recontratar
                                    no_exist.append([cc, sheet_name, 'RECONTRATAR'])
                        else:
                            no_exist.append([cc, sheet_name, 'INGRESAR'])

                # 🔹 Verificar empleados en BD que no están en el archivo
                workers = base_personal.objects.filter(activo=True).exclude(numero_identificacion__in=cedulas)
                for worker in workers:
                    no_exist.append([worker.numero_identificacion, worker.tipo_id, 'RETIRAR'])

                # 🔹 Crear archivo Excel de salida
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Cedula", "Origen", "Novedad"])
                for row in no_exist:
                    ws.append(row)

                filepath = settings.MEDIA_ROOT / "tmp/novedades_por_registrar.xlsx"
                wb.save(filepath)
                url = settings.MEDIA_URL + 'tmp/novedades_por_registrar.xlsx'

                return JsonResponse({
                    'class': 'success',
                    'msj': f'Puedes descargar las novedades pendientes por registrar <a href="{url}" target="_blank">Aquí</a>'
                })
            
            elif todo == 'add_descargo':
                numero_id = request.POST.get('numero_id')
                fecha_descargo = request.POST.get('fecha_descargo')
                
                worker = base_personal.objects.get(numero_identificacion=numero_id)
                
                descargos.objects.create(
                    trabajador = worker,
                    fecha = fecha_descargo,
                )
                
                data = {
                    'msj': f'Se registró el descargo para el empleado {worker.get_full_name()} con exito',
                    'class': 'success'
                }
                
                add_history(request.user, f'Registró un descargo a {worker.get_full_name()} ({worker.numero_identificacion})')
                
                return JsonResponse(data)
            
    context = {
        'workersForm':workersForm,
        'checkworkersForm': checkworkersForm,
        'auxiliosForm':auxiliosForm,
        'retiroForm':retiroForm,
        'descargosForms':descargosForms,
        'importMasiveForm': importMasiveForm
    }
    
    return render(request, 'personal_list.html', context)

@login_required
@rol_permission('Gestion humana')
def parameters(request):
    context = {
        'estructuras':estructura.objects.all(),
    }
    
    if request.method == 'GET':
        if request_is_ajax(request):
            todo = request.GET.get('todo')
            
            if todo == 'datatable-EIbyarea':
                area_ei = request.GET.get('area')
                
                ei = cargos_personal.objects.filter(
                    activo=True, area = area_ei
                )
                
                jsondata = JsonRender(ei, field_list = ('id','descripcion','cantidad_aprobada','criticidad'), 
                                      query_functions=('cantidad_actual','diferencia'))
                
                data = {
                    'data': jsondata.render()
                }
                
                
                return JsonResponse(data)
            
            elif todo == 'datatable-areas-by-estructura':
                estructura_id = request.GET.get('estructura')
                
                areas = area.objects.filter(estructura=estructura_id)
                
                jsondata = JsonRender(areas, field_list=('id','descripcion'),
                                      query_functions=('cantidad_actual',))
                
                data = {
                    'data': jsondata.render()
                }
                
                return JsonResponse(data)
                
    elif request.method == 'POST':
        if request_is_ajax(request):
            todo = request.POST.get('todo')
            
            if todo == 'update_ei':
                cargo_id = request.POST.get('cargo')
                cantidad = request.POST.get('cantidad')
                criticidad = request.POST.get('criticidad')
                
                cargo = cargos_personal.objects.get(pk=cargo_id)
                cantidad_actual = int(cargo.cantidad_aprobada)
                criticidad_actual = cargo.criticidad
                                
                
                
                
                
                if cantidad_actual != cantidad:
                    cargo.cantidad_aprobada = cantidad
                    action = f'Actualizó la estructura ideal del cargo {cargo.descripcion.upper()} de {cantidad_actual} a {cantidad}'
                    
                    historicos = historico_base_teorica.objects.filter(
                        cargo = cargo
                    )
                
                if criticidad_actual != criticidad:
                    cargo.criticidad = criticidad
                    action = f'Actualizó la criticidad del cargo {cargo.descripcion.upper()} de {criticidad_actual} a {criticidad}'
                    
                    historicos = historico_base_teorica.objects.filter(
                        cargo = cargo
                    )
                
                cargo.save()
                
                vigente_desde = None
                if historicos.exists():
                    vigente_desde = historicos.last().vigente_hasta + datetime.timedelta(days=1)
                    
                
                historico_base_teorica.objects.create(
                    cargo = cargo, cantidad = cantidad_actual,
                    vigente_desde = vigente_desde , 
                    vigente_hasta = datetime.date.today(),
                    usuario_crea = request.user
                )
                
                add_history(request.user, action)
                
                data = {
                    'class':'success',
                    'message': f'Se actualizó el cargo {cargo.descripcion} con éxito'
                }
                
                return JsonResponse(data)
            elif todo == 'add_cargo':
                area_ = request.POST.get('area')
                descripcion = request.POST.get('descripcion-cargo')
                cantidad_estructura = request.POST.get('cantidad-estructura-nuevo-cargo')
                criticidad = request.POST.get('criticidad_cargo')
                area_= area.objects.get(pk=area_)
                
                check_cargo = cargos_personal.objects.filter(
                    descripcion__iexact=descripcion,
                    area = area_,
                ).exists()
                
                if check_cargo:
                    data = {
                        'class': 'error',
                        'message': f'El cargo {descripcion.upper()} ya existe en el area {area_.descripcion}'
                    }
                    return JsonResponse(data) 
                
                cargo = cargos_personal.objects.create(
                    descripcion = descripcion.upper(),
                    area = area_,
                    cantidad_aprobada = cantidad_estructura,
                    criticidad = criticidad,
                )
                
                data = {
                    'class': 'success',
                    'message': f'Se agregó el cargo {descripcion.upper()} con éxito'
                }
                
                action = f'Creó un nuevo cargo {descripcion.upper()} en el area {area_.descripcion.upper()} '
                
                add_history(request.user, action)
                
                return JsonResponse(data)
            elif todo == 'add_area':
                descripcion = request.POST.get('descripcion-area')
                estructura_= estructura.objects.get(pk=request.POST.get('estructura'))
                
                check_estructura = area.objects.filter(
                    descripcion__iexact=descripcion,
                    estructura = estructura_
                ).exists()
                
                if check_estructura:
                    data = {
                        'class': 'error',
                        'message': f'El area {descripcion.upper()} ya existe en la estructura {estructura_.descripcion}'
                    }
                    return JsonResponse(data) 
                
                area_ = area.objects.create(
                    descripcion = descripcion.upper(),
                    estructura = estructura_,
                )
                
                data = {
                    'class': 'success',
                    'message': f'Se agregó el area {descripcion.upper()} con éxito'
                }
                
                action = f'Creó una nueva area {descripcion.upper()} en la estructura {estructura_.descripcion.upper()} '
                
                add_history(request.user, action)
                
                return JsonResponse(data)
                
            
    return render(request, 'gh_parameters.html', context)

@login_required
@rol_permission('Gestion humana')
def historico(request):
    
    
    context = {
        'acciones': historial.objects.all().order_by('-fecha')
    }
    
    return render(request, 'historial_gh.html',context)

@login_required
@rol_permission('Gestion humana')
def dashboard(request):
    
    if request.method == 'GET':
        todo = request.GET.get('todo')

        if todo == 'getHRdata':
            HR = base_personal.objects.filter(activo=True)
            nombres = ['nombre','sexo','fecha_inicio','tipo_ingreso','modalidad','tipo_contrato',
                       'empleador','temporal','cargo','area','estructura','dependencia',
                       'cceco','sede','map_sede','salario_base','auxilio_transporte','fecha_retiro','activo',
                       'rango_salario']
            data = []
            for p in HR:
               i = contratos_personal.objects.filter(activo=True, trabajador=p.pk).last()
               if i is None:
                   continue

               data.append([
                    p.get_full_name(),
                    p.sexo,
                    i.fecha_inicio,
                    i.tipo_ingreso,
                    i.modalidad_ingreso,
                    i.tipo_contrato,
                    i.empleador.nombre,
                    i.temporal.nombre if i.temporal != None else '',
                    i.cargo.descripcion,
                    i.area.descripcion,
                    i.area.estructura.descripcion,
                    i.area.estructura.dependecia.descripcion,
                    i.cceco.descripcion if i.cceco != None else '',
                    i.sede.descripcion,
                    i.sede.codigo_map,
                    i.salario_base,
                    i.auxilio_transporte,
                    i.fecha_retiro,
                    i.activo,
                    i.rango_salario(),
               ])
            data.insert(0,nombres)
            return JsonResponse(data,safe=False)
            
        elif todo == 'getpositionsData':
            
            cargos = cargos_personal.objects.filter(activo=True).annotate(
                actual = Count(Case(When(
                    contratos_personal__activo= True,
                    then = 1
                )))
            ).annotate(
                diferencia = F('cantidad_aprobada') - F('actual'),
                area_descripcion = F('area__descripcion'),
                estructura_descripcion = F('area__estructura__descripcion'),
                dependencia_descripcion = F('area__estructura__dependecia__descripcion'),
            ).values_list()
            
            cargos_list = list(cargos)
            nombres = ['id','descripcion','area_id','activo','cantidad_aprobada',
                       'criticidad','actual','diferencia',
                       'area','estructura','dependecia']
            cargos_list.insert(0,nombres)
            
            return JsonResponse(cargos_list,safe=False)    
    
    return HttpResponse()

@login_required
@rol_permission('Gestion humana')
def transitions(request):
    context = {
        
    }
    
    if request.method == 'GET':
        if request_is_ajax(request):
            todo =  request.GET.get('todo')
            
            if todo == 'datatable-transitions':
                tipo = request.GET.get('type')
                
                empalmes_ = empalmes.objects.filter(contrato_sale__fecha_retiro__isnull = True)
                
                if tipo == 'completed':
                    empalmes_ = empalmes.objects.all()
                
                
                jsondata = JsonRender(empalmes_,
                                      query_functions=('dias','dias_reales',
                                                       'quien_ingresa.get_full_name',
                                                       'quien_sale.get_full_name',
                                                       )).render()
                
                data = {
                    'data':jsondata
                }
                
                
                return JsonResponse(data)
            
    
    return render(request, 'gh_transitions.html', context)

@login_required
def errors_and_warnings(request):
    
    errors = []
    
    next_15 = datetime.date.today() + datetime.timedelta(days=15)
    next_month = datetime.date.today() + datetime.timedelta(days=30)
    
    empalmes_next = empalmes.objects.filter(
        contrato_sale__fecha_retiro__isnull = True,
        fecha_fin__lte = next_15,
        fecha_fin__gt = datetime.date.today()
    )
    
    vencimiento_contratos_next = contratos_personal.objects.filter(
        activo = True, fecha_fin__lte = next_month
    )
    
    empalmes_vencidos = empalmes.objects.filter(
        contrato_sale__fecha_retiro__isnull = True,
        fecha_fin__lt = datetime.date.today()
    )
    
    vencimiento_contratos_vencidos = contratos_personal.objects.filter(
        activo = True, 
        fecha_fin__lt = datetime.date.today()
    )
    
    
    cargos_oversized = cargos_personal.objects.annotate(
        cantidad_real = Count(Case(When(
                    contratos_personal__activo= True,
                    then = 1
                )))
        ).filter(
        activo = True,
        cantidad_real__gt = F('cantidad_aprobada')
    )
        
    data = {
        'warnings': {
            'empalmes':JsonRender(empalmes_next).render(),
            'contratos': JsonRender(vencimiento_contratos_next).render(),
        },
        'errors': {
            'empalmes':JsonRender(empalmes_vencidos).render(),
            'contratos':JsonRender(vencimiento_contratos_vencidos).render(),
            'cargos':JsonRender(cargos_oversized,annotates=('cantidad_real',)).render(),
        }
    }
    
    return JsonResponse(data)

def charts_reports(request):
    
    if request_is_ajax(request):
        if request.method == 'GET':
            todo = request.GET.get('todo')
            
            if todo == 'chart':
                report= request.GET.get('report')
                month = request.GET.get('month')
                year = request.GET.get('year')
                last_date = datetime.date.today()
                
                if report == 'quantybyareas':
                    
                    dataset_subareas = subarea.objects.filter(
                        base_personal__activo = True
                    )
                    
                    dataset_areas = area.objects.filter(
                        cargo__base_personal__activo = True
                    )
                    
                    if (month != "" and month is not None) and (year!= "" and year is not None):
                        day = calendar.monthrange(year, month)[1]

                        """ initial_date = datetime.datetime.strptime(f'01/{month}/{year}','%d/%m/%Y') """
                        last_date = datetime.datetime.strptime(f'{day}/{month}/{year}','%d/%m/%Y')
                        
                        dataset_subareas = subarea.objects.filter(
                            base_personal__fecha_ingreso__lte = last_date,
                            base_personal__fecha_retiro__gt = last_date
                        )
                        
                        dataset_areas = area.objects.filter(
                            cargo__base_personal__fecha_ingreso__lte = last_date,
                            cargo__base_personal__fecha_retiro__gt = last_date
                        )
                        
                    dataset_subareas = dataset_subareas.annotate(total = Count('base_personal'))
                    labels_subareas = [i.descripcion for i in dataset_subareas]
                    dataset_areas = dataset_areas.annotate(total = Count('cargo__base_personal'))
                    labels_areas = [i.descripcion for i in dataset_areas]
                    
                    data = {
                        'areas': {
                            'dataset_label':'Areas',
                            'labels':labels_areas,
                            'data': JsonRender(dataset_areas, 
                                               annotates=['total',], 
                                               query_functions=[f'getbudgets("{last_date}")',
                                                                f'budgetdifference("{last_date}")',
                                                                f'total_salarios("{last_date}")',]
                                               ).render(),
                        },
                        'subareas': {
                            'dataset_label':'Subareas',
                            'labels':labels_subareas,
                            'data': JsonRender(dataset_subareas,
                                               annotates=['total',], 
                                               query_functions=[f'getbudgets("{last_date}")',
                                                                f'budgetdifference("{last_date}")',
                                                                f'total_salarios("{last_date}")',
                                                                ]).render(),
                        }
                    }
                    
                    return JsonResponse(data)
                
                elif report == 'subareasbyareas':
                    dataset_subareas = subarea.objects.filter(
                        base_personal__activo = True
                    )
                    not_now = False
                    if (month != "" and month is not None) and (year!= "" and year is not None):
                        day = calendar.monthrange(year, month)[1]

                        last_date = datetime.datetime.strptime(f'{day}/{month}/{year}','%d/%m/%Y')
                        
                        dataset_subareas = subarea.objects.filter(
                            base_personal__fecha_ingreso__lte = last_date,
                            base_personal__fecha_retiro__gt = last_date
                        )
                        
                    dataset_subareas = dataset_subareas.annotate(total = Count('base_personal')
                                                                 ).order_by('-total')
                    labels_subareas = [i.descripcion for i in dataset_subareas]
                    labels_areas = [i.descripcion for i in area.objects.all()]
                    
                    data = {
                        'labels': [i.descripcion for i in area.objects.all()]
                    }
                    
                    for subar in dataset_subareas:
                        totales = [0,0,0,0]
                        totales[labels_areas.index(subar.area.descripcion)] +=  subar.total
                        
                        data[subar.descripcion] = {
                            'dataset_label' : subar.descripcion,
                            'data': totales
                        }
                        
                    return JsonResponse(data)
                    
                        
                
    
    context = {
        
    }
    
    return render(request, 'gerencial_resume.html', context)

def personalbudget(request):
    
    context = {
        
    }
    
    return render(request, 'personal_budget.html', context)


def parse_import_date(value):
    if not value:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    try:
        return datetime.date.fromisoformat(str(value)[:10])
    except Exception:
        return None


def build_import_response(records):
    importables = [r for r in records if r.get('can_import') and not r.get('ignored_row')]
    no_importables = [r for r in records if (not r.get('can_import')) and (not r.get('ignored_row'))]
    ignored_rows = [r for r in records if r.get('ignored_row')]
    return {
        'records': records,
        'importables': importables,
        'no_importables': no_importables,
        'ignored_rows': ignored_rows,
    }


def serialize_no_importables(no_importables, limit=None):
    items = no_importables if limit is None else no_importables[:limit]
    return [
        {
            'fila': item.get('row_number'),
            'id': item.get('numero_identificacion'),
            'estado': 'ACTIVO' if item.get('activo') is True else 'INACTIVO',
            'issues': item.get('issues', []),
        }
        for item in items
    ]


def process_import_records(records, progress_callback=None, should_stop=None):
    grouped = build_import_response(records)
    importables = grouped['importables']
    no_importables = grouped['no_importables']
    ignored_rows = grouped['ignored_rows']

    summary = {
        'creados': 0,
        'actualizados': 0,
        'omitidos': 0,
        'conflictos': 0,
        'errores': 0,
        'total_procesados': len(importables),
        'no_importables': len(records) - len(importables),
        'no_importables_detalle': len(no_importables),
        'ignored_rows': len(ignored_rows),
        'detalles_error': [],
        'detalles_conflicto': [],
        'detalles_no_importables': serialize_no_importables(no_importables),
        'stopped': False,
        'class': 'success',
    }

    if progress_callback:
        progress_callback(
            status='RUNNING',
            total_filas=len(records),
            filas_importables=len(importables),
            filas_procesadas=0,
            creados=0,
            actualizados=0,
            omitidos=0,
            conflictos=0,
            errores=0,
            no_importables=len(no_importables),
            ignoradas=len(ignored_rows),
            detalles_no_importables=summary['detalles_no_importables'],
        )

    for index, rec in enumerate(importables, start=1):
        if should_stop and should_stop():
            summary['stopped'] = True
            break
        try:
            bp_fields = {
                'tipo_id':              rec.get('tipo_id') or 'CC',
                'primer_nombre':        rec.get('primer_nombre'),
                'segundo_nombre':       rec.get('segundo_nombre'),
                'primer_apellido':      rec.get('primer_apellido'),
                'segundo_apellido':     rec.get('segundo_apellido'),
                'eps_id':               rec.get('eps_id'),
                'pension_id':           rec.get('pension_id'),
                'cesantias_id':         rec.get('cesantias_id'),
                'arl_id':               rec.get('arl_id'),
                'tipo_riesgo':          rec.get('tipo_riesgo') or '0',
                'ccf_id':               rec.get('ccf_id'),
                'email':                rec.get('email'),
                'direccion_residencia': rec.get('direccion_residencia'),
                'ciudad_id':            rec.get('ciudad_id'),
                'departamento_id':      rec.get('departamento_id'),
                'contacto':             str(rec.get('contacto')) if rec.get('contacto') else None,
                'contacto_otro':        str(rec.get('contacto_otro')) if rec.get('contacto_otro') else None,
                'tipo_vivienda':        rec.get('tipo_vivienda'),
                'nivel_educativo':      rec.get('nivel_educativo'),
                'titulo':               rec.get('titulo'),
                'sexo':                 rec.get('sexo'),
                'rh':                   rec.get('rh'),
                'estado_civil':         rec.get('estado_civil'),
                'fecha_nacimiento':     parse_import_date(rec.get('fecha_nacimiento')),
                'talla_camisa':         str(rec.get('talla_camisa')) if rec.get('talla_camisa') else None,
                'talla_pantalon':       str(rec.get('talla_pantalon')) if rec.get('talla_pantalon') else None,
                'talla_calzado':        str(rec.get('talla_calzado')) if rec.get('talla_calzado') else None,
                'codigo_sap':           str(rec.get('codigo_sap')) if rec.get('codigo_sap') else None,
                'activo':               rec.get('activo') if rec.get('activo') is not None else True,
            }

            fecha_inicio = parse_import_date(rec.get('fecha_inicio'))
            fecha_retiro = parse_import_date(rec.get('fecha_retiro'))
            es_activo = rec.get('activo') is True

            contrato_fields = {
                'tipo_ingreso':         rec.get('tipo_ingreso') or 'V',
                'modalidad_ingreso':    rec.get('modalidad_ingreso') or 'NUEVO INGRESO DIRECTO',
                'tipo_contrato':        rec.get('tipo_contrato') or 'Indefinido',
                'tipo_posicion':        rec.get('tipo_posicion'),
                'empleador_id':         rec.get('empleador_id'),
                'temporal_id':          rec.get('temporal_id'),
                'cargo_id':             rec.get('cargo_id'),
                'area_id':              rec.get('area_id'),
                'sede_id':              rec.get('sede_id'),
                'cceco_id':             rec.get('cceco_id'),
                'ciudad_laboral_id':    rec.get('ciudad_laboral_id'),
                'motivo_retiro_id':     rec.get('motivo_retiro_id'),
                'motivo_retiro_real':   rec.get('motivo_retiro_real'),
                'fecha_periodo_prueba': parse_import_date(rec.get('fecha_periodo_prueba')),
                'fecha_retiro':         fecha_retiro,
                'activo_desde':         fecha_inicio,
                'activo_hasta':         fecha_retiro if not es_activo else None,
                'activo':               es_activo,
                'salario_base':         rec.get('salario_base') or 0,
                'auxilio_transporte':   rec.get('auxilio_transporte'),
                'base_bonificacion':    rec.get('base_bonificacion'),
                'bonificacion':         rec.get('bonificacion'),
            }
            auxilio_specs = build_import_auxilio_specs(rec)

            worker = base_personal.objects.filter(
                numero_identificacion=rec['numero_identificacion']
            ).first()

            if worker is None:
                with transaction.atomic():
                    worker = base_personal.objects.create(
                        numero_identificacion=rec['numero_identificacion'],
                        **bp_fields,
                    )
                    contract_obj = contratos_personal.objects.create(
                        trabajador=worker,
                        fecha_inicio=fecha_inicio,
                        **contrato_fields,
                    )
                    sync_contract_auxilios(contract_obj, auxilio_specs)
                summary['creados'] += 1
            else:
                existing_contracts = list(
                    contratos_personal.objects.filter(trabajador=worker).order_by('-fecha_inicio', '-id')
                )
                contract_decision = resolve_contract_import(existing_contracts, fecha_inicio, contrato_fields)
                if contract_decision['action'] == 'conflict':
                    summary['conflictos'] += 1
                    conflict_target = contract_decision.get('target')
                    summary['detalles_conflicto'].append({
                        'fila': rec.get('row_number'),
                        'id': rec.get('numero_identificacion'),
                        'estado': 'ACTIVO' if rec.get('activo') is True else 'INACTIVO',
                        'motivo': contract_decision.get('reason'),
                        'contrato_id': getattr(conflict_target, 'pk', None),
                        'fecha_inicio_existente': comparable_import_value(getattr(conflict_target, 'fecha_inicio', None)) if conflict_target else None,
                    })
                else:
                    merged_bp_fields = merge_import_personal_fields(worker, bp_fields)
                    worker_changes = import_field_changes(worker, merged_bp_fields)
                    contract_changes = contract_decision.get('changes', {})
                    contract_changed = False
                    contract_created = False
                    auxilios_changed = False

                    with transaction.atomic():
                        if worker_changes:
                            for field_name, new_value in worker_changes.items():
                                setattr(worker, field_name, new_value)
                            worker.save(update_fields=list(worker_changes.keys()))

                        if contract_decision['action'] == 'create':
                            contract_obj = contratos_personal.objects.create(
                                trabajador=worker,
                                fecha_inicio=fecha_inicio,
                                **contrato_fields,
                            )
                            auxilios_changed = sync_contract_auxilios(contract_obj, auxilio_specs)
                            contract_created = True
                        elif contract_decision['action'] == 'update':
                            contract_obj = contract_decision['target']
                            if not contract_changes:
                                contract_changes = import_field_changes(contract_obj, contrato_fields)
                            if contract_changes:
                                for field_name, new_value in contract_changes.items():
                                    setattr(contract_obj, field_name, new_value)
                                contract_obj.save(update_fields=list(contract_changes.keys()))
                                contract_changed = True
                            auxilios_changed = sync_contract_auxilios(contract_obj, auxilio_specs)

                    if contract_created:
                        summary['creados'] += 1
                    elif worker_changes or contract_changed or auxilios_changed:
                        summary['actualizados'] += 1
                    else:
                        summary['omitidos'] += 1

        except Exception as e:
            summary['errores'] += 1
            summary['detalles_error'].append({
                'fila': rec.get('row_number'),
                'id': rec.get('numero_identificacion'),
                'estado': 'ACTIVO' if rec.get('activo') is True else 'INACTIVO',
                'error': str(e),
            })

        if progress_callback and (index == len(importables) or index % 10 == 0):
            progress_callback(
                status='RUNNING',
                total_filas=len(records),
                filas_importables=len(importables),
                filas_procesadas=index,
                creados=summary['creados'],
                actualizados=summary['actualizados'],
                omitidos=summary['omitidos'],
                conflictos=summary['conflictos'],
                errores=summary['errores'],
                no_importables=len(no_importables),
                ignoradas=len(ignored_rows),
                detalles_conflicto=summary['detalles_conflicto'][:50],
                detalles_error=summary['detalles_error'][:50],
                detalles_no_importables=summary['detalles_no_importables'][:50],
            )

    if summary['stopped'] and progress_callback:
        progress_callback(
            status='STOPPING',
            total_filas=len(records),
            filas_importables=len(importables),
            filas_procesadas=(
                summary['creados'] +
                summary['actualizados'] +
                summary['omitidos'] +
                summary['conflictos'] +
                summary['errores']
            ),
            creados=summary['creados'],
            actualizados=summary['actualizados'],
            omitidos=summary['omitidos'],
            conflictos=summary['conflictos'],
            errores=summary['errores'],
            no_importables=len(no_importables),
            ignoradas=len(ignored_rows),
            detalles_conflicto=summary['detalles_conflicto'][:50],
            detalles_error=summary['detalles_error'][:50],
            detalles_no_importables=summary['detalles_no_importables'][:50],
        )

    summary['class'] = 'success' if summary['conflictos'] == 0 and summary['errores'] == 0 else 'warning'
    return summary


def update_import_job(job_id, **fields):
    importacion_personal_job.objects.filter(pk=job_id).update(**fields)


def run_initial_import_job(job_id, user_id):
    close_old_connections()
    job = importacion_personal_job.objects.get(pk=job_id)
    try:
        update_import_job(job_id, status='MAPPING', fecha_inicio=timezone.now())
        job.refresh_from_db(fields=['stop_requested'])
        if job.stop_requested:
            update_import_job(job_id, status='STOPPED', fecha_fin=timezone.now())
            return
        from scripts.map_hr_snapshot import Mapper, load_rows

        source_rows = load_rows(job.ruta_archivo)
        job.refresh_from_db(fields=['stop_requested'])
        if job.stop_requested:
            update_import_job(
                job_id,
                status='STOPPED',
                fecha_fin=timezone.now(),
                total_filas=len(source_rows),
            )
            return
        mapper = Mapper()
        records = [mapper.map_row(raw_row, row_number) for row_number, raw_row in source_rows]

        def progress_callback(**kwargs):
            kwargs['fecha_actualizacion'] = timezone.now()
            update_import_job(job_id, **kwargs)

        def should_stop():
            return importacion_personal_job.objects.filter(pk=job_id, stop_requested=True).exists()

        summary = process_import_records(
            records,
            progress_callback=progress_callback,
            should_stop=should_stop,
        )
        update_import_job(
            job_id,
            status='STOPPED' if summary.get('stopped') else 'COMPLETED',
            fecha_fin=timezone.now(),
            total_filas=len(records),
            filas_importables=summary['total_procesados'],
            filas_procesadas=(
                summary['creados'] +
                summary['actualizados'] +
                summary['omitidos'] +
                summary['conflictos'] +
                summary['errores']
            ),
            creados=summary['creados'],
            actualizados=summary['actualizados'],
            omitidos=summary['omitidos'],
            conflictos=summary['conflictos'],
            errores=summary['errores'],
            no_importables=summary['no_importables_detalle'],
            ignoradas=summary['ignored_rows'],
            detalles_conflicto=summary['detalles_conflicto'],
            detalles_error=summary['detalles_error'],
            detalles_no_importables=summary['detalles_no_importables'],
        )
        user = User.objects.get(pk=user_id)
        if (summary['creados'] + summary['actualizados'] + summary['omitidos']) > 0:
            add_history(
                user,
                f'Importó personal desde carga inicial: {summary["creados"]} creados, {summary["actualizados"]} actualizados, {summary["omitidos"]} omitidos, {summary["conflictos"]} conflictos, {summary["errores"]} errores'
            )
    except Exception as e:
        update_import_job(
            job_id,
            status='FAILED',
            fecha_fin=timezone.now(),
            mensaje_error=str(e),
        )
    finally:
        try:
            if job.ruta_archivo and os.path.exists(job.ruta_archivo):
                os.unlink(job.ruta_archivo)
        except Exception:
            pass
        close_old_connections()

@login_required
def masivepartnercreation(request):
    if request.method == 'GET':
        todo = request.GET.get('todo')
        
        if todo == 'export-file':
            wb = openpyxl.Workbook()
            ws_data = wb.active
            ws_data.title = 'data'
            ws_params = wb.create_sheet('parametros')
            ws_params.title = "parametros"
            ws_params.sheet_state = 'hidden'
            
            ws_data.append(
                ['numero_id','tipo_id','codigo_sap','primer_nombre','segundo_nombre',
                 'primer_apellido','segundo_apellido','email','direccion',
                 'ciudad_residencia','celular','telefono','tipo_vivienda','sexo','rh','estado_civil',
                 'fecha_nacimiento','eps','fondo_pension','fondo_cesantias','arl',
                 'tipo_riesgo','caja_compensacion','talla_camisa','talla_pantalon',
                 'talla_calzado','nivel_educativo','tipo_ingreso','modalidad_ingreso',
                 'tipo_contrato','empleador','temporal',
                 'fecha_inicio_contrato','fecha_fin_periodo_prueba','fecha_fin_contrato',
                 'gerencia','area','cargo','ubicación','centro_costo','salario_base',
                 'auxilio_transporte','bonificacion','base_bonificacion','tipo_auxilio','auxilio',
                 'Titulo o profesion','ciudad_laboral']
            )
            
            grnc = estructura.objects.all()
            ar = area.objects.all()
            crgs = cargos_personal.objects.all() 
            
            col, row = 4, 0
            
            for a in grnc:
                row += 1
                rng_name = f'{a.descripcion.replace(" ","_")}_{a.pk}'
                rng_name = rng_name.replace("(","_")
                rng_name = rng_name.replace(")","_")
                ws_params.cell(row,col).value = rng_name
                
            rng = openpyxl.workbook.defined_name.DefinedName(
                'GERENCIAS',
                attr_text=f'parametros!$D$1:$D${row}')
            wb.defined_names.append(rng)
            
            col, row = 1, 0
            for g in grnc:
                row += 1
                row_ini = row + 1
                rng_name = f'{g.descripcion.replace(" ","_")}_{g.pk}'
                rng_name = rng_name.replace("(","_")
                rng_name = rng_name.replace(")","_")
                ws_params.cell(row,col).value = rng_name
                ar_x_grnc = ar.filter(estructura=g.pk)
                for a in ar_x_grnc:
                    row += 1
                    _name = f'{a.descripcion.replace(" ","_")}_{a.pk}'
                    _name = _name.replace("(","_")
                    _name = _name.replace(")","_")
                    ws_params.cell(row,col).value = _name
                
                rng = openpyxl.workbook.defined_name.DefinedName(
                    rng_name,
                    attr_text=f'parametros!$A${row_ini}:$A${row}')
                wb.defined_names.append(rng)
            
            col, row = 2, 1
            for a in ar:
                crg = crgs.filter(area = a.pk)
                rng_name = f'{a.descripcion.replace(" ","_")}_{a.pk}'
                rng_name = rng_name.replace("(","_")
                rng_name = rng_name.replace(")","_")
                ws_params.cell(row,col).value = rng_name
                row += 1
                row_ini = row
                for c in crg:
                    ws_params.cell(row,col).value = f'{c.descripcion}_{c.pk}'
                    row_end = row
                    row += 1
                rng = openpyxl.workbook.defined_name.DefinedName(
                    rng_name,
                    attr_text=f'parametros!$B${row_ini}:$B${row_end}')
                wb.defined_names.append(rng)
            
            col, row = 3, 0
            for c in ciudad.objects.all().order_by('departamento__nombre','nombre'):
                row += 1
                ws_params.cell(row,col).value = f'{c.nombre} ({c.departamento.nombre[:3].upper()})_{c.pk}'
            rng = openpyxl.workbook.defined_name.DefinedName(
                'CIUDADES',
                attr_text=f'parametros!$C$1:$C${row}')
            wb.defined_names.append(rng)
                
            col, row = 5, 0
            for i in eps.objects.all():
                row += 1
                ws_params.cell(row,col).value = f'{i.nombre}_{i.pk}'                
            rng = openpyxl.workbook.defined_name.DefinedName(
                'EPS',
                attr_text=f'parametros!$E$1:$E${row}')
            wb.defined_names.append(rng)
            
            col,row = 6,0
            for i in ccf.objects.all():
                row += 1
                ws_params.cell(row,col).value = f'{i.nombre}_{i.pk}'                
            rng = openpyxl.workbook.defined_name.DefinedName(
                'CCF',
                attr_text=f'parametros!$F$1:$F${row}')
            wb.defined_names.append(rng)
                        
            col,row = 7,0
            for i in sede_model.objects.all():
                row += 1
                ws_params.cell(row,col).value = f'{i.descripcion}_{i.pk}'                
            rng = openpyxl.workbook.defined_name.DefinedName(
                'UBICACIONES',
                attr_text=f'parametros!$G$1:$G${row}')
            wb.defined_names.append(rng)
            
            col,row = 8,0
            for i in cceco.objects.all():
                row += 1
                ws_params.cell(row,col).value = f'{i.codigo_sap} {i.descripcion}_{i.pk}'                
            rng = openpyxl.workbook.defined_name.DefinedName(
                'CECO',
                attr_text=f'parametros!$H$1:$H${row}')
            wb.defined_names.append(rng)
            
            nivel_educativo = ["BASICA PRIMARIA", "SECUNDARIA", "BACHILLER", "TECNICO",
                               "TECNOLOGO", "UNIVERSITARIO/PROFESIONAL","INGENIERO",
                               "ESPECIALISTA","MAGISTER","DOCTORADO"]
            col,row = 9,0
            for i in nivel_educativo:
                row += 1
                ws_params.cell(row,col).value = i                
            rng = openpyxl.workbook.defined_name.DefinedName(
                'NIVEL_EDUCATIVO',
                attr_text=f'parametros!$I$1:$I${row}')
            wb.defined_names.append(rng)
            
            modalidad_ingreso = ["Nuevo ingreso por temporal", "Nuevo ingreso directo", 
                               "Nuevo ingreso directo (temporal a directo)","Reingreso"]
            col,row = 10,0
            for i in modalidad_ingreso:
                row += 1
                ws_params.cell(row,col).value = i                
            rng = openpyxl.workbook.defined_name.DefinedName(
                'MODALIDAD_INGRESO',
                attr_text=f'parametros!$J$1:$J${row}')
            wb.defined_names.append(rng)
            
            col,row = 11,0
            for i in tipos_auxilio.objects.all():
                row += 1
                ws_params.cell(row,col).value = f'{i.descripcion}_{i.pk}'                
            rng = openpyxl.workbook.defined_name.DefinedName(
                'AUXILIOS',
                attr_text=f'parametros!$K$1:$K${row}')
            wb.defined_names.append(rng)
                    
            PENSION = str([f'{x.nombre}_{x.pk}' for x in fondo_pensiones.objects.all()]
                      ).replace("[",""
                    ).replace("]",""
                    ).replace("'",'')
            CESANTIAS = str([f'{x.nombre}_{x.pk}' for x in fondo_cesantias.objects.all()]
                      ).replace("[",""
                    ).replace("]",""
                    ).replace("'",'')
            ARL = str([f'{x.nombre}_{x.pk}' for x in arl.objects.all()]
                      ).replace("[",""
                    ).replace("]",""
                    ).replace("'",'')
            EMPLEADORES = str([f'{x.nombre}_{x.pk}' for x in empleadores.objects.all()]
                      ).replace("[",""
                    ).replace("]",""
                    ).replace("'",'')     
            TEMPORALES = str([f'{x.nombre}_{x.pk}' for x in temporales.objects.all()]
                      ).replace("[",""
                    ).replace("]",""
                    ).replace("'",'')
            
            validations = [
                ['"CC,CE,TI,PPT,PEP,PA"','B2'],
                ["=CIUDADES","J2"],
                ['"Propia,Familiar,Arrendada"','M2'],
                ['"M,F,O"','N2'],
                ['"A+,A-,B+,B-,AB+,AB-,O+,O-,ND"','O2'],
                ['"SOLTERO (A),CASADO (A),VIUDO (A),DIVORCIADO (A), UNION LIBRE"','P2'],
                ['=EPS',"R2"],
                [f'"{PENSION}"',"S2"],
                [f'"{CESANTIAS}"',"T2"],
                [f'"{ARL}"',"U2"],
                ['"0.522,1.044,4.35,NO APLICA,OTRO"','V2'],
                ["=CCF","W2"],
                ["=NIVEL_EDUCATIVO",'AA2'],
                ['"A,T,V"',"AB2"],
                ["=MODALIDAD_INGRESO","AC2"],
                ['"Indefinido,Obra o Labor,Aprendizaje"','AD2'],
                [f'"{EMPLEADORES}"',"AE2"],
                [f'"{TEMPORALES}"',"AF2"],
                ["=GERENCIAS","AJ2"],
                ["=INDIRECT(AJ2)","AK2"],
                ["=INDIRECT(AK2)","AL2"],
                ["=UBICACIONES","AM2"],
                ["=CECO","AN2"],
                ['"0.3,1.1,1.28,2.3,2.56,3,3.2,3.48,3.5,3.84,4.5,5.12"','AR2'],
                ["=AUXILIOS","AS2"],
                ["=CIUDADES","AV2"],
            ]
            
            for i in validations:
                dv = DataValidation(
                    type = 'list',
                    showDropDown=False,
                    formula1=i[0]
                )
                ws_data.add_data_validation(dv)
                dv.add(ws_data[i[1]])
            
            date_validations = ['Q2','AG2','AH2','AI2']
            for i in date_validations:
                dv = DataValidation(
                    type = 'date',
                    showDropDown=False,
                )
                ws_data.add_data_validation(dv)
                dv.add(ws_data[i])
                
            number_validations = ['AO2','AP2','AQ2','AT2']
            for i in number_validations:
                dv = DataValidation(
                    type = 'whole',
                    showDropDown=False,
                )
                ws_data.add_data_validation(dv)
                dv.add(ws_data[i])
            
            
            widths = [16,15,12,25,25,25,25,30,41,25,20,25,23,
                      10,10,18,23,23,23,23,18,21,23,11,11,11,24,25,
                      25,25,23,25,19,22,19,30,30,30,25,25,22,20,22,16,
                      22,22,25]
            
            for w, col in zip(widths, range(1,len(widths)+1)):
                col_letter = ws_data.cell(row=1, column=col).column_letter  # Obtener letra de la columna
                ws_data.column_dimensions[col_letter].width = w
            
            bold_font = Font(bold=True)  # Negrita
            fill_color = PatternFill(start_color="9ACBE6", end_color="9ACBE6", fill_type="solid")  # Amarillo
            border_style = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            center_alignment = Alignment(horizontal="center", vertical="center")  # Centrar texto
            # Aplicar estilos a la primera fila
            for cell in ws_data[1]:  # Fila 1
                cell.font = bold_font  # Negrita
                cell.fill = fill_color  # Fondo amarillo
                cell.border = border_style  # Borde delgado
                cell.alignment = center_alignment
                
            tmp_dir = settings.MEDIA_ROOT / 'tmp'
            tmp_dir.mkdir(parents=True, exist_ok=True)
            filepath = tmp_dir / 'Importar_masivo_terceros.xlsx'

            wb.save(filepath)
            
            data = {
                'url':'/media/tmp/Importar_masivo_terceros.xlsx'
            }
            
            return JsonResponse(data)
        
    elif request.method == 'POST':
        import tempfile
        from scripts.map_hr_snapshot import Mapper, load_rows

        file = request.FILES.get('archivo_importar')
        if file is None:
            return JsonResponse({'class': 'error', 'msj': 'Debes seleccionar el archivo de personal.'}, status=400)

        try:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                for chunk in file.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
        except Exception:
            return JsonResponse({'class': 'error', 'msj': 'Error al leer el archivo subido.'}, status=400)

        try:
            source_rows = load_rows(tmp_path)
        except Exception as e:
            return JsonResponse({'class': 'error', 'msj': f'Error al procesar el Excel: {e}'}, status=400)
        finally:
            os.unlink(tmp_path)

        mapper = Mapper()
        records = [mapper.map_row(raw_row, row_number) for row_number, raw_row in source_rows]
        summary = process_import_records(records)

        if (summary['creados'] + summary['actualizados'] + summary['omitidos']) > 0:
            add_history(
                request.user,
                f'Importó personal desde Excel: {summary["creados"]} creados, {summary["actualizados"]} actualizados, {summary["omitidos"]} omitidos, {summary["conflictos"]} conflictos, {summary["errores"]} errores'
            )

        return JsonResponse(summary)


@login_required
@rol_permission('Gestion humana')
def initial_personal_import(request):
    jobs = importacion_personal_job.objects.filter(usuario=request.user).order_by('-fecha_creacion')[:10]
    context = {
        'jobs': jobs,
    }
    return render(request, 'initial_personal_import.html', context)


@login_required
@rol_permission('Gestion humana')
def initial_personal_import_start(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    file = request.FILES.get('archivo_importar')
    if file is None:
        return JsonResponse({'error': 'Debes seleccionar el archivo de personal.'}, status=400)

    tmp_dir = settings.MEDIA_ROOT / 'tmp' / 'import_jobs'
    tmp_dir.mkdir(parents=True, exist_ok=True)
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S_%f')
    stored_path = tmp_dir / f'personal_import_{request.user.pk}_{timestamp}.xlsx'

    try:
        with open(stored_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)
    except Exception:
        return JsonResponse({'error': 'Error al guardar el archivo temporal.'}, status=400)

    job = importacion_personal_job.objects.create(
        usuario=request.user,
        nombre_archivo=file.name,
        ruta_archivo=str(stored_path),
        status='PENDING',
    )

    worker = threading.Thread(
        target=run_initial_import_job,
        args=(job.pk, request.user.pk),
        daemon=True,
    )
    worker.start()

    return JsonResponse({
        'job_id': job.pk,
        'status': job.status,
    })


@login_required
@rol_permission('Gestion humana')
def initial_personal_import_status(request, job_id):
    job = importacion_personal_job.objects.get(pk=job_id, usuario=request.user)
    return JsonResponse({
        'job_id': job.pk,
        'status': job.status,
        'stop_requested': job.stop_requested,
        'nombre_archivo': job.nombre_archivo,
        'total_filas': job.total_filas,
        'filas_importables': job.filas_importables,
        'filas_procesadas': job.filas_procesadas,
        'creados': job.creados,
        'actualizados': job.actualizados,
        'omitidos': job.omitidos,
        'conflictos': job.conflictos,
        'errores': job.errores,
        'no_importables': job.no_importables,
        'ignoradas': job.ignoradas,
        'detalles_conflicto': job.detalles_conflicto[:10],
        'detalles_error': job.detalles_error[:10],
        'detalles_no_importables': job.detalles_no_importables[:10],
        'mensaje_error': job.mensaje_error,
        'fecha_inicio': job.fecha_inicio.isoformat() if job.fecha_inicio else None,
        'fecha_fin': job.fecha_fin.isoformat() if job.fecha_fin else None,
    })


@login_required
@rol_permission('Gestion humana')
def initial_personal_import_stop(request, job_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    updated = importacion_personal_job.objects.filter(
        pk=job_id,
        usuario=request.user,
        status__in=['PENDING', 'MAPPING', 'RUNNING', 'STOPPING'],
    ).update(
        stop_requested=True,
        status='STOPPING',
        fecha_actualizacion=timezone.now(),
    )

    if not updated:
        return JsonResponse({'error': 'La importación no se puede detener en su estado actual.'}, status=400)

    return JsonResponse({
        'job_id': job_id,
        'status': 'STOPPING',
    })


@login_required
@rol_permission('Gestion humana')
def initial_personal_import_download(request, job_id):
    job = importacion_personal_job.objects.get(pk=job_id, usuario=request.user)

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Resumen'
    ws_summary.append(['Campo', 'Valor'])
    summary_rows = [
        ('Job ID', job.pk),
        ('Archivo', job.nombre_archivo),
        ('Estado', job.status),
        ('Total filas', job.total_filas),
        ('Filas importables', job.filas_importables),
        ('Filas procesadas', job.filas_procesadas),
        ('Creados', job.creados),
        ('Actualizados', job.actualizados),
        ('Omitidos', job.omitidos),
        ('Conflictos', job.conflictos),
        ('Errores', job.errores),
        ('No importables', job.no_importables),
        ('Ignoradas', job.ignoradas),
        ('Fecha inicio', job.fecha_inicio.isoformat() if job.fecha_inicio else ''),
        ('Fecha fin', job.fecha_fin.isoformat() if job.fecha_fin else ''),
        ('Mensaje error', job.mensaje_error or ''),
    ]
    for row in summary_rows:
        ws_summary.append(row)

    def fill_sheet(ws, headers, rows):
        ws.append(headers)
        for row in rows:
            ws.append(row)

    ws_conflicts = wb.create_sheet('Conflictos')
    fill_sheet(
        ws_conflicts,
        ['Fila', 'ID', 'Estado', 'Motivo', 'Contrato ID', 'Fecha inicio existente'],
        [
            [
                item.get('fila'),
                item.get('id'),
                item.get('estado'),
                item.get('motivo'),
                item.get('contrato_id'),
                item.get('fecha_inicio_existente'),
            ]
            for item in job.detalles_conflicto
        ],
    )

    ws_errors = wb.create_sheet('Errores')
    fill_sheet(
        ws_errors,
        ['Fila', 'ID', 'Estado', 'Error'],
        [
            [
                item.get('fila'),
                item.get('id'),
                item.get('estado'),
                item.get('error'),
            ]
            for item in job.detalles_error
        ],
    )

    ws_no_import = wb.create_sheet('No importables')
    fill_sheet(
        ws_no_import,
        ['Fila', 'ID', 'Estado', 'Issues'],
        [
            [
                item.get('fila'),
                item.get('id'),
                item.get('estado'),
                ', '.join(item.get('issues', [])),
            ]
            for item in job.detalles_no_importables
        ],
    )

    for worksheet in wb.worksheets:
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                value = '' if cell.value is None else str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="importacion_personal_job_{job.pk}_detalle.xlsx"'
    )
    wb.save(response)
    return response
        
def safe_value(value):
     return "" if value is None else value


def comparable_import_value(value):
    if isinstance(value, datetime.datetime):
        value = value.date()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float):
        return int(value) if value.is_integer() else round(value, 6)
    return value


def import_field_changes(instance, field_values):
    changes = {}
    for field_name, new_value in field_values.items():
        current_value = getattr(instance, field_name)
        if comparable_import_value(current_value) != comparable_import_value(new_value):
            changes[field_name] = new_value
    return changes


def merge_import_personal_fields(instance, incoming_fields):
    merged = {}
    for field_name, incoming_value in incoming_fields.items():
        current_value = getattr(instance, field_name)

        if field_name == 'activo':
            merged[field_name] = bool(current_value) or bool(incoming_value)
            continue

        if incoming_value in (None, ''):
            merged[field_name] = current_value
            continue

        merged[field_name] = incoming_value

    return merged


def normalize_import_label(value):
    if value is None:
        return ''
    value = str(value)
    value = ''.join(ch for ch in unicodedata.normalize('NFKD', value) if not unicodedata.combining(ch))
    value = re.sub(r'\s+', ' ', value).strip()
    return value.upper()


AUXILIO_TYPE_ALIASES = {
    'AUIXLIO POR INVENTARIOS': 'AUXILIO POR INVENTARIOS',
    'COMISIONES COMERCIALES': 'COMISION COMERCIALES',
    'BONIFICACION RENDIMIENTO DE PLANTA': 'BONIFICACION POR RENDIMIENTO DE PLANTA',
    'RENDIMIENTO PLANTA': 'BONIFICACION POR RENDIMIENTO DE PLANTA',
}


def canonical_auxilio_description(description):
    normalized = normalize_import_label(description)
    if not normalized:
        return None
    return AUXILIO_TYPE_ALIASES.get(normalized, normalized)


def build_import_auxilio_specs(record):
    specs = []
    for item in record.get('auxilios', []):
        description = canonical_auxilio_description(item.get('descripcion'))
        value = item.get('valor')
        if not description or value in (None, 0):
            continue
        specs.append({
            'descripcion': description,
            'valor': float(value),
        })
    return specs


def sync_contract_auxilios(contract, auxilio_specs):
    desired = sorted(
        [(spec['descripcion'], comparable_import_value(spec['valor'])) for spec in auxilio_specs],
        key=lambda item: (item[0], item[1]),
    )
    existing_qs = auxilios_contrato.objects.filter(contrato=contract).select_related('tipo').order_by('id')
    existing = sorted(
        [(canonical_auxilio_description(item.tipo.descripcion), comparable_import_value(item.valor)) for item in existing_qs],
        key=lambda item: (item[0], item[1]),
    )
    if existing == desired:
        return False

    existing_qs.delete()
    for spec in auxilio_specs:
        tipo_obj, _ = tipos_auxilio.objects.get_or_create(descripcion=spec['descripcion'])
        auxilios_contrato.objects.create(
            contrato=contract,
            tipo=tipo_obj,
            valor=spec['valor'],
        )
    return True


def resolve_contract_import(existing_contracts, fecha_inicio, contract_fields):
    exact_matches = [
        contract for contract in existing_contracts
        if comparable_import_value(getattr(contract, 'fecha_inicio', None)) == comparable_import_value(fecha_inicio)
    ]
    if len(exact_matches) > 1:
        return {
            'action': 'conflict',
            'reason': 'multiple_contracts_same_start',
            'target': None,
        }
    if exact_matches:
        target = exact_matches[0]
        changes = import_field_changes(target, contract_fields)
        return {
            'action': 'update' if changes else 'noop',
            'reason': None,
            'target': target,
            'changes': changes,
        }

    active_contracts = [contract for contract in existing_contracts if getattr(contract, 'activo', False)]
    incoming_is_active = contract_fields.get('activo') is True

    if len(active_contracts) > 1:
        return {
            'action': 'conflict',
            'reason': 'multiple_active_contracts',
            'target': None,
        }

    # Historical inactive contracts can be created even if the worker already has
    # a different active contract. The real conflict is trying to create a
    # second active contract with a different start date.
    if incoming_is_active and active_contracts:
        return {
            'action': 'conflict',
            'reason': 'active_contract_exists_with_different_start',
            'target': active_contracts[0],
        }

    return {
        'action': 'create',
        'reason': None,
        'target': None,
    }


@login_required
def dependentlist(request):
    
    if request_is_ajax(request):
        if request.method == 'GET':
            todo = request.GET.get('todo')
            
            if todo == 'areasbyestructura':
                estructura = request.GET.get('estructura')
                if estructura == '': 
                    data = {
                        'data':[]
                    }
                else:
                    areas = area.objects.filter(estructura=estructura
                                            ).order_by('descripcion')                                            
                    data = {
                        'data':[{'name':i.descripcion,'value':i.pk} for i in areas]
                    }
                
            elif todo == 'cargosbyarea':
                areacargo = request.GET.get('area')
                
                if areacargo == '': 
                    data = {
                        'data':[]
                    }
                    
                else:
                    cargos = cargos_personal.objects.filter(area=areacargo, activo = True
                                            ).order_by('descripcion')
                                            
                    data = {
                        'data':[{'name':i.descripcion,'value':i.pk} for i in cargos]
                    }
            
            elif todo == 'empleadoresbytype':
                tipo = request.GET.get('tipo')
                if tipo == '':
                    data = {
                        'data':[]
                    }
                else:
                    temporal = True if tipo == 'T' else False
                    _empleadores = empleadores.objects.all().order_by('nombre')
                        
                    data = {
                        'data':[{'name':i.nombre,'value':i.pk} for i in _empleadores]
                    }  
                
            elif todo == 'citiesbystates':
                estado = request.GET.get('estado')
                if estado == '':
                    data = {
                        'data':[]
                    }
                else:     
                    ciudades = ciudad.objects.filter(departamento=estado).order_by('nombre')
                    
                    data = {
                        'data':[{'name':i.nombre,'value':i.pk} for i in ciudades]
                    }  
            
            return JsonResponse(data, safe=False)

def add_history(user,message):
    
    historial.objects.create(
        usuario = user,
        texto = message,
    )
    
    return True

def request_is_ajax(request):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest': 
        return True
    return False
