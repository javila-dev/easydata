import datetime
import importlib.util
import io
import os
import tempfile
from types import SimpleNamespace

from django.contrib.auth.models import User
from django.test import SimpleTestCase, TestCase
from openpyxl import Workbook, load_workbook

from human_resources.models import (
    arl,
    area,
    base_personal,
    cargo,
    ccf,
    ciudad,
    contratos_personal,
    departamento,
    dependencia,
    empleadores,
    estructura,
    motivos_retiro,
    sede,
)
from human_resources.views import (
    build_import_auxilio_specs,
    canonical_auxilio_description,
    import_field_changes,
    merge_import_personal_fields,
    resolve_contract_import,
)


class ImportHelpersTests(SimpleTestCase):
    def test_canonical_auxilio_description_normalizes_known_aliases(self):
        self.assertEqual(
            canonical_auxilio_description("AUIXLIO POR INVENTARIOS"),
            "AUXILIO POR INVENTARIOS",
        )
        self.assertEqual(
            canonical_auxilio_description("Rendimiento planta"),
            "BONIFICACION POR RENDIMIENTO DE PLANTA",
        )

    def test_build_import_auxilio_specs_includes_movilizacion_and_declared_auxilio(self):
        specs = build_import_auxilio_specs({
            "auxilios": [
                {"descripcion": "AUXILIO DE MOVILIZACION", "valor": 300000},
                {"descripcion": "AUIXLIO POR INVENTARIOS", "valor": 50000},
            ]
        })

        self.assertEqual(specs, [
            {"descripcion": "AUXILIO DE MOVILIZACION", "valor": 300000.0},
            {"descripcion": "AUXILIO POR INVENTARIOS", "valor": 50000.0},
        ])

    def test_import_field_changes_detects_only_modified_values(self):
        worker = SimpleNamespace(
            primer_nombre="ANA",
            primer_apellido="PEREZ",
            activo=True,
            fecha_nacimiento=datetime.date(1990, 1, 1),
        )

        changes = import_field_changes(worker, {
            "primer_nombre": "ANA",
            "primer_apellido": "GOMEZ",
            "activo": True,
            "fecha_nacimiento": datetime.date(1990, 1, 1),
        })

        self.assertEqual(changes, {"primer_apellido": "GOMEZ"})

    def test_merge_import_personal_fields_preserves_non_empty_existing_values(self):
        worker = SimpleNamespace(
            primer_nombre="ANA",
            talla_camisa="M",
            talla_pantalon="32",
            talla_calzado="39",
            email=None,
            activo=False,
        )

        merged = merge_import_personal_fields(worker, {
            "primer_nombre": "ANA MARIA",
            "talla_camisa": None,
            "talla_pantalon": "",
            "talla_calzado": None,
            "email": "ana@example.com",
            "activo": True,
        })

        self.assertEqual(merged["primer_nombre"], "ANA MARIA")
        self.assertEqual(merged["talla_camisa"], "M")
        self.assertEqual(merged["talla_pantalon"], "32")
        self.assertEqual(merged["talla_calzado"], "39")
        self.assertEqual(merged["email"], "ana@example.com")
        self.assertTrue(merged["activo"])

    def test_resolve_contract_import_returns_noop_for_identical_retry(self):
        contract = SimpleNamespace(
            pk=1,
            fecha_inicio=datetime.date(2026, 1, 1),
            tipo_ingreso="V",
            modalidad_ingreso="NUEVO INGRESO DIRECTO",
            tipo_contrato="Indefinido",
            tipo_posicion="HC",
            empleador_id=1,
            temporal_id=None,
            cargo_id=10,
            area_id=20,
            sede_id=30,
            cceco_id=None,
            ciudad_laboral_id=None,
            motivo_retiro_id=None,
            motivo_retiro_real=None,
            fecha_periodo_prueba=None,
            fecha_retiro=None,
            activo_desde=datetime.date(2026, 1, 1),
            activo_hasta=None,
            activo=True,
            salario_base=2000000,
        )

        decision = resolve_contract_import([contract], datetime.date(2026, 1, 1), {
            "tipo_ingreso": "V",
            "modalidad_ingreso": "NUEVO INGRESO DIRECTO",
            "tipo_contrato": "Indefinido",
            "tipo_posicion": "HC",
            "empleador_id": 1,
            "temporal_id": None,
            "cargo_id": 10,
            "area_id": 20,
            "sede_id": 30,
            "cceco_id": None,
            "ciudad_laboral_id": None,
            "motivo_retiro_id": None,
            "motivo_retiro_real": None,
            "fecha_periodo_prueba": None,
            "fecha_retiro": None,
            "activo_desde": datetime.date(2026, 1, 1),
            "activo_hasta": None,
            "activo": True,
            "salario_base": 2000000,
        })

        self.assertEqual(decision["action"], "noop")
        self.assertEqual(decision["target"].pk, 1)

    def test_resolve_contract_import_returns_update_for_same_start_date(self):
        contract = SimpleNamespace(
            pk=1,
            fecha_inicio=datetime.date(2026, 1, 1),
            tipo_ingreso="V",
            modalidad_ingreso="NUEVO INGRESO DIRECTO",
            tipo_contrato="Indefinido",
            tipo_posicion="HC",
            empleador_id=1,
            temporal_id=None,
            cargo_id=10,
            area_id=20,
            sede_id=30,
            cceco_id=None,
            ciudad_laboral_id=None,
            motivo_retiro_id=None,
            motivo_retiro_real=None,
            fecha_periodo_prueba=None,
            fecha_retiro=None,
            activo_desde=datetime.date(2026, 1, 1),
            activo_hasta=None,
            activo=True,
            salario_base=2000000,
        )

        decision = resolve_contract_import([contract], datetime.date(2026, 1, 1), {
            "tipo_ingreso": "V",
            "modalidad_ingreso": "NUEVO INGRESO DIRECTO",
            "tipo_contrato": "Indefinido",
            "tipo_posicion": "HC",
            "empleador_id": 1,
            "temporal_id": None,
            "cargo_id": 10,
            "area_id": 20,
            "sede_id": 30,
            "cceco_id": None,
            "ciudad_laboral_id": None,
            "motivo_retiro_id": None,
            "motivo_retiro_real": None,
            "fecha_periodo_prueba": None,
            "fecha_retiro": None,
            "activo_desde": datetime.date(2026, 1, 1),
            "activo_hasta": None,
            "activo": True,
            "salario_base": 2500000,
        })

        self.assertEqual(decision["action"], "update")
        self.assertEqual(decision["changes"], {"salario_base": 2500000})

    def test_resolve_contract_import_returns_conflict_for_different_active_contract(self):
        contract = SimpleNamespace(
            pk=1,
            fecha_inicio=datetime.date(2026, 1, 1),
            tipo_ingreso="V",
            modalidad_ingreso="NUEVO INGRESO DIRECTO",
            tipo_contrato="Indefinido",
            tipo_posicion="HC",
            empleador_id=1,
            temporal_id=None,
            cargo_id=10,
            area_id=20,
            sede_id=30,
            cceco_id=None,
            ciudad_laboral_id=None,
            motivo_retiro_id=None,
            motivo_retiro_real=None,
            fecha_periodo_prueba=None,
            fecha_retiro=None,
            activo_desde=datetime.date(2026, 1, 1),
            activo_hasta=None,
            activo=True,
            salario_base=2000000,
        )

        decision = resolve_contract_import([contract], datetime.date(2026, 2, 1), {
            "tipo_ingreso": "V",
            "modalidad_ingreso": "NUEVO INGRESO DIRECTO",
            "tipo_contrato": "Indefinido",
            "tipo_posicion": "HC",
            "empleador_id": 1,
            "temporal_id": None,
            "cargo_id": 99,
            "area_id": 20,
            "sede_id": 30,
            "cceco_id": None,
            "ciudad_laboral_id": None,
            "motivo_retiro_id": None,
            "motivo_retiro_real": None,
            "fecha_periodo_prueba": None,
            "fecha_retiro": None,
            "activo_desde": datetime.date(2026, 2, 1),
            "activo_hasta": None,
            "activo": True,
            "salario_base": 2000000,
        })

        self.assertEqual(decision["action"], "conflict")
        self.assertEqual(decision["reason"], "active_contract_exists_with_different_start")

    def test_resolve_contract_import_allows_historical_inactive_contract_when_active_exists(self):
        active_contract = SimpleNamespace(
            pk=1,
            fecha_inicio=datetime.date(2026, 3, 1),
            tipo_ingreso="V",
            modalidad_ingreso="NUEVO INGRESO DIRECTO",
            tipo_contrato="Indefinido",
            tipo_posicion="HC",
            empleador_id=1,
            temporal_id=None,
            cargo_id=10,
            area_id=20,
            sede_id=30,
            cceco_id=None,
            ciudad_laboral_id=None,
            motivo_retiro_id=None,
            motivo_retiro_real=None,
            fecha_periodo_prueba=None,
            fecha_retiro=None,
            activo_desde=datetime.date(2026, 3, 1),
            activo_hasta=None,
            activo=True,
            salario_base=2000000,
        )

        decision = resolve_contract_import([active_contract], datetime.date(2025, 1, 1), {
            "tipo_ingreso": "T",
            "modalidad_ingreso": "NUEVO INGRESO POR TEMPORAL",
            "tipo_contrato": "Obra o Labor",
            "tipo_posicion": "HC",
            "empleador_id": 2,
            "temporal_id": 3,
            "cargo_id": 11,
            "area_id": 21,
            "sede_id": 31,
            "cceco_id": None,
            "ciudad_laboral_id": None,
            "motivo_retiro_id": 4,
            "motivo_retiro_real": "FIN DE CONTRATO",
            "fecha_periodo_prueba": None,
            "fecha_retiro": datetime.date(2025, 6, 1),
            "activo_desde": datetime.date(2025, 1, 1),
            "activo_hasta": datetime.date(2025, 6, 1),
            "activo": False,
            "salario_base": 1500000,
        })

        self.assertEqual(decision["action"], "create")


class HumanResourcesRegressionTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(
            username="admin",
            email="admin@example.com",
            password="secret123",
        )
        self.client.force_login(self.user)

    def _create_worker(self, numero_identificacion=12345):
        depto = departamento.objects.create(nombre=f"Cundinamarca-{numero_identificacion}")
        city = ciudad.objects.create(nombre=f"Bogota-{numero_identificacion}", departamento=depto)
        arl_obj = arl.objects.create(nombre=f"ARL {numero_identificacion}")
        ccf_obj = ccf.objects.create(nombre=f"CCF {numero_identificacion}")

        return base_personal.objects.create(
            numero_identificacion=numero_identificacion,
            tipo_id="CC",
            primer_nombre="Ana",
            primer_apellido="Perez",
            arl=arl_obj,
            tipo_riesgo="0.522",
            ccf=ccf_obj,
            ciudad=city,
            departamento=depto,
            sexo="F",
        )

    def _create_active_contract(self, worker):
        dep = dependencia.objects.create(descripcion=f"Gerencia {worker.numero_identificacion}")
        struct = estructura.objects.create(
            descripcion=f"Estructura {worker.numero_identificacion}",
            dependecia=dep,
        )
        area_obj = area.objects.create(
            descripcion=f"Area {worker.numero_identificacion}",
            estructura=struct,
        )
        cargo_obj = cargo.objects.create(
            descripcion=f"Cargo {worker.numero_identificacion}",
            area=area_obj,
            cantidad_aprobada=1,
        )
        employer = empleadores.objects.create(nombre=f"Empresa {worker.numero_identificacion}")
        site = sede.objects.create(descripcion=f"Sede {worker.numero_identificacion}")

        return contratos_personal.objects.create(
            trabajador=worker,
            fecha_inicio="2026-01-01",
            tipo_ingreso="V",
            modalidad_ingreso="NUEVO INGRESO DIRECTO",
            tipo_contrato="Indefinido",
            empleador=employer,
            cargo=cargo_obj,
            area=area_obj,
            sede=site,
            salario_base=2000000,
            activo_desde="2026-01-01",
            activo=True,
        )

    def test_get_full_name_handles_missing_optional_names(self):
        worker = self._create_worker()

        self.assertEqual(worker.get_full_name(), "ANA PEREZ")

    def test_getworkers_supports_serverside_pagination(self):
        worker_one = self._create_worker(numero_identificacion=11111)
        self._create_active_contract(worker_one)

        worker_two = self._create_worker(numero_identificacion=22222)
        self._create_active_contract(worker_two)

        response = self.client.get(
            "/humanresources/workers/",
            {
                "todo": "getworkers",
                "active": "true",
                "retired": "false",
                "draw": 3,
                "start": 0,
                "length": 1,
                "order[0][column]": 0,
                "order[0][dir]": "asc",
                "search[value]": "",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["draw"], 3)
        self.assertEqual(payload["recordsTotal"], 2)
        self.assertEqual(payload["recordsFiltered"], 2)
        self.assertEqual(len(payload["data"]), 1)
        self.assertEqual(payload["data"][0]["numero_identificacion"], 11111)

    def test_getworkers_applies_searchbuilder_filters(self):
        worker_one = self._create_worker(numero_identificacion=11111)
        self._create_active_contract(worker_one)

        worker_two = self._create_worker(numero_identificacion=22222)
        self._create_active_contract(worker_two)

        response = self.client.get(
            "/humanresources/workers/",
            {
                "todo": "getworkers",
                "active": "true",
                "retired": "false",
                "draw": 4,
                "start": 0,
                "length": 25,
                "order[0][column]": 1,
                "order[0][dir]": "asc",
                "search[value]": "",
                "searchBuilder[logic]": "AND",
                "searchBuilder[criteria][0][condition]": "contains",
                "searchBuilder[criteria][0][origData]": "contrato_activo.cargo.descripcion",
                "searchBuilder[criteria][0][type]": "string",
                "searchBuilder[criteria][0][value][0]": "11111",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["recordsTotal"], 2)
        self.assertEqual(payload["recordsFiltered"], 1)
        self.assertEqual(len(payload["data"]), 1)
        self.assertEqual(payload["data"][0]["numero_identificacion"], 11111)

    def test_export_workers_excel_respects_datatable_filters(self):
        worker_one = self._create_worker(numero_identificacion=11111)
        self._create_active_contract(worker_one)

        worker_two = self._create_worker(numero_identificacion=22222)
        self._create_active_contract(worker_two)

        response = self.client.get(
            "/humanresources/workers/",
            {
                "todo": "export_workers_excel",
                "active": "true",
                "retired": "false",
                "order[0][column]": 0,
                "order[0][dir]": "asc",
                "search[value]": "11111",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
        worksheet = workbook["BD PERSONAL"]
        headers = [worksheet.cell(row=1, column=i).value for i in range(1, 67)]

        self.assertEqual(headers[0], "EMPLEADOR")
        self.assertEqual(headers[4], "ID")
        self.assertEqual(headers[5], "NOMBRE COMPLETO")
        self.assertEqual(headers[65], "MOTIVO REAL TERMINACION(CONFIDENCIAL)")
        self.assertEqual(worksheet.max_row, 2)
        self.assertEqual(worksheet.cell(row=2, column=5).value, 11111)

    def test_export_workers_excel_ignores_datatable_pagination(self):
        worker_one = self._create_worker(numero_identificacion=11111)
        self._create_active_contract(worker_one)

        worker_two = self._create_worker(numero_identificacion=22222)
        self._create_active_contract(worker_two)

        response = self.client.get(
            "/humanresources/workers/",
            {
                "todo": "export_workers_excel",
                "active": "true",
                "retired": "false",
                "order[0][column]": 0,
                "order[0][dir]": "asc",
                "search[value]": "",
                "start": 1,
                "length": 1,
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
        worksheet = workbook["BD PERSONAL"]

        self.assertEqual(worksheet.max_row, 3)
        self.assertEqual(worksheet.cell(row=2, column=5).value, 11111)
        self.assertEqual(worksheet.cell(row=3, column=5).value, 22222)

    def test_retire_worker_persists_real_reason(self):
        worker = self._create_worker(numero_identificacion=54321)
        contract = self._create_active_contract(worker)
        reason = motivos_retiro.objects.create(descripcion="Renuncia")

        response = self.client.post(
            "/humanresources/workers/",
            {
                "todo": "retire_worker",
                "numero_id": worker.numero_identificacion,
                "fecha_retiro": "2026-03-10",
                "motivo_retiro": reason.pk,
                "motivo_real": "Renuncia por oferta externa",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        worker.refresh_from_db()
        contract.refresh_from_db()
        self.assertFalse(worker.activo)
        self.assertFalse(contract.activo)
        self.assertEqual(contract.motivo_retiro_real, "Renuncia por oferta externa")

    def test_empleadoresbytype_returns_employer_list(self):
        empleadores.objects.create(nombre="Empresa Uno")
        empleadores.objects.create(nombre="Empresa Dos")

        response = self.client.get(
            "/humanresources/utils/dependentlist",
            {"todo": "empleadoresbytype", "tipo": "T"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(len(payload["data"]), 2)
        self.assertEqual(payload["data"][0]["name"], "Empresa Dos")
        self.assertEqual(payload["data"][1]["name"], "Empresa Uno")


class SyncHrCargosFromExcelTests(TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        spec = importlib.util.spec_from_file_location(
            "sync_hr_cargos_from_excel",
            "/code/scripts/sync_hr_cargos_from_excel.py",
        )
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        cls.sync_module = module

    def setUp(self):
        dep = dependencia.objects.create(descripcion="Dep Sync")
        self.struct_oper = estructura.objects.create(descripcion="OPERACIONES", dependecia=dep)
        self.area_sedes = area.objects.create(descripcion="O16 SEDES", estructura=self.struct_oper)

        self.used_cargo = cargo.objects.create(
            descripcion="OPERARIO DE SEDE",
            area=self.area_sedes,
            cantidad_aprobada=10,
            criticidad="MEDIA",
            tipo_posicion="HC",
        )
        self.excel_update_cargo = cargo.objects.create(
            descripcion="AUXILIAR DE SEDE",
            area=self.area_sedes,
            cantidad_aprobada=3,
            criticidad="MEDIA",
            tipo_posicion="HC",
        )
        self.delete_cargo = cargo.objects.create(
            descripcion="CARGO SOLO DB",
            area=self.area_sedes,
            cantidad_aprobada=1,
            criticidad="MEDIA",
            tipo_posicion="HC",
        )

        depto = departamento.objects.create(nombre="Depto Sync")
        city = ciudad.objects.create(nombre="Ciudad Sync", departamento=depto)
        arl_obj = arl.objects.create(nombre="ARL Sync")
        ccf_obj = ccf.objects.create(nombre="CCF Sync")
        worker = base_personal.objects.create(
            numero_identificacion=99901,
            tipo_id="CC",
            primer_nombre="ANA",
            primer_apellido="PEREZ",
            arl=arl_obj,
            tipo_riesgo="0.522",
            ccf=ccf_obj,
            ciudad=city,
            departamento=depto,
            sexo="F",
        )
        employer = empleadores.objects.create(nombre="Empresa Sync")
        site = sede.objects.create(descripcion="Sede Sync")
        contratos_personal.objects.create(
            trabajador=worker,
            fecha_inicio="2026-01-01",
            tipo_ingreso="V",
            modalidad_ingreso="NUEVO INGRESO DIRECTO",
            tipo_contrato="Indefinido",
            empleador=employer,
            cargo=self.used_cargo,
            area=self.area_sedes,
            sede=site,
            salario_base=2000000,
            activo_desde="2026-01-01",
            activo=True,
        )

    def _build_excel(self, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "BD PERSONAL"
        ws.append(["ESTRUCTURA", "AREA", "CARGO", "ESTADO", "NOMBRE COMPLETO", "TAP", "TIPO DE POSICION"])
        for row in rows:
            ws.append(row)

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()
        self.addCleanup(lambda: os.unlink(tmp.name))
        return tmp.name

    def test_build_snapshot_counts_active_and_vacante_rows(self):
        excel_path = self._build_excel([
            ["OPERATIVO", "O16 SEDES", "AUXILIAR DE SEDE", "ACTIVO", "ANA PEREZ", "MEDIO", "HC"],
            ["OPERATIVO", "O16 SEDES", "AUXILIAR DE SEDE", "VACANTE", "VACANTE", "MEDIO", "HC"],
            ["OPERATIVO", "O16 SEDES", "AUXILIAR DE SEDE", "INACTIVO VACANTE", "PERSONA", "ALTO", "HC PLUS"],
            ["OPERATIVO", "O16 SEDES", "AUXILIAR DE SEDE", "INACTIVO", "VACANTE AUXILIAR", "ALTO", "HC PLUS"],
            ["OPERATIVO", "O16 SEDES", "AUXILIAR DE SEDE", "INACTIVO", "PERSONA", "BAJO", "HC"],
            ["OPERATIVO", "O16 SEDES", "", "ACTIVO", "SIN CARGO", "", ""],
            ["SIN MAPEO", "O16 SEDES", "IGNORADO", "ACTIVO", "TEST", "", ""],
        ])

        snapshot = self.sync_module.build_snapshot(excel_path, "BD PERSONAL")
        key = ("OPERACIONES", "O16 SEDES", "AUXILIAR DE SEDE")

        self.assertEqual(snapshot["countable_by_key"][key], 4)
        self.assertEqual(snapshot["rows_by_structure"]["OPERACIONES"], 6)
        self.assertEqual(self.sync_module.most_common_value(snapshot["tap_by_key"][key]), "ALTO")
        self.assertEqual(self.sync_module.most_common_value(snapshot["tipo_posicion_by_key"][key]), "HC PLUS")

    def test_plan_sync_classifies_create_update_delete_and_keep(self):
        excel_path = self._build_excel([
            ["OPERATIVO", "O16 SEDES", "AUXILIAR DE SEDE", "ACTIVO", "ANA PEREZ", "MEDIO", "HC"],
            ["OPERATIVO", "O16 SEDES", "AUXILIAR DE SEDE", "VACANTE", "VACANTE", "MEDIO", "HC"],
            ["OPERATIVO", "O16 SEDES", "CARGO NUEVO", "INACTIVO", "VACANTE CARGO", "ALTO", "HC PLUS"],
        ])

        snapshot = self.sync_module.build_snapshot(excel_path, "BD PERSONAL")
        sync_plan = self.sync_module.plan_sync(snapshot)

        self.assertEqual(sync_plan["summary"]["to_create"], 1)
        self.assertEqual(sync_plan["summary"]["to_update"], 1)
        self.assertEqual(sync_plan["summary"]["to_delete"], 1)
        self.assertEqual(sync_plan["summary"]["used_outside_excel"], 1)
        self.assertEqual(sync_plan["to_create"][0]["cargo"], "CARGO NUEVO")
        self.assertEqual(sync_plan["to_create"][0]["cantidad_aprobada"], 1)
        self.assertEqual(sync_plan["to_create"][0]["tipo_posicion"], "HC PLUS")
        self.assertEqual(sync_plan["to_update"][0]["new_cantidad_aprobada"], 2)
        self.assertEqual(sync_plan["to_delete"][0]["cargo"], "CARGO SOLO DB")
        self.assertEqual(sync_plan["kept_used_outside_excel"][0]["cargo"], "OPERARIO DE SEDE")

    def test_apply_sync_updates_creates_and_deletes(self):
        excel_path = self._build_excel([
            ["OPERATIVO", "O16 SEDES", "AUXILIAR DE SEDE", "ACTIVO", "ANA PEREZ", "MEDIO", "HC"],
            ["OPERATIVO", "O16 SEDES", "AUXILIAR DE SEDE", "VACANTE", "VACANTE", "MEDIO", "HC"],
            ["OPERATIVO", "O16 SEDES", "CARGO NUEVO", "INACTIVO", "VACANTE CARGO", "ALTO", "HC PLUS"],
        ])
        snapshot = self.sync_module.build_snapshot(excel_path, "BD PERSONAL")
        sync_plan = self.sync_module.plan_sync(snapshot)

        with tempfile.TemporaryDirectory() as backup_dir:
            result = self.sync_module.apply_sync(sync_plan, backup_dir)

        self.excel_update_cargo.refresh_from_db()
        self.assertEqual(self.excel_update_cargo.cantidad_aprobada, 2)
        self.assertTrue(cargo.objects.filter(descripcion="CARGO NUEVO", area=self.area_sedes).exists())
        self.assertFalse(cargo.objects.filter(pk=self.delete_cargo.pk).exists())
        self.assertTrue(cargo.objects.filter(pk=self.used_cargo.pk).exists())
        self.assertEqual(result["updated_cargos"], 1)
        self.assertEqual(result["created_cargos"], 1)
        self.assertEqual(result["deleted_cargos"], 1)
        self.assertEqual(result["skipped_protected_deletes"], [])
